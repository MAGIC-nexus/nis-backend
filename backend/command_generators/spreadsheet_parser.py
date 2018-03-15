import openpyxl

import io
# import koala # An Excel files parser elaborating a graph allowing automatic evaluation
import regex as re  # Improvement over standard "re"

from backend import Issue
from backend.command_generators.json import create_command
from backend.command_generators.spreadsheet_command_parsers.analysis.indicators_spreadsheet_parse import parse_indicators_command
from backend.command_generators.spreadsheet_command_parsers.external_data.mapping_spreadsheet_parse import parse_mapping_command
from backend.command_generators.spreadsheet_command_parsers.external_data.etl_external_dataset_spreadsheet_parse import parse_etl_external_dataset_command
from backend.command_generators.spreadsheet_command_parsers.external_data.parameters_spreadsheet_parse import parse_parameters_command
from backend.command_generators.spreadsheet_command_parsers.specification.hierarchy_spreadsheet_parser import parse_hierarchy_command
from backend.command_generators.spreadsheet_command_parsers.specification.metadata_spreadsheet_parse import parse_metadata_command
from backend.command_generators.spreadsheet_command_parsers.specification.data_input_spreadsheet_parse import parse_data_input_command
from backend.command_generators.spreadsheet_command_parsers.specification.pedigree_matrix_spreadsheet_parse import \
    parse_pedigree_matrix_command
from backend.command_generators.spreadsheet_command_parsers.specification.references_spreadsheet_parser import \
    parse_references_command
from backend.command_generators.spreadsheet_command_parsers.specification.upscale_spreadsheet_parse import parse_upscale_command
from backend.command_generators.spreadsheet_command_parsers.specification.structure_spreadsheet_parser import parse_structure_command
from backend.command_generators.spreadsheet_utils import binary_mask_from_worksheet, obtain_rectangular_submatrices

# Most complex name
# [namespace::][type:]var(.var)*[(@|#)var] : [namespace] + [type] + var + [attribute or tag]
var_name = "([a-zA-Z][a-zA-Z0-9_-]*)"
cplex_var = "((" + var_name + "::)?" + var_name + "(\\." + var_name + ")*)"

# ############################### #
#  Main function                  #
# ############################### #


def commands_generator_from_ooxml_file(input, state):
    """
    It reads an Office Open XML input
    Yields a sequence of command_executors

Hoja comando
* Lex+Parse
* Producir JSON ~ AST
* Enumerar problemas sintácticos
* Producir hoja a partir de JSON??

Comando
* Analizar JSON
* Ejecutar
* Enumerar problemas semánticos

    :param input: A bytes input
    :param state: State used to check variables
    :return:
    """
    # Start the Excel reader
    xl_in = openpyxl.load_workbook(io.BytesIO(input), data_only=True)

    # Regular expressions for the different commands
    flags = re.IGNORECASE
    re_metadata = re.compile(r"^Metadata", flags=flags)
    re_processors = re.compile(r"(Processors|Proc)[ _]+" + var_name, flags=flags)
    re_hierarchy = re.compile(r"(Taxonomy|Tax|Composition|Comp)[ _]([cpf])[ ]" + var_name, flags=flags)
    re_upscale = re.compile(r"(Upscale|Up)[ _](" + var_name + "[ _]" + var_name + ")?", flags=flags)
    re_relations = re.compile(r"(Grammar|Relations|Rel)([ _]+" + var_name+")?", flags=flags)
    re_transform = re.compile(r"(Transform|Tx)[ _]" + var_name + "[ _]" + var_name, flags=flags)
    re_pedigree_template = re.compile(r"(Pedigree|Ped|NUSAP\.PM)[ _]+" + var_name, flags=flags)
    re_references = re.compile(r"(References|Ref)[ _]" + var_name, flags=flags)
    re_parameters = re.compile(r"(Parameters|Params)([ _]" + var_name + ")?", flags=flags)

    re_enum = re.compile(r"(Dataset|DS)[ _]" + var_name + "[ _](Enumerate|Enum)", flags=flags)
    re_meta = re.compile(r"(Metadata|MD)[ _]" + var_name, flags=flags)
    re_data = re.compile(r"(Dataset|DS)[ _]" + var_name, flags=flags)
    re_mapping = re.compile(r"^(Mapping|Map)([ _]" + cplex_var + "[ _]" + cplex_var + ")?", flags=flags)

    re_indicators = re.compile(r"(Indicators|KPI)([ _]" + var_name + ")?", flags=flags)

    # For each worksheet, get the command type, convert into primitive JSON
    for c, sh_name in enumerate(xl_in.sheetnames):
        issues = []
        total_issues = []  # type: List[Issue]
        sh_in = xl_in[sh_name]

        c_type = None
        c_label = None
        c_content = None

        # Extract worksheet matrices
        m = binary_mask_from_worksheet(sh_in, False)
        t = obtain_rectangular_submatrices(m)
        t = t[0]  # Take just the first element, a tuple (top, bottom, left, right) representing a rectangular region
        t = (t[0]+1, t[1]+1, t[2]+1, t[3]+1)  # Indices start at 1
        # v = worksheet_to_numpy_array(sh_in)

        name = sh_in.title
        # EXTERNAL DATASETS
        if re_enum.search(name):
            datasource = re_enum.search(name).group(1)
            c_type = "enumerate_datasource_datasets"
            c_content = {"data_source": datasource}

            # TODO "Enumerate dataset member of a Datasource" command
            # TODO The enumeration will be stored in the state
            # TODO A read command is needed to obtain it
            # if not sh_in.cell(row=t[0], column=t[2]).value:
            #     for s in sources:
            #         if s.lower() == datasource.lower():
            #             get_codes_all_statistical_datasets(s, dsm)
            #             break
        elif re_meta.search(name):
            dataset = re_meta.search(name).group(1)
            c_type = "dataset_metadata"
        elif re_data.search(name):
            dataset = re_data.search(name).group(2)
            c_type = "etl_dataset"
            if sh_in.cell(row=t[0], column=t[2]).value:
                t = (1, m.shape[0]+1, 1, m.shape[1]+1)
                # Parse to read parameters
                issues, c_label, c_content = parse_etl_external_dataset_command(sh_in, t, dataset, state)
            else:
                # Syntax error: it seems there are no parameters
                issue = {"sheet_number": c, "sheet_name": sh_name, "c_type": c_type, "type": 3,
                         "message": "It seems there are no parameters for the dataset import command at worksheet '" + sh_name + "'"}
                total_issues.append(issue)
        elif re_mapping.search(name):
            c_type = "mapping"
            g = re_mapping.search(name).groups()
            if g[2] and g[8]:
                origin = g[2]
                destination = g[8]
            elif not g[2] and not g[8]:
                origin = None
                destination = None
            else:
                issue = {"sheet_number": c, "sheet_name": sh_name, "c_type": c_type, "type": 3, "message": "Either origin or destination are not correctly specified in the sheet name '"+sh_name+"'"}
                total_issues.append(issue)
            issues, c_label, c_content = parse_mapping_command(sh_in, t, origin, destination)
        elif re_parameters.search(name):  # Parameters are not external datasets but are information coming from outside, somehow
            name = re_parameters.search(name).group(1)
            c_type = "parameters"
            issues, c_label, c_content = parse_parameters_command(sh_in, t)
        # SPECIFICATIONS
        elif re_metadata.search(name):
            c_type = "metadata"
            issues, c_label, c_content = parse_metadata_command(sh_in, t)
        elif re_processors.search(name):
            # Read ALL the content
            c_type = "data_input"
            c_label = re_processors.search(name).group(2)
            issues, c_label, c_content = parse_data_input_command(sh_in, t, c_label, None)
        elif re_upscale.search(name):
            c_type = "upscale"
            res = re_upscale.search(name)
            child = res.group(2)
            parent = res.group(3)
            c_label = "Upscale child '"+child+"' into parent '"+parent+"'"
            issues, c_label, c_content = parse_upscale_command(sh_in, t)
        elif re_hierarchy.search(name):
            # Read the hierarchy
            c_type = "hierarchy"
            res = re_hierarchy.search(name)
            h_type = res.group(2)
            c_label = res.group(3)
            issues, _, c_content = parse_hierarchy_command(sh_in, t, c_label, h_type)
        elif re_relations.search(name):
            # Read the content
            c_type = "structure"
            issues, c_label, c_content = parse_structure_command(sh_in, t)
        elif re_transform.search(name):
            c_type = "scale_conversion"
        elif re_pedigree_template.search(name):
            # Read the content
            c_type = "pedigree_matrix"
            res = re_pedigree_template.search(name)
            pm_name = res.group(2)
            issues, c_label, c_content = parse_pedigree_matrix_command(sh_in, t, pm_name)
        elif re_references.search(name):
            # Read the content
            c_type = "references"
            issues, c_label, c_content = parse_references_command(sh_in, t)
        # ANALYSIS COMMANDS
        elif re_indicators.search(name):  # Indicators
            name = re_parameters.search(name).group(1)
            c_type = "indicators"
            issues, c_label, c_content = parse_indicators_command(sh_in, t)

        # Append issues
        errors = 0
        if len(issues) > 0:
            for i in issues:
                if i[0] == 3:
                    errors += 1
                issue = {"sheet_number": c, "sheet_name": sh_name, "c_type": c_type, "type": i[0], "message": i[1]}
                total_issues.append(issue)
        if errors == 0:
            try:
                cmd, issues = create_command(c_type, c_label, c_content)
            except:
                cmd = None
                issues = [(3, "Could not create command of type '"+c_type+"'")]
            if issues:
                for i in issues:
                    issue = {"sheet_number": c, "sheet_name": sh_name, "c_type": c_type, "type": i[0], "message": i[1]}
                    total_issues.append(issue)
        else:
            cmd, _ = create_command(c_type, c_label, {})

        yield cmd, total_issues
    # yield from []  # Empty generator


# def get_codes_all_statistical_datasets(source, dataset_manager):
#     """
#     Obtain a list of datasets available from a source
#     If no source is specified, all the sources are queried
#     For each dataset, the source, the name, the periods available, an example command and a description are obtained
#
#     :param source:
#     :param dataset_manager: It is a DataSourceManager
#     :return: A Dataframe with the list of datasets
#     """
#     lst2 = []
#     # TODO Probably "get_datasets" will not work as expected. It returns a tuple (Source, list of datasets)
#     for r, k in enumerate(dataset_manager.get_datasets(source)):
#         if len(k) == 4:
#             src = k[3]
#         else:
#             src = ""
#         lst2.append((k[0], k[1], k[2], src))
#     return pd.DataFrame(data=lst2, columns=["Dataset ID", "Description", "URN", "Data Source"])

