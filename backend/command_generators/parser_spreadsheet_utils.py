import numpy as np
import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill

global_fill = PatternFill("none")

# #################################### #
#  Worksheet related helper functions  #
# #################################### #


def worksheet_to_numpy_array(sh_in):
    """
    Obtain a replica of the worksheet into a Numpy NDArray, with combined cells (combined cells are repeated)

    :param sh_in:
    :return: The numpy array with the values of the worksheet
    """
    m = np.zeros((sh_in.max_row, sh_in.max_column)).astype(object)
    for r in range(sh_in.max_row):
        for c in range(sh_in.max_column):
            v = sh_in.cell(row=r + 1, column=c + 1).value
            if v:
                m[r, c] = v
            else:
                m[r, c] = 0.0

    # Merged cells
    for ra in sh_in.merged_cell_ranges:
        t = openpyxl.utils.range_boundaries(ra)  # min col, min row, max col, max row (max's included)
        mc = (t[1]-1, t[3]-1, t[0]-1, t[2]-1)  # Rearrange and subtract one
        v = m[mc[0], mc[2]]
        m[mc[0]:mc[1]+1, mc[2]:mc[3]+1] = v

    return m


def binary_mask_from_worksheet(sh_in, only_numbers=True):
    """
    Sweep the worksheet, considering merged cells, elaborate a mask for those cells which
    are not empty or contain a number

    :param sh_in:
    :param only_numbers:
    :return:
    """
    m = np.zeros((sh_in.max_row, sh_in.max_column), dtype=bool)
    for r in range(sh_in.max_row):
        for c in range(sh_in.max_column):
            v = sh_in.cell(row=r + 1, column=c + 1).value
            if v:
                if only_numbers:
                    if isinstance(v, int) or isinstance(v, float):
                        m[r, c] = 1
                else:
                    m[r, c] = 1

    # Merged cells
    for ra in sh_in.merged_cell_ranges:
        t = openpyxl.utils.range_boundaries(ra)  # min col, min row, max col, max row (max's included)
        mc = (t[1]-1, t[3]-1, t[0]-1, t[2]-1)  # Rearrange and subtract one
        v = m[mc[0], mc[2]]
        m[mc[0]:mc[1]+1, mc[2]:mc[3]+1] = v

    return m


def obtain_rectangular_submatrices(mask, region=None, only_remove_empty_bottom=False):
    """
    Obtain rectangular submatrices of mask
    IMPORTANT: currently it only obtains ONE region

    :param mask: The original matrix, numpy.NDArray, containing only 0/1 (1 is "some content")
    :param region: A tuple (top, bottom, left, right) with indices to search. bottom and right are not included
    :return: The list of rectangular regions as tuples (top, bottom, left, right)
    """

    def nonzero_sequences(a):
        # Create an array that is 1 where a is non-zero, and pad each end with an extra 0.
        isnonzero = np.concatenate(([0], a != 0, [0]))
        absdiff = np.abs(np.diff(isnonzero))
        # Runs start and end where absdiff is 1.
        ranges = np.where(absdiff == 1)[0].reshape(-1, 2)
        return ranges

    lst = []
    if not region:
        region = (0, mask.shape[0], 0, mask.shape[1])  # All the mask
    submask = mask[region[0]:region[1], region[2]:region[3]]
    offset_col, offset_row = (region[2], region[0])
    # Accumulation of elements by row (resulting in a column vector)
    row_sum = np.sum(submask, axis=1)
    # Accumulation of elements by column (resulting in a row vector)
    col_sum = np.sum(submask, axis=0)

    # Ranges
    rs = nonzero_sequences(row_sum.flatten())
    cs = nonzero_sequences(col_sum.flatten())
    if only_remove_empty_bottom:
        if len(rs) > 0:
            lst.append((rs[0][0], rs[-1][1], cs[0][0], cs[-1][1]))
    else:
        # Take the first rectangle
        if len(rs) > 0:
            lst.append((rs[0][0], rs[0][1], cs[0][0], cs[0][1]))

    return lst


def reset_cell_format(sh_writable, r, c):
    """
    When writing, reset cell's format

    :param sh_writable: Output worksheet
    :param r: Row number
    :param c: Col number
    :return:
    """
    cell = sh_writable.cell(row=r, column=c)
    cell.fill = global_fill
    cell.comment = None


def reset_cells_format(sh_writable):
    """
    When writing, reset all worksheet cells format

    :param sh_writable: Output worksheet
    :return:
    """
    for r in range(sh_writable.max_row):
        for c in range(sh_writable.max_column):
            reset_cell_format(sh_writable, r + 1, c + 1)


def cell_content_to_str(v):
    """
    Convert the value of a cell to string

    :param v: Value of a cell
    :return:
    """
    if v:
        if isinstance(v, float) or isinstance(v, int):
            return str(int(v))
        else:
            return str(v).strip()
    else:
        return None


def show_message(sh, r, c, message, type="error", accumulate=True):
    """
    It serves to show a cell in a worksheet
    It shows some type of error (warning or error) with a message in a comment
    The name of the sheet is changed with a prefix indicating there is at least an issue to be solved

    :param sh:
    :param r:
    :param c:
    :param message:
    :param type:
    :return:
    """
    cell = sh.cell(row=r, column=c)
    fill = cell.fill
    if type == "error":
        fill = PatternFill("solid", fgColor="CC0000")
    elif type == "warning":
        fill = PatternFill("solid", fgColor="FFFF33")
    elif type == "info":
        fill = PatternFill("solid", fgColor="87CEEB")
    cell.fill = fill
    if accumulate:
        comment = cell.comment
        if comment:
            comment.text += "\n" + message
        else:
            comment = Comment(message, "NIS")
    else:
        comment = Comment(message, "NIS")
    cell.comment = comment
    # if type == "error":
    #     sh.title = "!" + sh.title


def rewrite_xlsx_file(xl):
    """
    Regenerates the worksheets of the input file. The aim is to calculate correctly the "dimension@ref" attribute of
    each of the Worksheets, in order to have it correctly processed by the Kendo UI Spreadsheet

    :param xl: A Workbook object, constructed with OpenPyXL
    :return: Nothing, the "xl" object is modified inplace
    """
    sn = xl.sheetnames
    for c, sh_name in enumerate(sn):
        source = xl.get_sheet_by_name(sh_name)
        tmp = xl.copy_worksheet(source)
        xl.remove_sheet(source)
        tmp.title = sh_name
