import openpyxl
import io
# import koala # An Excel files parser elaborating a graph allowing automatic evaluation
import numpy as np
import pandas as pd
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill
import regex as re  # Improvement over standard "re"

from backend import Issue
from backend.command_generators.json import create_command
from backend.command_generators.spreadsheet_command_parsers.external_data.mapping_spreadsheet_parse import parse_mapping_command
from backend.command_generators.spreadsheet_command_parsers.external_data.etl_external_dataset_spreadsheet_parse import parse_etl_external_dataset_command
from backend.command_generators.spreadsheet_command_parsers.specification.metadata_spreadsheet_parse import parse_metadata_command
from backend.command_generators.spreadsheet_command_parsers.specification.data_input_spreadsheet_parse import parse_data_input_command
from backend.command_generators.spreadsheet_command_parsers.specification.upscale_spreadsheet_parse import parse_upscale_command

global_fill = PatternFill("none")

# Most complex name
# [namespace::][type:]var(.var)*[(@|#)var] : [namespace] + [type] + var + [attribute or tag]
var_name = "([a-zA-Z][a-zA-Z0-9_-]*)"
cplex_var = "((" + var_name + "::)?" + var_name + "(\\." + var_name + ")*)"

# ############################### #
#  Main function                  #
# ############################### #


def commands_generator_from_ooxml_file(input, state):
    """
    It reads an Office Open XML input
    Yields a sequence of command_executors

Hoja comando
* Lex+Parse
* Producir JSON ~ AST
* Enumerar problemas sintácticos
* Producir hoja a partir de JSON??

Comando
* Analizar JSON
* Ejecutar
* Enumerar problemas semánticos

    :param input: A bytes input
    :param state: State used to check variables
    :return:
    """
    # Start the Excel reader
    xl_in = openpyxl.load_workbook(io.BytesIO(input), data_only=True)

    # Regular expressions for the different commands
    flags = re.IGNORECASE
    re_metadata = re.compile(r"^Metadata", flags=flags)
    re_processors = re.compile(r"(Processors|Proc)[ _]" + var_name, flags=flags)
    re_hierarchy = re.compile(r"(Taxonomy|Tax|Composition|Comp)[ _]([cpf])[ ]" + var_name, flags=flags)
    re_upscale = re.compile(r"(Upscale|Up)[ _]" + var_name + "[ _]" + var_name, flags=flags)
    re_relations = re.compile(r"(Relations|Rel)" + var_name, flags=flags)
    re_transform = re.compile(r"(Transform|Tx)[ _]" + var_name + "[ _]" + var_name, flags=flags)
    re_pedigree_template = re.compile(r"(Pedigree|Ped)[ _]" + var_name, flags=flags)
    re_references = re.compile(r"(References|Ref)[ _]" + var_name, flags=flags)

    re_enum = re.compile(r"(Dataset|DS)[ _]" + var_name + "[ _](Enumerate|Enum)", flags=flags)
    re_meta = re.compile(r"(Metadata|MD)[ _]" + var_name, flags=flags)
    re_data = re.compile(r"(Dataset|DS)[ _]" + var_name, flags=flags)
    re_mapping = re.compile(r"^(Mapping|Map)([ _]" + cplex_var + "[ _]" + cplex_var + ")?", flags=flags)

    # For each worksheet, get the command type, convert into primitive JSON
    for c, sh_name in enumerate(xl_in.get_sheet_names()):
        issues = []
        total_issues = []  # type: List[Issue]
        sh_in = xl_in.get_sheet_by_name(sh_name)

        c_type = None
        c_label = None
        c_content = None

        # Extract worksheet matrices
        m = binary_mask_from_worksheet(sh_in, False)
        t = obtain_rectangular_submatrices(m)
        t = t[0]  # Take just the first element, a tuple (top, bottom, left, right) representing a rectangular region
        t = (t[0]+1, t[1]+1, t[2]+1, t[3]+1)  # Indices start at 1
        # v = worksheet_to_numpy_array(sh_in)

        name = sh_in.title
        # EXTERNAL DATASETS
        if re_enum.search(name):
            datasource = re_enum.search(name).group(1)
            c_type = "enumerate_datasource_datasets"
            c_content = {"data_source": datasource}

            # TODO "Enumerate dataset member of a Datasource" command
            # TODO The enumeration will be stored in the state
            # TODO A read command is needed to obtain it
            # if not sh_in.cell(row=t[0], column=t[2]).value:
            #     for s in sources:
            #         if s.lower() == datasource.lower():
            #             get_codes_all_statistical_datasets(s, dsm)
            #             break
        elif re_meta.search(name):
            # TODO
            dataset = re_meta.search(name).group(1)
            c_type = "dataset_metadata"
            c_content = None
        elif re_data.search(name):
            dataset = re_data.search(name).group(2)
            c_type = "etl_dataset"
            if sh_in.cell(row=t[0], column=t[2]).value:
                t = (1, m.shape[0]+1, 1, m.shape[1]+1)
                # Parse to read parameters
                issues, c_label, c_content = parse_etl_external_dataset_command(sh_in, t, dataset, state)
            else:
                # Syntax error: it seems there are no parameters
                issue = {"sheet_number": c, "sheet_name": sh_name, "c_type": c_type, "type": 3,
                         "message": "It seems there are no parameters for the dataset import command at worksheet '" + sh_name + "'"}
                total_issues.append(issue)
        elif re_mapping.search(name):
            c_type = "mapping"
            g = re_mapping.search(name).groups()
            if g[2] and g[8]:
                origin = g[2]
                destination = g[8]
            elif not g[2] and not g[8]:
                origin = None
                destination = None
            else:
                issue = {"sheet_number": c, "sheet_name": sh_name, "c_type": c_type, "type": 3, "message": "Either origin or destination are not correctly specified in the sheet name '"+sh_name+"'"}
                total_issues.append(issue)
            issues, c_label, c_content = parse_mapping_command(sh_in, t, origin, destination)
        # SPECIFICATIONS
        elif re_metadata.search(name):
            c_type = "metadata"
            issues, c_label, c_content = parse_metadata_command(sh_in, t)
        elif re_processors.search(name):
            # Read ALL the content
            c_type = "data_input"
            c_label = re_processors.search(name).group(2)
            issues, c_label, c_content = parse_data_input_command(sh_in, t, c_label, None)
        elif re_upscale.search(name):
            # TODO Upscale
            c_type = "upscale"
            res = re_upscale.search(name)
            child = res.group(2)
            parent = res.group(3)
            c_label = "Upscale child '"+child+"' into parent '"+parent+"'"
            issues, c_label, c_content = parse_upscale_command(sh_in, t, c_label, None)
        elif re_hierarchy.search(name):
            # TODO Read the hierarchy
            c_type = "hierarchy"
            res = re_hierarchy.search(name)
            c_label = res.group(1) + ":" + res.group(2) + ":" + res.group(3)
        elif re_relations.search(name):
            # TODO Read the content
            c_type = "structure"
        elif re_transform.search(name):
            c_type = "scale_conversion"
        elif re_pedigree_template.search(name):
            # TODO Read the content
            c_type = "pedigree_matrix"
        elif re_references.search(name):
            # TODO Read the content
            c_type = "referenceable_data"

        # Append issues
        errors = 0
        if len(issues) > 0:
            for i in issues:
                if i[0] == 3:
                    errors += 1
                issue = {"sheet_number": c, "sheet_name": sh_name, "c_type": c_type, "type": i[0], "message": i[1]}
                total_issues.append(issue)
        if errors == 0:
            cmd, issues = create_command(c_type, c_label, c_content)
            if issues:
                for i in issues:
                    issue = {"sheet_number": c, "sheet_name": sh_name, "c_type": c_type, "type": i[0], "message": i[1]}
                    total_issues.append(issue)
        else:
            cmd, _ = create_command(c_type, c_label, {})

        yield cmd, total_issues
    # yield from []  # Empty generator

# ################################## #
# Worksheet related helper functions #
# ################################## #


def reset_cell_format(sh_writable, r, c):
    """
    When writing, reset cell's format

    :param sh_writable: Output worksheet
    :param r: Row number
    :param c: Col number
    :return:
    """
    cell = sh_writable.cell(row=r, column=c)
    cell.fill = global_fill
    cell.comment = None


def reset_cells_format(sh_writable):
    """
    When writing, reset all worksheet cells format

    :param sh_writable: Output worksheet
    :return:
    """
    for r in range(sh_writable.max_row):
        for c in range(sh_writable.max_column):
            reset_cell_format(sh_writable, r + 1, c + 1)


def cell_content_to_str(v):
    """
    Convert the value of a cell to string

    :param v: Value of a cell
    :return:
    """
    if v:
        if isinstance(v, float) or isinstance(v, int):
            return str(int(v))
        else:
            return str(v).strip()
    else:
        return None


def show_message(sh, r, c, message, type="error", accumulate=True):
    """
    It serves to show a cell in a worksheet
    It shows some type of error (warning or error) with a message in a comment
    The name of the sheet is changed with a prefix indicating there is at least an issue to be solved

    :param sh:
    :param r:
    :param c:
    :param message:
    :param type:
    :return:
    """
    cell = sh.cell(row=r, column=c)
    fill = cell.fill
    if type == "error":
        fill = PatternFill("solid", fgColor="CC0000")
    elif type == "warning":
        fill = PatternFill("solid", fgColor="FFFF33")
    elif type == "info":
        fill = PatternFill("solid", fgColor="87CEEB")
    cell.fill = fill
    if accumulate:
        comment = cell.comment
        if comment:
            comment.text += "\n" + message
        else:
            comment = Comment(message, "NIS")
    else:
        comment = Comment(message, "NIS")
    cell.comment = comment
    # if type == "error":
    #     sh.title = "!" + sh.title


def obtain_rectangular_submatrices(mask, region=None):
    """
    Obtain rectangular submatrices of mask
    IMPORTANT: currently it only obtains ONE region

    :param mask: The original matrix, numpy.NDArray, containing only 0/1 (1 is "some content")
    :param region: A tuple (top, bottom, left, right) with indices to search. bottom and right are not included
    :return: The list of rectangular regions as tuples (top, bottom, left, right)
    """

    def nonzero_sequences(a):
        # Create an array that is 1 where a is non-zero, and pad each end with an extra 0.
        isnonzero = np.concatenate(([0], a != 0, [0]))
        absdiff = np.abs(np.diff(isnonzero))
        # Runs start and end where absdiff is 1.
        ranges = np.where(absdiff == 1)[0].reshape(-1, 2)
        return ranges

    lst = []
    if not region:
        region = (0, mask.shape[0], 0, mask.shape[1])  # All the mask
    submask = mask[region[0]:region[1], region[2]:region[3]]
    offset_col, offset_row = (region[2], region[0])
    # Accumulation of elements by row (resulting in a column vector)
    row_sum = np.sum(submask, axis=1)
    # Accumulation of elements by column (resulting in a row vector)
    col_sum = np.sum(submask, axis=0)

    # Ranges
    rs = nonzero_sequences(row_sum.flatten())
    cs = nonzero_sequences(col_sum.flatten())
    lst.append((rs[0][0], rs[0][1], cs[0][0], cs[0][1]))

    return lst


def worksheet_to_numpy_array(sh_in):
    """
    Obtain a replica of the worksheet into a Numpy NDArray, with combined cells (combined cells are repeated)

    :param sh_in:
    :return: The numpy array with the values of the worksheet
    """
    m = np.zeros((sh_in.max_row, sh_in.max_column)).astype(object)
    for r in range(sh_in.max_row):
        for c in range(sh_in.max_column):
            v = sh_in.cell(row=r + 1, column=c + 1).value
            if v:
                m[r, c] = v
            else:
                m[r, c] = 0.0

    # Merged cells
    for ra in sh_in.merged_cell_ranges:
        t = openpyxl.utils.range_boundaries(ra)  # min col, min row, max col, max row (max's included)
        mc = (t[1]-1, t[3]-1, t[0]-1, t[2]-1)  # Rearrange and subtract one
        v = m[mc[0], mc[2]]
        m[mc[0]:mc[1]+1, mc[2]:mc[3]+1] = v

    return m


def binary_mask_from_worksheet(sh_in, only_numbers=True):
    """
    Sweep the worksheet, considering merged cells, elaborate a mask for those cells which
    are not empty or contain a number

    :param sh_in:
    :param only_numbers:
    :return:
    """
    m = np.zeros((sh_in.max_row, sh_in.max_column), dtype=bool)
    for r in range(sh_in.max_row):
        for c in range(sh_in.max_column):
            v = sh_in.cell(row=r + 1, column=c + 1).value
            if v:
                if only_numbers:
                    if isinstance(v, int) or isinstance(v, float):
                        m[r, c] = 1
                else:
                    m[r, c] = 1

    # Merged cells
    for ra in sh_in.merged_cell_ranges:
        t = openpyxl.utils.range_boundaries(ra)  # min col, min row, max col, max row (max's included)
        mc = (t[1]-1, t[3]-1, t[0]-1, t[2]-1)  # Rearrange and subtract one
        v = m[mc[0], mc[2]]
        m[mc[0]:mc[1]+1, mc[2]:mc[3]+1] = v

    return m


def get_codes_all_statistical_datasets(source, dataset_manager):
    """
    Obtain a list of datasets available from a source
    If no source is specified, all the sources are queried
    For each dataset, the source, the name, the periods available, an example command and a description are obtained

    :param source:
    :param dataset_manager: It is a DataSourceManager
    :return: A Dataframe with the list of datasets
    """
    lst2 = []
    for r, k in enumerate(dataset_manager.get_datasets(source)):
        if len(k) == 4:
            src = k[3]
        else:
            src = ""
        lst2.append((k[0], k[1], k[2], src))
    return pd.DataFrame(data=lst2, columns=["Dataset ID", "Description", "URN", "Data Source"])

