from typing import List, Tuple, Optional, Dict

from openpyxl.worksheet.worksheet import Worksheet

from backend import CommandField, IssuesLabelContentTripleType, AreaTupleType
from backend.command_generators import Issue, parser_field_parsers, IssueLocation, IType


def check_columns(sh, name: str, area: Tuple, cols: List[CommandField], command_name: str, ignore_not_found=False):
    """
    When parsing of a command starts, check columns
    Try to match each column with declared column fields. If a column is not declared, raise an error (or ignore it)
    If mandatory columns are not found, raise an error

    :param sh: The worksheet being analyzed
    :param name: The name of the worksheet
    :param area: Area inside the worksheet that will be scanned
    :param cols: List of CommandField
    :param command_name: A string with the name of the command
    :param ignore_not_found: True if a column not matching declared ones has to be ignored, False if an error has to be raised in this case
    :return: The map column name to column index (or indices for multiply declared columns); The issues found
    """

    issues: List[Issue] = []

    # Set of mandatory columns
    mandatory_not_found = set([c.name for c in cols if c.mandatory])

    # Check columns
    col_map = {}  # From CommandField to a list of column index
    for c in range(area[2], area[3]):  # For each column of row 0 (Header Row)
        ##val = sh.get((area[0], c), None)
        val = sh.cell(row=area[0], column=c).value
        if not val:
            continue
        col_name = val.strip()
        for col in cols:  # Find matching CommandField from the attribute "regex_allowed_names"
            if col.regex_allowed_names.match(col_name):
                # Found matching CommandField "col". Process
                if "@" in col_name:  # In case of use of "@", remove prefix
                    col_name = col_name[col_name.index("@")+1:]
                # Column Name to Column Index
                if not col.many_appearances:  # Column appears once
                    if col in col_map:
                        issues.append(Issue(itype=IType.ERROR,
                                            description="The column '"+col.name+"' should not appear more than one time",
                                            location=IssueLocation(sheet_name=name, row=1, column=c)))
                    col_map[col] = [(col_name, c)]
                else:  # Column appears one or more times
                    if col not in col_map:
                        col_map[col] = []
                    col_map[col].append((col_name, c))
                # Mandatory found (good)
                if col.name in mandatory_not_found:
                    mandatory_not_found.discard(col.name)
                break
        else:  # No match for the column "col_name"
            if not ignore_not_found:
                issues.append(Issue(itype=IType.ERROR,
                                    description="The column name '" + col_name + "' does not match any of the allowed column names for the command '" + command_name + "'",
                                    location=IssueLocation(sheet_name=name, row=1, column=c)))

    if len(mandatory_not_found) > 0:
        issues.append(Issue(itype=IType.ERROR,
                            description="Mandatory columns: " + ", ".join(
                                mandatory_not_found) + " have not been specified",
                            location=IssueLocation(sheet_name=name, row=1, column=None)))

    return col_map, issues


def read_worksheet(sh: Worksheet) -> Dict:
    rows = sh.rows
    data = {}
    for r, row in enumerate(rows):
        for c, cell in enumerate(row):
            if cell.data_type == 's':
                data[(r+1, c+1)] = cell.value.strip()
            else:
                data[(r+1, c+1)] = cell.value
    return data


def parse_command(sh: Worksheet, area: AreaTupleType, name: Optional[str], cmd_name: str) -> IssuesLabelContentTripleType:
    """
    Parse command in general
    Generate a JSON
    Generate a list of issues

    :param sh: Worksheet to read
    :param area: Area of the worksheet
    :param name: Name of the worksheet
    :param cmd_name: Name of the command. Key to access "command_fields" variable. Also, shown in issue descriptions
    :return: issues List, None, content (JSON)
    """

    issues: List[Issue] = []

    from backend.command_field_definitions import command_fields

    cols = command_fields[cmd_name]  # List of CommandField that will guide the parsing
    ##sh_dict = read_worksheet(sh)
    ##col_map, local_issues = check_columns(sh_dict, name, area, cols, cmd_name)
    col_map, local_issues = check_columns(sh, name, area, cols, cmd_name)

    if any([i.itype == IType.ERROR for i in local_issues]):
        return local_issues, None, None

    issues.extend(local_issues)

    # "mandatory" can be defined as expression depending on other base fields (like in RefBibliographic command fields)
    # Elaborate a list of fields having this "complex" mandatory property
    complex_mandatory_cols = [c for c in cols if isinstance(c.mandatory, str)]

    content = []  # The output JSON
    # Parse each Row
    for r in range(area[0] + 1, area[1]):
        line = {}
        expandable = False  # The line contains at least one field implying expansion into multiple lines
        complex = False  # The line contains at least one field with a complex rule (which cannot be evaluated with a simple cast)

        # Constant mandatory values
        mandatory_not_found = set([c.name for c in cols if c.mandatory and isinstance(c.mandatory, bool)])

        # Each "field"
        for col in col_map.keys():
            cname = col.name
            # Appearances of field (normally just once, there attributes allowing more than one appearance)
            for col_name, col_idx in col_map[col]:
                # Read and prepare "value"
                ##value = sh_dict.get((r, col_idx), None)
                value = sh.cell(row=r, column=col_idx).value
                if value is not None:
                    if not isinstance(value, str):
                        value = str(value)
                    value = value.strip()
                else:
                    continue

                if col.allowed_values:  # If the CommandField checks for a list of allowed values
                    if value.lower() not in [v.lower() for v in col.allowed_values]:  # TODO Case insensitive CI
                        issues.append(
                            Issue(itype=IType.ERROR,
                                  description=f"Field '{col_name}' of command '{cmd_name}' has invalid value '{value}'."
                                              f" Allowed values are: {', '.join(col.allowed_values)}.",
                                  location=IssueLocation(sheet_name=name, row=r, column=col_idx)))
                    else:
                        line[cname] = value
                else:  # Instead of a list of values, check if a syntactic rule is met by the value
                    if col.parser:  # Parse, just check syntax (do not store the AST)
                        try:
                            ast = parser_field_parsers.string_to_ast(col.parser, value)
                            # Rules are in charge of informing if the result is expandable and if it complex
                            if "expandable" in ast and ast["expandable"]:
                                expandable = True
                            if "complex" in ast and ast["complex"]:
                                complex = True

                            # With many appearances, just a "Key-Value list" syntax is permitted
                            if col.many_appearances:
                                if cname in line:
                                    line[cname] += ", " + col_name + "='" + value + "'"
                                else:
                                    line[cname] = col_name + "='" + value + "'"
                            else:
                                if cname in line:
                                    line[cname] += ", " + value
                                else:
                                    line[cname] = value  # Store the value
                        except:
                            ##col_header = sh_dict.get((1, col_idx), None)
                            col_header = sh.cell(row=1, column=col_idx).value
                            issues.append(Issue(itype=IType.ERROR,
                                                description="The value in field '" + col_header + "' of command '" + cmd_name + "' is not syntactically correct. Entered: " + value,
                                                location=IssueLocation(sheet_name=name, row=r, column=col_idx)))
                    else:
                        line[cname] = value  # No parser, just store blindly the value

            if col.name in mandatory_not_found:
                mandatory_not_found.discard(col.name)

        if len(line) == 0:
            continue  # Empty line (allowed)

        # Flags to accelerate the second evaluation, during execution
        line["_row"] = r
        line["_expandable"] = expandable
        line["_complex"] = complex

        # Append if all mandatory fields have been filled
        may_append = True
        if len(mandatory_not_found) > 0:
            issues.append(Issue(itype=IType.ERROR,
                                description="Mandatory columns: " + ", ".join(
                                    mandatory_not_found) + " have not been specified",
                                location=IssueLocation(sheet_name=name, row=r, column=None)))
            may_append = False

        # Check varying mandatory fields (fields depending on the value of other fields)
        for c in complex_mandatory_cols:
            col = c.name  # next(c2 for c2 in col_map if strcmp(c.name, c2.name))
            if isinstance(c.mandatory, str):
                # Evaluate
                mandatory = eval(c.mandatory, None, line)
                may_append = (mandatory and col in line) or (not mandatory)
                if mandatory and col not in line:
                    issues.append(Issue(itype=IType.ERROR,
                                        description="Mandatory column: " + col + " has not been specified",
                                        location=IssueLocation(sheet_name=name, row=r, column=None)))

        if may_append:
            content.append(line)

    return issues, None, {"items": content, "command_name": name}
