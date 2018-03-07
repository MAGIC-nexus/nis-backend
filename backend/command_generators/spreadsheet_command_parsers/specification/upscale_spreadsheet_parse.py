import openpyxl
import numpy as np

from backend.command_generators.spreadsheet_utils import worksheet_to_numpy_array, obtain_rectangular_submatrices


def parse_upscale_command(sh, area):
    """
    Analyze the input area
    Obtain the numerical part
    Obtain the tags identifying parent and child processors. Some tags may be used for both

    Read
    Most "parse" methods are mostly syntactic (as opposed to semantic). They do not check existence of names.
    But in this case, the valid field names are fixed beforehand, so they are checked at this time.
    Some of the fields will be controlled also, according to some

    :param sh: Input worksheet
    :param area: Tuple (top, bottom, left, right) representing the rectangular area of the input worksheet where the
    command is present
    :return: list of issues (issue_type, message), command label, command content
    """

    some_error = False
    issues = []
    # Read base fields
    # Parent Processor type	Child Processor type	Scaled Factor	Source
    parent_processor_type = None
    child_processor_type = None
    scaled_factor = None
    source = None
    for c in range(area[2], area[3]):  # Columns
        key = sh.cell(row=1, column=c).value
        value = sh.cell(row=2, column=c).value
        if key.lower() in ["parent"]:
            parent_processor_type = value
        elif key.lower() in ["child"]:
            child_processor_type = value
        elif key.lower() in ["scaled factor"]:
            scaled_factor = value
        elif key.lower() in ["source"]:  # "Observer"
            source = value

    # Detect the matrix defining scales
    m = binary_mask_from_worksheet(sh, True)  # "True" is to focus on cells containing numbers
    # Locate the matrix with numbers. Assume this defines the labels to consider, they will be around the matrix
    t = obtain_rectangular_submatrices(m)[0]  # Take just the first element
    v = worksheet_to_numpy_array(sh)
    for t in [(t[0], t[1], t[2], t[3]), (t[0] + 1, t[1], t[2], t[3]), (t[0], t[1], t[2] + 1, t[3])]:
        f = v[t[0]:t[1], t[2]:t[3]].astype(np.float64)
        row_sum = np.sum(f, axis=1)  # A column vector. If "all ones", the container will be along rows
        col_sum = np.sum(f, axis=0)  # A row vector. If "all ones", the container will be along columns
        container_situation = None
        if np.allclose(row_sum, 1, 1e-2):
            container_situation = "in_rows"
        if np.allclose(col_sum, 1, 1e-2):
            if container_situation:
                some_error = True
                issues.append((3, "Both rows and columns should not add up to 1.0"))
            container_situation = "in_columns"
        if container_situation:
            break
    if not container_situation:
        issues.append((2, "Neither the sum of rows nor of columns is summing to one"))

    # TODO Detect the factors

    content = {"parent_processor_type": parent_processor_type,
               "child_processor_type": child_processor_type,
               "scaled_factor": scaled_factor,
               "source": source,
               "column_headers": column_headers,  # List of lists
               "row_headers": row_headers,  # List of lists
               "scales": scales  # Matrix of scales, row by row
               }
    return issues, None, content


def binary_mask_from_worksheet(sh_in, only_numbers=True):
    """
    Sweep the worksheet, considering merged cells, elaborate a mask for those cells which
    are not empty or contain a number

    :param sh_in:
    :param only_numbers:
    :return:
    """
    m = np.zeros((sh_in.max_row, sh_in.max_column), dtype=bool)
    for r in range(sh_in.max_row):
        for c in range(sh_in.max_column):
            v = sh_in.cell(row=r + 1, column=c + 1).value
            if v:
                if only_numbers:
                    if isinstance(v, int) or isinstance(v, float):
                        m[r, c] = 1
                else:
                    m[r, c] = 1

    # Merged cells
    for ra in sh_in.merged_cell_ranges:
        t = openpyxl.utils.range_boundaries(ra)  # min col, min row, max col, max row (max's included)
        mc = (t[1]-1, t[3]-1, t[0]-1, t[2]-1)  # Rearrange and subtract one
        v = m[mc[0], mc[2]]
        m[mc[0]:mc[1]+1, mc[2]:mc[3]+1] = v

    return m
