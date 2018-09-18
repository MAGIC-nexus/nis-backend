import openpyxl

import io
# import koala # An Excel files parser elaborating a graph allowing automatic evaluation
import regex as re  # Improvement over standard "re"
from backend import Issue
import backend
from backend.command_executors.analysis.indicators_command import IndicatorsCommand
from backend.command_executors.external_data.etl_external_dataset_command import ETLExternalDatasetCommand
from backend.command_executors.external_data.mapping_command import MappingCommand
from backend.command_executors.external_data.parameters_command import ParametersCommand
from backend.command_executors.specification.data_input_command import DataInputCommand
from backend.command_executors.specification.hierarchy_command import HierarchyCommand
from backend.command_executors.specification.metadata_command import MetadataCommand
from backend.command_executors.specification.pedigree_matrix_command import PedigreeMatrixCommand
from backend.command_executors.specification.references_command import ReferencesCommand
from backend.command_executors.specification.scale_conversion_command import ScaleConversionCommand
from backend.command_executors.specification.structure_command import StructureCommand
from backend.command_executors.specification.upscale_command import UpscaleCommand
from backend.command_executors.version2.hierarchy_categories_command import HierarchyCategoriesCommand
from backend.command_executors.version2.hierarchy_mapping_command import HierarchyMappingCommand
from backend.command_executors import create_command, DatasetDataCommand, DatasetQryCommand, AttributeTypesCommand, \
    AttributeSetsCommand, InterfaceTypesCommand, ProcessorsCommand, InterfacesAndQualifiedQuantitiesCommand, \
    RelationshipsCommand, InstantiationsCommand, ScaleConversionV2Command, DatasetDefCommand
from backend.command_generators.spreadsheet_command_parsers.analysis.indicators_spreadsheet_parse import parse_indicators_command
from backend.command_generators.spreadsheet_command_parsers.external_data.mapping_spreadsheet_parse import parse_mapping_command
from backend.command_generators.spreadsheet_command_parsers.external_data.etl_external_dataset_spreadsheet_parse import parse_etl_external_dataset_command
from backend.command_generators.spreadsheet_command_parsers.external_data.parameters_spreadsheet_parse import parse_parameters_command
from backend.command_generators.spreadsheet_command_parsers.specification.hierarchy_spreadsheet_parser import parse_hierarchy_command
from backend.command_generators.spreadsheet_command_parsers.specification.metadata_spreadsheet_parse import parse_metadata_command
from backend.command_generators.spreadsheet_command_parsers.specification.data_input_spreadsheet_parse import parse_data_input_command
from backend.command_generators.spreadsheet_command_parsers.specification.pedigree_matrix_spreadsheet_parse import \
    parse_pedigree_matrix_command
from backend.command_generators.spreadsheet_command_parsers.specification.references_spreadsheet_parser import \
    parse_references_command
from backend.command_generators.spreadsheet_command_parsers.specification.scale_conversion_spreadsheet_parse import \
    parse_scale_conversion_command
from backend.command_generators.spreadsheet_command_parsers.specification.upscale_spreadsheet_parse import parse_upscale_command
from backend.command_generators.spreadsheet_command_parsers.specification.structure_spreadsheet_parser import parse_structure_command
from backend.command_generators.parser_spreadsheet_utils import binary_mask_from_worksheet, obtain_rectangular_submatrices
from backend.command_generators.spreadsheet_command_parsers_v2.dataset_data_spreadsheet_parse import \
    parse_dataset_data_command
from backend.command_generators.spreadsheet_command_parsers_v2.dataset_qry_spreadsheet_parse import \
    parse_dataset_qry_command
from backend.command_generators.spreadsheet_command_parsers_v2.simple_parsers import parse_cat_hierarchy_command, \
    parse_hierarchy_mapping_command, parse_parameters_command_v2, parse_attribute_sets_command, \
    parse_attribute_types_command, parse_datasetdef_command, parse_interface_types_command, parse_processors_v2_command, \
    parse_interfaces_command, parse_relationships_command, parse_instantiations_command, parse_scale_changers_command, \
    parse_shared_elements_command, parse_reused_elements_command, parse_indicators_v2_command

# Most complex name
# [namespace::][type:]var(.var)*[(@|#)var] : [namespace] + [type] + var + [attribute or tag]
var_name = "([a-zA-Z][a-zA-Z0-9_-]*)"
hvar_name = "("+var_name + r"(\." + var_name + ")*)"
cplex_var = "((" + var_name + "::)?" + hvar_name + ")"

# ############################### #
#  Main function                  #
# ############################### #


def commands_generator_from_ooxml_file(input, state):
    """
    It reads an Office Open XML input
    Yields a sequence of command_executors

Hoja comando
* Lex+Parse
* Producir JSON ~ AST
* Enumerar problemas sintácticos
* Producir hoja a partir de JSON??

Comando
* Analizar JSON
* Ejecutar
* Enumerar problemas semánticos

    :param input: A bytes input
    :param state: State used to check variables
    :return:
    """
    # Start the Excel reader
    xl_in = openpyxl.load_workbook(io.BytesIO(input), data_only=True)

    # Regular expressions for the different commands
    flags = re.IGNORECASE
    optional_alphanumeric = "([ a-zA-Z0-9_-]*)?"  # Whitespace also allowed

    #
    # NOTE #1: Different command states, regarding command development evolution:
    # * Old supported. A new approach is taken
    # * Old deprecated
    # * Old renewed
    # * New
    #
    # NOTE #2: If a command name is used for a newer command, the old version would be preceded by a version number:
    # * Processors -> V1Processors
    #
    # NOTE #3: Some of the commands have a special form, preventing its inclusion in the "cmds" variable below. Those are:
    # * re_data - etl_dataset
    # * re_mapping - mapping
    # * re_hierarchy - hierarchy

    # WORKSHEET NAME REGULAR EXPRESSIONS
    re_metadata = re.compile(r"^(Metadata)" + optional_alphanumeric, flags=flags)  # Shared by V1 and V2
    re_processors = re.compile(r"(Processors|Proc)[ _]+" + var_name, flags=flags)  # V1. Deprecated in V2 (use "re_interfaces")
    re_hierarchy = re.compile(r"(Taxonomy|Tax|Composition|Comp)[ _]([cpf])[ ]" + var_name, flags=flags)  # V1. Deprecated in V2 (use "re_hierarchies"). SPECIAL
    re_upscale = re.compile(r"(Upscale|Up)[ _](" + var_name + "[ _]" + var_name + ")?", flags=flags)  # V1. Deprecated in V2 (use "re_instantiations")
    re_relations = re.compile(r"(Grammar|Structure|Relations|Rel)([ _]+" + var_name+")?", flags=flags)  # V1. Deprecated in V2 (use "re_relationships")
    re_scale_conversion = re.compile(r"Scale", flags=flags)  # V1. Deprecated in V2 (use "re_scale_changers")
    re_pedigree_template = re.compile(r"(Pedigree|Ped|NUSAP\.PM)[ _]+" + var_name, flags=flags)  # V1. Deprecated in V2 (use "re_pedigree_matrices") ???? TO BE SOLVED <<<<<<< MAYBE V1 is better option
    re_references = re.compile(r"(References|Ref)[ _]" + var_name, flags=flags)  # V1. Maybe valid also in V2 (if deprecated, use the different specialized "re_ref*" commands)
    re_parameters = re.compile(r"(Parameters|Params)" + optional_alphanumeric, flags=flags)  # Shared by V1 and V2 (not used in V1, so V2 directly)
    re_data = re.compile(r"(Dataset|DS)[ _]" + hvar_name, flags=flags)  # V1. Deprecated in V2 (use "re_datasetqry"). SPECIAL
    re_mapping = re.compile(r"^(Mapping|Map)([ _]" + cplex_var + "[ _]" + cplex_var + ")?", flags=flags)  # V1. Deprecated in V2 (use "re_hierarchies_mapping"). SPECIAL
    re_indicators = re.compile(r"(Indicators|KPI)([ _]" + var_name + ")?", flags=flags)  # Shared by V1 and V2 (not used in V1, so v2 directly)
    # Version 2 commands
    re_hierarchies = re.compile(r"(CatHierarchies|Categories)" + optional_alphanumeric, flags=flags)  # Hierarchies for categories (formerly "Taxonomy_C")
    re_hierarchies_mapping = re.compile(r"(CatHierarchiesMapping|CategoriesMap)" + optional_alphanumeric, flags=flags)
    re_attributes = re.compile(r"(AttributeTypes)" + optional_alphanumeric, flags=flags)  # Declaration of attributes used in different elements
    re_datasetdef = re.compile(r"(DatasetDef)" + optional_alphanumeric, flags=flags)  # Dataset metadata
    re_datasetdata = re.compile(r"(DatasetData)" + optional_alphanumeric, flags=flags)  # Dataset data
    re_attribute_sets = re.compile(r"(AttributeSets)" + optional_alphanumeric, flags=flags)
    # re_parameters
    re_datasetqry = re.compile(r"(DatasetQry)" + optional_alphanumeric, flags=flags)  # Dataset Query
    re_interfacetypes = re.compile(r"(InterfaceTypes)" + optional_alphanumeric, flags=flags)  # Declaration of Interface types and hierarchies
    re_processors_v2 = re.compile(r"(Processors2)" + optional_alphanumeric, flags=flags)  # It is NOT the next version of "re_processors" (which is "re_interfaces")
    re_interfaces = re.compile(r"(Interfaces)" + optional_alphanumeric, flags=flags)  # Interfaces and data. V2 of "re_processors"
    re_relationships = re.compile(r"(Flows|Relationships)" + optional_alphanumeric, flags=flags)
    re_instantiations = re.compile(r"(Instantiations)" + optional_alphanumeric, flags=flags)
    re_scale_changers = re.compile(r"(ScaleChangers) + optional_alphanumeric", flags=flags)
    re_shared_elements = re.compile(r"(SharedElements)" + optional_alphanumeric, flags=flags)
    re_reused_elements = re.compile(r"(ReusedElements)" + optional_alphanumeric, flags=flags)
    re_pedigree_matrices = re.compile(r"PedigreeMatrices" + optional_alphanumeric, flags=flags)
    re_refbibliographic = re.compile(r"RefBibliographic" + optional_alphanumeric, flags=flags)
    re_refsource = re.compile(r"RefSource" + optional_alphanumeric, flags=flags)
    re_refgeographical = re.compile(r"RefGeographical" + optional_alphanumeric, flags=flags)
    re_refprovenance = re.compile(r"RefProvenance" + optional_alphanumeric, flags=flags)
    # re_indicators
    # re_indicators_benchmark NOT DEFINED, NOT IMPLEMENTED
    re_problem_statement = re.compile(r"ProblemStatement" + optional_alphanumeric, flags=flags)

    # List of tuples defining each command parsing:
    # - Regular expression
    # - Command type, used later to obtain the IExecutableCommand instance
    # - Parse function
    # - Number of arguments received by parse function
    cmds = [(re_metadata, "metadata", parse_metadata_command, 2, MetadataCommand),  # V1 and V2
            (re_processors, "data_input", parse_data_input_command, 3, DataInputCommand),  # V1
            (re_relations, "structure", parse_structure_command, 2, StructureCommand),  # V1
            (re_scale_conversion, "scale_conversion", parse_scale_conversion_command, 2, ScaleConversionCommand),  # V1
            (re_upscale, "upscale", parse_upscale_command, 2, UpscaleCommand),  # V1
            (re_pedigree_template, "pedigree_matrix", parse_pedigree_matrix_command, 3, PedigreeMatrixCommand),  # V1 and V2
            (re_references, "references", parse_references_command, 2, ReferencesCommand),  # V1 MAYBE V2
            # Special (declared but not used)
            (re_hierarchy, "hierarchy", 0, HierarchyCommand),
            (re_data, "etl_dataset", 0, ETLExternalDatasetCommand),
            (re_mapping, "mapping", 0, MappingCommand),
            # V2 commands
            (re_hierarchies_mapping, "cat_hier_mapping", parse_hierarchy_mapping_command, 3, HierarchyMappingCommand),  # TODO (2***)
            (re_hierarchies, "cat_hierarchies", parse_cat_hierarchy_command, 3, HierarchyCategoriesCommand),
             (re_attributes, "attribute_types", parse_attribute_types_command, 2, AttributeTypesCommand),  # TODO Attribute Types (1***)
             (re_attribute_sets, "attribute_sets", parse_attribute_sets_command, 3, AttributeSetsCommand),  # TODO (2***)
            (re_datasetdef, "datasetdef", parse_datasetdef_command, 2, DatasetDefCommand),  # TODO Dataset Metadata (3***)
            (re_datasetdata, "datasetdata", parse_dataset_data_command, 2, DatasetDataCommand),  # TODO Dataset Data   (3***)
            (re_parameters, "parameters", parse_parameters_command_v2, 3, ParametersCommand),  # The old function was "parse_parameters_command"
            (re_datasetqry, "datasetqry", parse_dataset_qry_command, 2, DatasetQryCommand),  # TODO Dataset Query. Very similar to "etl_dataset" IExecutableCommand (3***)
            (re_interfacetypes, "interface_types", parse_interface_types_command, 2, InterfaceTypesCommand),  # TODO (2***)
            (re_processors_v2, "processors", parse_processors_v2_command, 2, ProcessorsCommand),  # TODO (4***)
            (re_interfaces, "interfaces_and_qq", parse_interfaces_command, 2, InterfacesAndQualifiedQuantitiesCommand),  # TODO (4***)(evolution of "re_processors" "data_input")
            (re_relationships, "relationships", parse_relationships_command, 2, RelationshipsCommand),  # TODO (4***)(evolution of "re_relations" "structure")
            (re_instantiations, "instantiations", parse_instantiations_command, 2, InstantiationsCommand),  # TODO (5***)(evolution of "re_upscale" "upscale")
            (re_scale_changers, "scale_conversion_v2", parse_scale_changers_command, 2, ScaleConversionV2Command),  # TODO (5***)Relations of conversion between interface types
            (re_problem_statement, "problem_statement",),  # TODO

            (re_shared_elements, "shared_elements", parse_shared_elements_command, 2, ),  # TODO
            (re_reused_elements, "reused_elements", parse_reused_elements_command, 2, ),  # TODO
            (re_indicators, "indicators", parse_indicators_v2_command, 2, IndicatorsCommand),  # (V1 and) V2

            (re_pedigree_matrices, "ref_pedigree_matrices", ),  # TODO
            (re_refbibliographic, "ref_bibliographic", ),  # TODO
            (re_refsource, "ref_source", ),  # TODO
            (re_refgeographical, "ref_geographical", ),  # TODO
            (re_refprovenance, "ref_provenance", ),  # TODO
            ]

    # For each worksheet, get the command type, convert into primitive JSON
    for c, sh_name in enumerate(xl_in.sheetnames):
        issues = []
        total_issues = []  # type: List[Issue]
        sh_in = xl_in[sh_name]

        c_type = None
        c_label = None
        c_content = None

        name = sh_in.title

        # Extract worksheet matrices
        m = binary_mask_from_worksheet(sh_in, False)
        t = obtain_rectangular_submatrices(m, only_remove_empty_bottom=True)
        if len(t) == 0:  # No data
            continue

        t = t[0]  # Take just the first element, a tuple (top, bottom, left, right) representing a rectangular region
        t = (t[0] + 1, t[1] + 1, t[2] + 1, t[3] + 1)  # Indices start at 1

        # v = worksheet_to_numpy_array(sh_in)

        # COMMAND parse
        if re_data.search(name):
            dataset = re_data.search(name).group(2)
            c_type = "etl_dataset"
            if sh_in.cell(row=t[0], column=t[2]).value:
                t = (1, m.shape[0]+1, 1, m.shape[1]+1)
                # Parse to read parameters
                issues, c_label, c_content = parse_etl_external_dataset_command(sh_in, t, dataset, state)
            else:
                # Syntax error: it seems there are no parameters
                issue = {"sheet_number": c, "sheet_name": sh_name, "c_type": c_type, "type": 3,
                         "message": "It seems there are no parameters for the dataset import command at worksheet '" + sh_name + "'"}
                total_issues.append(issue)
        elif re_datasetqry.search(name):
            c_type = "datasetqry"
            issues, c_label, c_content = parse_dataset_qry_command(sh_in, t, sh_name, state)
        elif re_datasetdata.search(name):
            c_type = "datasetdata"
            issues, c_label, c_content = parse_dataset_data_command(sh_in, t, sh_name, state)
        elif re_mapping.search(name):
            c_type = "mapping"
            g = re_mapping.search(name).groups()
            if g[2] and g[8]:
                origin = g[2]
                destination = g[8]
            elif not g[2] and not g[8]:
                origin = None
                destination = None
            else:
                issue = {"sheet_number": c, "sheet_name": sh_name, "c_type": c_type, "type": 3, "message": "Either origin or destination are not correctly specified in the sheet name '"+sh_name+"'"}
                total_issues.append(issue)
            issues, c_label, c_content = parse_mapping_command(sh_in, t, origin, destination)
        elif re_hierarchy.search(name):
            # Read the hierarchy
            c_type = "hierarchy"
            res = re_hierarchy.search(name)
            h_type = res.group(2)
            c_label = res.group(3)
            issues, _, c_content = parse_hierarchy_command(sh_in, t, c_label, h_type)
        else:  # GENERIC command parser. Based on template of columns (fields)
            for cmd in cmds:
                if cmd[0].search(name):
                    c_type = cmd[1]
                    if cmd[3] == 2:  # Call parser, TWO parameters: the worksheet & the rectangular area to parse
                        issues, c_label, c_content = cmd[2](sh_in, t)
                    elif cmd[3] == 3:  # Call parser, THREE params: the worksheet, area to parse AND the worksheet Name
                        name2 = cmd[0].search(name).group(2)
                        issues, c_label, c_content = cmd[2](sh_in, t, name2)

        # Append issues
        errors = 0
        if len(issues) > 0:
            for i in issues:
                if isinstance(i, backend.command_generators.Issue):
                    if i.itype == 3:
                        errors += 1
                    issue = {"sheet_number": c, "sheet_name": sh_name, "c_type": c_type, "type": i.itype, "message": i.description}
                else:
                    if i[0] == 3:
                        errors += 1
                    issue = {"sheet_number": c, "sheet_name": sh_name, "c_type": c_type, "type": i[0], "message": i[1]}
                total_issues.append(issue)
        if errors == 0:
            try:
                cmd, issues = create_command(c_type, c_label, c_content, sh_name)
            except:
                cmd = None
                issues = [(3, "Could not create command of type '"+c_type+"'")]
            if issues:
                for i in issues:
                    if isinstance(i, backend.command_generators.Issue):
                        issue = {"sheet_number": c, "sheet_name": sh_name, "c_type": c_type, "type": i.itype,
                                 "message": i.description}
                    else:
                        issue = {"sheet_number": c, "sheet_name": sh_name, "c_type": c_type, "type": i[0],
                                 "message": i[1]}

                    total_issues.append(issue)
        else:
            cmd = None  # cmd, _ = create_command(c_type, c_label, {}, sh_name)

        yield cmd, total_issues
    # yield from []  # Empty generator


# def get_codes_all_statistical_datasets(source, dataset_manager):
#     """
#     Obtain a list of datasets available from a source
#     If no source is specified, all the sources are queried
#     For each dataset, the source, the name, the periods available, an example command and a description are obtained
#
#     :param source:
#     :param dataset_manager: It is a DataSourceManager
#     :return: A Dataframe with the list of datasets
#     """
#     lst2 = []
#     # TODO Probably "get_datasets" will not work as expected. It returns a tuple (Source, list of datasets)
#     for r, k in enumerate(dataset_manager.get_datasets(source)):
#         if len(k) == 4:
#             src = k[3]
#         else:
#             src = ""
#         lst2.append((k[0], k[1], k[2], src))
#     return pd.DataFrame(data=lst2, columns=["Dataset ID", "Description", "URN", "Data Source"])

