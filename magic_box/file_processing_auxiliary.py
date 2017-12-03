# -*- coding: utf-8 -*-


import json


import sqlalchemy.orm
import sqlalchemy.schema
from magic_box.model import DBSession, ORMBase
import io
import openpyxl
import openpyxl.utils
from openpyxl.comments import Comment
import pandas as pd
import numpy as np
import requests
from multidict import MultiDict, CIMultiDict
import pint  # Units management
import re
import collections
import copy
import pandasdmx
from magic_box import app
from magic_box.source_eurostat_bulk import get_eurostat_filtered_dataset_into_dataframe
from openpyxl.styles import PatternFill

# GLOBAL VARIABLES
case_sensitive = False
ureg = pint.UnitRegistry()


# #################################################################
# CASE SeNsItIvE or INSENSITIVE names (flows, funds, processors, ...)
#

class CaseInsensitiveDict(collections.MutableMapping):
    """
    A dictionary with case insensitive Keys.
    Prepared also to support TUPLES as keys, required because compound keys are required
    """
    def __init__(self, data=None, **kwargs):
        from collections import OrderedDict
        self._store = OrderedDict()
        if data is None:
            data = {}
        self.update(data, **kwargs)

    def __setitem__(self, key, value):
        # Use the lowercased key for lookups, but store the actual
        # key alongside the value.
        if not isinstance(key, tuple):
            self._store[key.lower()] = (key, value)
        else:
            self._store[tuple([k.lower() for k in key])] = (key, value)

    def __getitem__(self, key):
        if not isinstance(key, tuple):
            return self._store[key.lower()][1]
        else:
            return self._store[tuple([k.lower() for k in key])][1]

    def __delitem__(self, key):
        if not isinstance(key, tuple):
            del self._store[key.lower()]
        else:
            del self._store[tuple([k.lower() for k in key])]

    def __iter__(self):
        return (casedkey for casedkey, mappedvalue in self._store.values())

    def __len__(self):
        return len(self._store)

    def lower_items(self):
        """Like iteritems(), but with all lowercase keys."""
        return (
            (lowerkey, keyval[1])
            for (lowerkey, keyval)
            in self._store.items()
        )

    def __eq__(self, other):
        if isinstance(other, collections.Mapping):
            other = CaseInsensitiveDict(other)
        else:
            return NotImplemented
        # Compare insensitively
        return dict(self.lower_items()) == dict(other.lower_items())

    # Copy is required
    def copy(self):
        return CaseInsensitiveDict(self._store.values())

    def __repr__(self):
        return str(dict(self.items()))


def create_dictionary(multi_dict=False):
    """
    Factory to create dictionaries used by the prototype

    It reads the "case_sensitive" global variable

    :param multi_dict: True to create a "MultiDict", capable of storing several values
    :return:
    """

    if not multi_dict:
        if case_sensitive:
            return {}
        else:
            return CaseInsensitiveDict()
    else:
        if case_sensitive:
            return MultiDict()
        else:
            return CIMultiDict()


def strcmp(s1, s2):
    """
    Compare two strings for equality or not, considering a flag for case sensitiveness or not

    It also removes leading and trailing whitespace from both strings, so it is not sensitive to this possible
    difference, which can be a source of problems

    :param s1:
    :param s2:
    :return:
    """
    if case_sensitive:
        return s1.strip() == s2.strip()
    else:
        return s1.strip().lower() == s2.strip().lower()
#
#
# #################################################################


def create_after_worksheet(sh, title):
    """
    Create a new worksheet after the worksheet passed as parameter

    :param sh:
    :param title:
    :return: The new worksheet
    """
    def get_worksheet_index():
        for i, sh_name in enumerate(wb.get_sheet_names()):
            if sh_name == sh.title:
                break
        return i

    wb = sh.parent
    i = get_worksheet_index()
    return wb.create_sheet(title, i+1)


def reset_worksheet(sh):
    """
    Reset a worksheet by deleting and creating it from scratch

    :param sh:
    :return: The new worksheet
    """
    def get_worksheet_index():
        for i, sh_name in enumerate(wb.get_sheet_names()):
            if sh_name == sh.title:
                break
        return i

    wb = sh.parent
    i = get_worksheet_index()
    tmp = sh.title
    wb.remove_sheet(sh)
    return wb.create_sheet(tmp, i)


global_fill = PatternFill("none")


def reset_cell_format(sh, r, c):
    cell = sh.cell(row=r, column=c)
    cell.fill = global_fill
    cell.comment = None


def reset_cells_format(sh_writable):
    for r in range(sh_writable.max_row):
        for c in range(sh_writable.max_column):
            reset_cell_format(sh_writable, r + 1, c + 1)


def cell_content_to_str(v):
    """
    Auxiliary function

    :param v:
    :return:
    """
    if v:
        if isinstance(v, float) or isinstance(v, int):
            return str(int(v))
        else:
            return str(v).strip()
    else:
        return None


def show_message(sh, r, c, message, type="error", accumulate=True):
    """
    It serves to show a cell with some type of error (warning or error)
    A message is shown in the comment
    The name of the sheet is changed with a prefix indicating there is at least an issue to be solved

    :param sh:
    :param r:
    :param c:
    :param type:
    :param message:
    :return:
    """
    cell = sh.cell(row=r, column=c)
    fill = cell.fill
    if type == "error":
        fill = PatternFill("solid", fgColor="CC0000")
    elif type == "warning":
        fill = PatternFill("solid", fgColor="FFFF33")
    elif type == "info":
        fill = PatternFill("solid", fgColor="87CEEB")
    cell.fill = fill
    if accumulate:
        comment = cell.comment
        if comment:
            comment.text += "\n" + message
        else:
            comment = Comment(message, "NIS")
    else:
        comment = Comment(message, "NIS")
    cell.comment = comment
    # if type == "error":
    #     sh.title = "!" + sh.title


def obtain_rectangular_submatrices(mask, region=None):
    """
    Obtain rectangular submatrices of mask
    IMPORTANT: currently it only obtains ONE region

    :param mask: The original matrix, numpy.NDArray, containing only 0/1 (1 is "some content")
    :param region: A tuple (top, bottom, left, right) with indices to search. bottom and right are not included
    :return: The list of rectangular regions as tuples (top, bottom, left, right)
    """

    def nonzero_sequences(a):
        # Create an array that is 1 where a is non-zero, and pad each end with an extra 0.
        isnonzero = np.concatenate(([0], a != 0, [0]))
        absdiff = np.abs(np.diff(isnonzero))
        # Runs start and end where absdiff is 1.
        ranges = np.where(absdiff == 1)[0].reshape(-1, 2)
        return ranges

    lst = []
    if not region:
        region = (0, mask.shape[0], 0, mask.shape[1])  # All the mask
    submask = mask[region[0]:region[1], region[2]:region[3]]
    offset_col, offset_row = (region[2], region[0])
    # Accumulation of elements by row (resulting in a column vector)
    row_sum = np.sum(submask, axis=1)
    # Accumulation of elements by column (resulting in a row vector)
    col_sum = np.sum(submask, axis=0)

    # Ranges
    rs = nonzero_sequences(row_sum.flatten())
    cs = nonzero_sequences(col_sum.flatten())
    lst.append((rs[0][0], rs[0][1], cs[0][0], cs[0][1]))

    return lst


def worksheet_to_numpy_array(sh_in):
    """
    Obtain a replica of the worksheet into a Numpy NDArray, with combined cells (combined cells are repeated)

    :param sh_in:
    :return: The numpy array with the values of the worksheet
    """
    m = np.zeros((sh_in.max_row, sh_in.max_column)).astype(object)
    for r in range(sh_in.max_row):
        for c in range(sh_in.max_column):
            v = sh_in.cell(row=r + 1, column=c + 1).value
            if v:
                m[r, c] = v
            else:
                m[r, c] = 0.0

    # Merged cells
    for ra in sh_in.merged_cell_ranges:
        t = openpyxl.utils.range_boundaries(ra)  # min col, min row, max col, max row (max's included)
        mc = (t[1]-1, t[3]-1, t[0]-1, t[2]-1)  # Rearrange and subtract one
        v = m[mc[0], mc[2]]
        m[mc[0]:mc[1]+1, mc[2]:mc[3]+1] = v

    return m


def binary_mask_from_worksheet(sh_in, only_numbers=True):
    """
    Sweep the worksheet, considering merged cells, elaborate a mask for those cells which
    are not empty or contain a number

    :param sh_in:
    :param only_numbers:
    :return:
    """
    m = np.zeros((sh_in.max_row, sh_in.max_column), dtype=bool)
    for r in range(sh_in.max_row):
        for c in range(sh_in.max_column):
            v = sh_in.cell(row=r + 1, column=c + 1).value
            if v:
                if only_numbers:
                    if isinstance(v, int) or isinstance(v, float):
                        m[r, c] = 1
                else:
                    m[r, c] = 1

    # Merged cells
    for ra in sh_in.merged_cell_ranges:
        t = openpyxl.utils.range_boundaries(ra)  # min col, min row, max col, max row (max's included)
        mc = (t[1]-1, t[3]-1, t[0]-1, t[2]-1)  # Rearrange and subtract one
        v = m[mc[0], mc[2]]
        m[mc[0]:mc[1]+1, mc[2]:mc[3]+1] = v

    return m
