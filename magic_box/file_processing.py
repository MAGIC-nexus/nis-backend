# -*- coding: utf-8 -*-
import copy


import re

import numpy as np
import openpyxl
import openpyxl.utils
import pandas as pd
import regex as re

from backend.common.create_database import create_pg_database_engine, create_monet_database_engine
from backend.ie_imports.data_source_manager import DataSourceManager
from backend.ie_imports.data_sources.eurostat_bulk import Eurostat
from backend.ie_imports.data_sources.faostat import FAOSTAT
import backend

from magic_box import app, the_registry
from magic_box.file_processing_auxiliary import (
    create_dictionary, strcmp, create_after_worksheet, reset_worksheet, reset_cells_format, cell_content_to_str,
    show_message, obtain_rectangular_submatrices, worksheet_to_numpy_array, binary_mask_from_worksheet,
)
from magic_box.file_processing_datasets import (
    get_codes_all_statistical_datasets, get_statistical_dataset_structure, map_codelists,
    obtain_reverse_codes
)


#                                 else:
#                                     # TODO If the param is organized as a tree, prioritize: latest wins
# """
# * taxonomy
# * additive upscaling
# * transform dataset to processor facts
#
# * trees
#   - It cannot be assumed that the hierarchy of codes result in aggregations
#   - So the mapping cannot benefit from this
# * sequences
#   - need some Excel
#   - connect two ends. If one of them has a value, provide the value to the other side. Accumulate if several are specified
#     - source
#     - end
#     - factor
#     - ¿use expressions? because some will have values (constraints) the others need to compute
#
# * user's guide
# * análisis:
#   - modelo MuSIASEM. Modelar la elaboración
#   - gestión de casos de estudio
#
# * gis data. Inspire. Local IDEs: spain, canarias, other regions
# * gis data processing
# * other data sources:
#   - local, sectoral
#
# * salvador: considerations on economic
#
# * ANGULAR2. User interface in general. ¿VAADIN?
# """


# ------------------------------------------------------------------------


def process_metadata(sh_in):
    """
    Analyze metadata

    :param sh_in: Input/output sheet containing metadata
    :return:
    """
    def read_metadata():
        for r in range(sh_in.max_row):
            k = None
            for c in range(sh_in.max_column):
                cell = sh_in.cell(row=r + 1, column=c + 1)
                if c == 0:  # Key
                    k = cell.value
                else:
                    v = cell.value
                    if k in d:
                        l_values = d[k]
                    else:
                        l_values = []
                        d[k] = l_values
                    l_values.append(v)

    def write_and_validate_metadata():
        # Process each field then remove from the dictionary
        r = 0  # First row
        # Simple DC fields not covered:
        #  type (controlled),
        #  format (controlled),
        #  rights (controlled),
        #  publisher,
        #  contributor,
        #  relation
        #
        # XML Dublin Core: http://www.dublincore.org/documents/dc-xml-guidelines/
        # Exhaustive list: http://dublincore.org/documents/dcmi-type-vocabulary/

        # Fields: ("<field label in Spreadsheet file>", "<field name in Dublin Core>", Mandatory?, Controlled?)
        # TODO For Spreadsheet fields going to the same DC field AND with no controlled vocabulary, it is not possible to
        #      do the reverse transformation, from DC to Spreadsheet file, immediately. Only if some kind of syntax is enforced
        #      for each of these fields
        k_list = [("Case study name", "title", True, False),  # DEPRECATED
                  ("Case study code", "title", True, False),
                  ("Title", "title", True, False),
                  ("Subject, topic and/or keywords", "subject", False, True),
                  ("Description", "description", False, False),
                  ("Level", "description", False, True),
                  ("Dimensions", "subject", True, True),
                  ("Reference documentation", "source", False, False),
                  ("Authors", "creator", True, False),
                  ("Date of elaboration", "date", True, False),
                  ("Temporal situation", "coverage", True, False),
                  ("Geographical location", "coverage", True, True),
                  ("DOI", "identifier", False, False),
                  ("Language", "language", True, True)
                  ]
        for k in k_list:
            if k[0] in d:
                first = True
                for c, v in enumerate(d[k[0]]):
                    if isinstance(v, int) or (isinstance(v, str) and v.strip() != ""):
                        # TODO Control possible values of "v" if k[3] is TRUE. v in controlled_vocabulary?
                        # sh_in.cell(row=r + 1, column=0 + 1).value = k[0]
                        # sh_in.cell(row=r + 1, column=c + 1 + 1).value = v
                        dc.append((k[1], v))
                    else:
                        if first and k[2]:
                            # sh_in.cell(row=r + 1, column=0 + 1).value = k[0]
                            s = "'" + k[0] + "' was not specified"
                            # sh_in.cell(row=r + 1, column=2 + 1).value = s
                            issues.append(s)
                            break
                    first = False
            else:
                if k[2]:
                    # sh_in.cell(row=r + 1, column=0 + 1).value = k[0]
                    s = "'" + k[0] + "' field should be present. Inserted automatically."
                    # sh_in.cell(row=r + 1, column=2 + 1).value = s
                    issues.append(s)
            r += 1
        return dc, issues

    def generate_dublin_core_xml():
        s = """
<?xml version="1.0"?>

<caseStudyMetadata
  xmlns="http://magic-nexus.org/dmp/"
  xmlns:dc="http://purl.org/dc/elements/1.1/">
        """
        for t in dc:
            s += "  <dc:" + t[0] + ">" + str(t[1]) + "</dc:" + t[0] + ">\n"

        s += "</caseStudyMetadata>"

        return s

    # -----------------------------------------------------------------

    # Elaborate a record
    d = create_dictionary()
    # Read information into dictionary
    read_metadata()
    # Put information back, validated (controlled vocabularies only)
    issues = []
    dc = []
    write_and_validate_metadata()

    xml = generate_dublin_core_xml()

    sh_in.cell(row=30, column=1).value = xml


def clone_processor(p):
    """
    Create a new processor with the same properties as the original
    Recursive copy child processors

    :param p:
    :return:
    """
    return copy.deepcopy(p)


def adimensionally_scale_processor(p, number):
    """
    Multiply ALL flows and funds of a processor and its descendants by "number"
    If a unit is specified, for each intensive processor with the same

    :param p:
    :param number:
    :return:
    """
    # If a unit is specified and the processor is intensive
    if not p["intensive"]:
        print("ERROR: in order to be able to multiply by a factor, the processor has to be intensive, not extensive.")
        return

    if "factors" in p:
        p["factors"] += " * " + str(number)
    else:
        p["factors"] = str(number)

    flows_funds = p["ff"]
    for ff in flows_funds:
        # Store the different factors as an expression which can be evaluated later (using "eval"). Append up-scales
        if flows_funds[ff]["unit"]:  # Only if there is a unit. If it is adimensional, do not apply the factor
            flows_funds[ff]["value"] *= number
        # TODO Register the scaling operation

    # Recurse into descendants
    if "children" in p:
        for ch in p["children"]:
            adimensionally_scale_processor(ch, number)


def read_processors(sh_in, sh_out, registry, dfs):
    """
    Read a set of processors
    They will come with flows and funds (with value, unit, date, source, comment), types and name
    They need also a relation with the processor which originates the specialization, the combination matrix (or other structure) and so

    :param sh_in:
    :param sh_out:
    :param registry: Processors
    :params dfs: Dictionary of datasets (including PivotTable's)
    :return: String with the name of the entry added to the registry
    """
    def process_row(curr, rr):
        error = False
        # Copy value to output
        for c in range(sh_in.max_column):
            if curr[c] is not None:  # If empty, assume the last one
                sh_out.cell(row=rr + 1, column=c + 1, value=str(curr[c]))  # Copy to out worksheet

        # Full taxon
        taxon = tuple([first_taxonomic_rank] + [cell_content_to_str(curr[c]) for c in type_cols if curr[c] != ""])
        # TODO Check: there cannot exist an empty type value followed by a type to its right
        if taxon not in proc_taxonomy:
            proc_taxonomy[taxon] = None
        # Add subtypes to their corresponding taxonomic ranks
        for i, c in enumerate(type_cols):
            # Store a new entry (curr[c]) in the taxonomic rank (type_cols_names[i])
            if cell_content_to_str(curr[c]) not in taxonomic_ranks[type_cols_names[i]]:
                taxonomic_ranks[type_cols_names[i]][cell_content_to_str(curr[c])] = None

        # Second, the processor, which includes the types and the name
        if taxon in procs:
            p = procs[taxon]
        else:
            p = {"intensive": intensive_processors_specified,
                 "relative_to_unit": intensive_processors_unit,
                 "full_name": taxon,
                 "name": taxon[-1]  # Last level MAY BE defining the processor
                 }
            procs[taxon] = p
        # Third, the fund or flow columns
        # Flows in the registry of flows for this type of processor
        ff_name = curr[cols["FF_Name"]]
        if ff_name in ffs:
            ff_type = ffs[ff_name]
        else:
            ff_type = create_dictionary()
            ffs[ff_name] = ff_type

        # Flows and funds in a processor
        if "ff" in p:
            ff = p["ff"]
        else:
            ff = create_dictionary()
            p["ff"] = ff

        # Prepare the FF type

        # If the columns VAR_L1, VAR_L2, VAR_L3 were defined, convert to FF_Type
        if "FF_Type" not in cols and "Var_L1" in cols and "Var_L2" in cols and "Var_L3" in cols:
            v = [sh_in.cell(row=0 + 1, column=cols["Var_L" + str(i)] + 1).value for i in range(1, 4)]
            int_ext = None
            if strcmp(v[0][:3], "Int"):
                int_ext = "Int_"
            elif strcmp(v[0][:3], "Ext"):
                int_ext = "Ext_"
            in_out = None
            if strcmp(v[1][:3], "Inp"):
                int_ext = "In_"
            elif strcmp(v[1][:3], "Out"):
                int_ext = "Out_"
            flow_fund = None
            if strcmp(v[2][:3], "Flow"):
                int_ext = "Flow"
            elif strcmp(v[2][:3], "Fund"):
                int_ext = "Fund"
            if int_ext and in_out and flow_fund:
                ff_type = int_ext + in_out + flow_fund
            else:
                # TODO Improve
                sh_out.cell(row=rr + 1, column=len(sh_in.max_column) + 1,
                            value="Variables defining a flow or fund type must be all defined correctly")
                error = True
        else:
            ff_type = curr[cols["FF_Type"]]

        if ff_type.lower() not in ["int_in_flow", "int_in_fund", "int_out_flow", "ext_in_flow", "ext_out_flow"]:
            # TODO Improve
            sh_out.cell(row=rr + 1, column=len(sh_in.max_column) + error_count + 1,
                        value="FF type not allowed ('Int_Out_Fund', 'Ext_In_Fund', 'Ext_Out_Fund' combinations do not apply)")
            error = True

        # Add the flow or fund
        if ff_name not in p["ff"]:
            if curr[cols[unit_col_name]] and curr[cols[unit_col_name]].strip() != "-":
                unit = curr[cols[unit_col_name]]
                if strcmp(unit, "m3"):
                    unit = "m**3"
                elif strcmp(unit, "m2"):
                    unit = "m**2"

                unit = unit if not intensive_processors_specified else unit + "/" + intensive_processors_unit
            else:
                unit = ""
            ff[ff_name] = {"name": ff_name,
                           "type": ff_type,
                           "value": curr[cols["FF_Value"]],
                           "unit": unit,
                           "scale": curr[cols["Scale"]] if "Scale" in cols else None,
                           "source": curr[cols["FF_Source"]] if "FF_Source" in cols else None,
                           "comment": curr[cols["FF_Comment"]] if "FF_Comment" in cols else None
                           }
        else:
            sh_out.cell(row=rr + 1, column=len(sh_in.max_column) + 1,
                        value="'" + ff_name + "' in processor '"+str(taxon)+"' is repeated. Row skipped.")
            error = True

        return error

    # Analyze column names, to know what property is in each
    first_taxonomic_rank = sh_in.title[11:]
    cols = create_dictionary()  # Name to Index map
    intensive_processors_specified = None  # True if intensive processors are being specified
    intensive_processors_unit = None  # Unit name of the fund that can realize a processor (intensive to extensive)
    error_count = 0
    some_known = False  # The first columns will be taxonomic ranks, with free names
    type_cols = []  # Indices of columns with type information
    type_cols_names = []
    taxonomic_ranks = create_dictionary()  # Dictionary of sets (but dictionaries are used to allow case insensitivity if desired)
    # TODO If "FF_Type" is found, Var_L1 and the other are ignored, show a warning
    for c in range(sh_in.max_column):
        cell = sh_in.cell(row=0 + 1, column=c + 1)
        col_name = cell.value
        if not col_name:
            continue
        col_name = col_name.replace(" ", "")
        if strcmp(col_name, "FF_Type") or strcmp(col_name, "Var_Type"):
            col_name = "FF_Type"
            some_known = True
        elif strcmp(col_name, "Var_L1"):
            col_name = "Var_L1"
            some_known = True
        elif strcmp(col_name, "Var_L2"):
            col_name = "Var_L2"
            some_known = True
        elif strcmp(col_name, "Var_L3"):
            col_name = "Var_L3"
            some_known = True
        elif strcmp(col_name, "FF_Name") or strcmp(col_name, "Var"):
            col_name = "FF_Name"
            some_known = True
        elif strcmp(col_name, "FF_Value") or strcmp(col_name, "Value"):
            col_name = "FF_Value"
            some_known = True
        elif strcmp(col_name[:8], "FF_Unit_") or strcmp(col_name[:5], "Unit_") or \
             strcmp(col_name[:8], "FF_Unit/") or strcmp(col_name[:5], "Unit/"):
            col_name = "FF_Unit_" + (col_name[8:] if strcmp(col_name[:1], "F") else col_name[5:])
            unit_col_name = col_name  # Store specially
            intensive_processors_specified = True
            intensive_processors_unit = col_name[8:]
            some_known = True
            # TODO Consider the link of the unit to the kind of fund, for instance "Land Use"
        elif strcmp(col_name[:8], "FF_Unit") or strcmp(col_name[:5], "Unit"):
            col_name = "FF_Unit"
            unit_col_name = col_name
            intensive_processors_specified = False
            some_known = True
        elif strcmp(col_name, "Scale") or strcmp(col_name, "Data_Scale"):
            col_name = "Scale"
            some_known = True
        elif strcmp(col_name, "FF_Source") or strcmp(col_name, "Source") or strcmp(col_name, "Data_Source"):
            col_name = "FF_Source"
            some_known = True
        elif strcmp(col_name[:10], "FF_Comment") or strcmp(col_name[:7], "Comment"):
            col_name = "FF_Comment"
            some_known = True
        elif strcmp(col_name, ""):
            continue  # Just ignore it
        else:  # Unknown column name
            if some_known:  # If some known column was processed previously --> warning (the column will be ignored)
                if not col_name.endswith("(not recognized)"):  # Avoid concatenating
                    col_name += " (not recognized)"
            else:
                # Assume a taxonomic rank is being defined in this column. Store it as a part of the type
                taxonomic_ranks[col_name] = create_dictionary()
                type_cols.append(c)
                type_cols_names.append(col_name)

        if col_name in cols:
            # TODO Error sh_out.write(1 + error_count + 1, 0 + 1, "'" + col_name + "' column can appear only one time")
            error_count += 1

        cols[col_name] = c
        # TODO Before: sh_out.write(0, c, col_name)

    if "FF_Type" not in cols and "Var_L1" in cols and "Var_L2" in cols and "Var_L3" in cols:
        cols["FF_Type"] = -1  # Just to pass the following test

    # There should exist at least one column indicating the type of the processor
    if len(type_cols) == 0:
        print("ERROR: no column identifying the processor type or name was found")

    # Check presence of mandatory columns
    mandatory = ["FF_Type", "FF_Name", "FF_Value", unit_col_name]
    for m in mandatory:
        if m not in cols:
            # TODO Error: sh_out.cell(1+error_count+1, 0+1, "There is no '" + m + "' column")
            error_count += 1
    if error_count > 0:
        # TODO Error: sh_out.cell(2+error_count+1, 0+1, "The processing has been stopped")
        return

    # Dictionary used to implement a shortcut to avoid typing values in a column: if the value of the current
    # row is equal to the value in the previous row, it is not needed to specify it.
    previous_line = {}
    current_line = {}

    proc_taxonomy = create_dictionary()  # Register processors taxonomy. All the different types
    procs = create_dictionary()  # Register processor instances (intensive or extensive)
    ff_taxonomy = create_dictionary()  # Register all flow fund types
    ffs = create_dictionary()

    some_error = False
    # Read each line, which will be both a processor and a flow/fund
    rr = 1
    for r in range(1, sh_in.max_row):
        error_count = 0 + 1  # 1 is for an offset, errors are located in columns AFTER the information
        # Read values of current line
        for c in range(sh_in.max_column):
            cell = sh_in.cell(row=r + 1, column=c + 1)
            v = cell.value
            current_line[c] = v
            previous_line[c] = v  # Overwrite last value for the column "c"

        # PROCESS ROW
        # If it is a special "#" row, expand, looping through the origin
        # Else, just pass the value
        first_hash = True
        source_ds = None
        col2dim = dict()
        for c in range(len(current_line)):
            if isinstance(current_line[c], str) and current_line[c].startswith("#"):
                if first_hash:
                    first_hash = False
                    source_ds, dim = current_line[c][1:].lower().split(".")
                else:
                    dim = current_line[c][1:].lower()
                col2dim[dim] = c
        if not first_hash and source_ds:  # Process a dataset
            if "value" not in col2dim:
                # TODO Show error because no VALUE column was specified
                pass
            else:
                # Find dataset
                if source_ds in dfs:
                    df = dfs[source_ds]
                    # TODO Find type: flat dataset or PivotTable
                    pivot_table = True
                    # Find all the dimensions
                    dims = []
                    not_found_dims = set([k.lower() for k in col2dim])
                    for i, c in enumerate(df.columns.names):
                        if c and c.lower() in not_found_dims:
                            not_found_dims.remove(c.lower())
                            dims.append((c.lower(), "col", i))
                    if isinstance(df.index, pd.MultiIndex):
                        for i, c in enumerate(df.index.names):
                            if c and c.lower() in not_found_dims:
                                not_found_dims.remove(c.lower())
                                dims.append((c.lower(), "row", i))
                    else:
                        if df.index.name.lower() in not_found_dims:
                            not_found_dims.remove(df.index.name.lower())
                            dims.append((df.index.name.lower(), "row", 0))

                    # Value
                    if "value" in not_found_dims:
                        dims.append(("value", "", 0))
                        not_found_dims.remove("value")

                    if len(not_found_dims) > 0:
                        # TODO Error. One or more dimensions were not found
                        pass
                    else:
                        # TODO Find value. If it is organized in facts form, find value column. If not, all cells are "value"
                        # TODO in this case, only one aggregator is allowed.
                        if pivot_table:
                            # Copy curr vector
                            c_line = current_line.copy()
                            # iterate through all rows then columns
                            for r in range(df.shape[0]):
                                for c in range(df.shape[1]):
                                    # Modify the original at dimension indices, with the
                                    # dimension realizations for each cell (to process uniformly,
                                    # "value" is considered another item inside the dims vector)
                                    for d in dims:
                                        if d[1] == "col":
                                             v = df.columns.values[c][d[2]]
                                        elif d[1] == "row":
                                            if isinstance(df.index, pd.MultiIndex):
                                                v = df.index.values[r][d[2]]
                                            else:
                                                v = df.index[d[2]]
                                        else:
                                            # Get the value
                                            v = df.iloc[r, c]
                                        # Prepare the curr vector. Modify the original at dimension indices
                                        c_line[col2dim[d[0]]] = v

                                    # Issue a new row ------------
                                    error = process_row(c_line, rr)
                                    if error:
                                        some_error = True
                                        error_count += 1
                                    rr += 1

                        else:
                            # TODO iterate through rows
                            # TODO Find dimension realizations
                            # TODO Get the value
                            # TODO prepare the curr vector. Modify the original at dimension indices
                            pass
                else:
                    # TODO Show error in the row where the
                    pass
        else:
            error = process_row(current_line, rr)
            if error:
                some_error = True
                error_count += 1

        rr += 1
        # taxon = tuple([first_taxonomic_rank] + [cell_content_to_str(current_line[c]) for c in type_cols if current_line[c] != ""])
        # if taxon not in proc_taxonomy:
        #     proc_taxonomy[taxon] = None
        # # Add subtypes to their corresponding taxonomic ranks
        # for i, c in enumerate(type_cols):
        #     # Store a new entry (current_line[c]) in the taxonomic rank (type_cols_names[i])
        #     if cell_content_to_str(current_line[c]) not in taxonomic_ranks[type_cols_names[i]]:
        #         taxonomic_ranks[type_cols_names[i]][cell_content_to_str(current_line[c])] = None
        #
        # # Second, the processor, which includes the types and the name
        # if taxon in procs:
        #     p = procs[taxon]
        # else:
        #     p = {"intensive": intensive_processors_specified,
        #          "relative_to_unit": intensive_processors_unit,
        #          "full_name": taxon,
        #          "name": taxon[-1]  # Last level MAY BE defining the processor
        #          }
        #     procs[taxon] = p
        # # Third, the fund or flow columns
        # # Flows in the registry of flows for this type of processor
        # ff_name = current_line[cols["FF_Name"]]
        # if ff_name in ffs:
        #     ff_type = ffs[ff_name]
        # else:
        #     ff_type = create_dictionary()
        #     ffs[ff_name] = ff_type
        #
        # # Flows and funds in a processor
        # if "ff" in p:
        #     ff = p["ff"]
        # else:
        #     ff = create_dictionary()
        #     p["ff"] = ff
        #
        # # Prepare the FF type
        #
        # # If the columns VAR_L1, VAR_L2, VAR_L3 were defined, convert to FF_Type
        # if "FF_Type" not in cols and "Var_L1" in cols and "Var_L2" in cols and "Var_L3" in cols:
        #     v = [sh_in.cell(row=0 + 1, column=cols["Var_L" + str(i)] + 1).value for i in range(1, 4)]
        #     int_ext = None
        #     if strcmp(v[0][:3], "Int"):
        #         int_ext = "Int_"
        #     elif strcmp(v[0][:3], "Ext"):
        #         int_ext = "Ext_"
        #     in_out = None
        #     if strcmp(v[1][:3], "Inp"):
        #         int_ext = "In_"
        #     elif strcmp(v[1][:3], "Out"):
        #         int_ext = "Out_"
        #     flow_fund = None
        #     if strcmp(v[2][:3], "Flow"):
        #         int_ext = "Flow"
        #     elif strcmp(v[2][:3], "Fund"):
        #         int_ext = "Fund"
        #     if int_ext and in_out and flow_fund:
        #         ff_type = int_ext + in_out + flow_fund
        #     else:
        #         # TODO Improve
        #         sh_out.cell(row=r + 1, column=len(sh_in.max_column) + error_count + 1,
        #                     value="Variables defining a flow or fund type must be all defined correctly")
        #         error_count += 1
        #         some_error = True
        # else:
        #     ff_type = current_line[cols["FF_Type"]]
        #
        # if ff_type not in ["Int_In_Flow", "Int_In_Fund", "Int_Out_Flow", "Ext_In_Flow", "Ext_Out_Flow"]:
        #     # TODO Improve
        #     sh_out.cell(row=r + 1, column=len(sh_in.ncols) + error_count + 1,
        #                 value="FF type not allowed ('Int_Out_Fund', 'Ext_In_Fund', 'Ext_Out_Fund' combinations do not apply)")
        #     error_count += 1
        #     some_error = True
        #
        # # Add the flow or fund
        # if ff_name not in p["ff"]:
        #     if current_line[cols[unit_col_name]] and current_line[cols[unit_col_name]].strip() != "-":
        #         unit = current_line[cols[unit_col_name]]
        #         if strcmp(unit, "m3"):
        #             unit = "m**3"
        #         elif strcmp(unit, "m2"):
        #             unit = "m**2"
        #
        #         unit = unit if not intensive_processors_specified else unit + "/" + intensive_processors_unit
        #     else:
        #         unit = ""
        #     ff[ff_name] = {"name": ff_name,
        #                    "type": ff_type,
        #                    "value": current_line[cols["FF_Value"]],
        #                    "unit": unit,
        #                    "scale": current_line[cols["Scale"]],
        #                    "source": current_line[cols["FF_Source"]],
        #                    "comment": current_line[cols["FF_Comment"]]
        #                    }
        # else:
        #     sh_out.cell(row=r + 1, column=len(sh_in.ncols) + error_count + 1,
        #                 value="'" + ff_name + "' in processor '"+str(taxon)+"' is repeated. Row skipped.")
        #     error_count += 1
        #     some_error = True

    # Loop finished!

    # Last step is to register the results for following operations, if no error occurred
    if not some_error:
        registry[first_taxonomic_rank] = {"processors_taxonomy": proc_taxonomy,  # All the taxonomic types of processors, except the first level
                                          "processors": procs,  # The processors, also indexed by the full taxonomic types
                                          "taxonomic_rank_levels": type_cols_names,  # Names of the taxonomic ranks
                                          "taxonomic_ranks": taxonomic_ranks  # Names contained in each of the taxonomic ranks
                                          }
        return first_taxonomic_rank
    else:
        return None

# -----------------------------------------------
# COMBINE
# -----------------------------------------------


def convert_intensive_processor_to_extensive(p, fund):
    """
    Convert the current processor from intensive to extensive
    The conversion is assumed (even if no flow or fund need the scaling, the processor will be extensive after this call)
    Also apply the conversion to descendants, recursively (children are converted first)

    :param p:
    :param fund: Fund, whose value is the total fund to be applied. The processor will contain scalings, so it
    will be multiplied by this factor
    :return:
    """

    if "children" in p:
        for c in p["children"]:
            convert_intensive_processor_to_extensive(c, fund)

    if p["intensive"]:
        fund["unit_intensive"] = "-"
        fund["value_intensive"] = "1"
        base_unit = 1 / backend.ureg(fund["unit"])
        t = base_unit.dimensionality
        # TODO Is the fund already there? WARNING, because it should be
        # Multiply funds and flows which have the same unit in the denominator
        # print(p["full_name"])
        ff = p["ff"]
        for k in ff:
            # Analyze units of the current flow / fund
            un = ff[k]["unit"]
            dims = backend.ureg(ff[k]["unit"]).to_tuple()
            found = False
            for d in dims[1]:
                v = backend.ureg(d[0])**d[1]
                if v.dimensionality == t:
                    found = True
                    break
            if found:
                # Do a dimensional multiplication of the factors
                v = ff[k]["value"] * backend.ureg(ff[k]["unit"]) * fund["value"] * backend.ureg(fund["unit"])
                # Format using short unit names, then split in two parts (magnitude and unit)
                s = "{:~P}".format(v).split(" ", 1)
                ff[k]["value_intensive"] = ff[k]["value"]
                ff[k]["unit_intensive"] = ff[k]["unit"]
                ff[k]["value"] = float(s[0])
                ff[k]["unit"] = s[1]  # Unit is changed
                # print(un + " -> " + s[1])
            else:
                # print(un + "(not converted)")
                pass

        # Clone the fund, scale it using the accumulated scaling factors, then inject it into the processor
        f2 = copy.deepcopy(fund)
        f2["value"] *= eval(p["factors"])
        ff[f2["name"]] = f2
        # Mark the processor as "extensive"
        p["intensive"] = False
        p["relative_to_unit"] = None


def list_processors_set(registry, entry, processor_type):
    """
    Traverse the entry and obtain a list of flows and funds, accompanied with the preceding typologies, showing for each
    flow and/or fund the hierarchy needed to get to the processor and the processor itself

    :param registry: The registry of processor layers
    :param entry: Which registry entry specifically to read
    :param processor_type: A processor type (the registry will be analyzed to recover related taxonomic ranks)
    :return: pd.DataFrame containing the desired structure
    """

    def add():
        nonlocal idx, reg, lst, distinct_fields
        t = []
        for l in lst:
            for m2 in l:
                if m2[0] not in distinct_fields:
                    distinct_fields[m2[0]] = idx
                    idx += 1
                t.append((m2[0], m2[1]))
        reg.append(t)

    def add_itself_and_children(p):
        nonlocal m, lst
        tr = m[p["full_name"][0]]
        tmp = []  # ("ProcessorType", p["full_name"][0])
        tmp.extend([(st, sti) for st, sti in zip(tr, p["full_name"][1:])])
        if "factors" in p:
            tmp.append((p["full_name"][0] + "_factors", p["factors"]))
        lst.append(tmp)  # Add a pack of fields for the level
        if p["full_name"][0] != processor_type:
            if "children" in p:
                for p2 in p["children"]:
                    add_itself_and_children(p2)
        else:
            lst.append([("ProcessorType", p["full_name"][0])])
            for k in p["ff"]:
                lst.append([(l, p["ff"][k][l]) for l in p["ff"][k] if l in
                            ["name", "unit", "value", "type", "value_intensive", "unit_intensive"]])
                # Add to reg
                add()
                del lst[-1]
            del lst[-1]
        del lst[-1]  # Remove fields for the level

    # Find taxonomic rank names for the different processors types. Elaborate a map of lists
    distinct_fields = dict()
    idx = 0
    m = {}
    lst = []
    reg = []
    for k in registry:
        m[k] = registry[k]["taxonomic_rank_levels"]
    for p in entry["processors"]:
        add_itself_and_children(entry["processors"][p])
    # List of fields
    fn = np.zeros((len(distinct_fields))).astype(object)
    for k in distinct_fields:
        fn[distinct_fields[k]] = k
    # Convert reg to Dataframe
    d = np.zeros((len(reg), len(distinct_fields))).astype(object)
    for r, lst in enumerate(reg):
        for lst2 in lst:
            d[r, distinct_fields[lst2[0]]] = lst2[1]
    # Create a Dataframe
    df = pd.DataFrame(d, columns=fn)
    # Convert "Value" column (the last column) to numeric
    df['value'] = df['value'].apply(lambda x: pd.to_numeric(x, errors='coerce'))

    return df


def add_child_to(contained, container, factor):
    """
    Add contained to container, as CHILD ("children" list or dict)
    CONTAINER inherits all flows and funds. The accounted value is reset to zero when a new flow are fund are created
    The linked values need to be updated on request, so when two FF's (from container and contained) are linked, the value
    at the container will be updated automatically.

    :param contained:
    :param container:
    :param factor:
    :return:
    """
    if "children" in container:
        ch = container["children"]
    else:
        ch = []
        container["children"] = ch

    ch.append(contained)
    # Cascade "factor" to contained
    adimensionally_scale_processor(contained, factor)
    # Cascade "extensive" maker to contained with "intensive" nature
    if not container["intensive"] and contained["intensive"]:
        # Obtain the right number
        unit = contained["relative_to_unit"]
        fund = None
        for ff in container["ff"]:
            if container["ff"][ff]["unit"].lower() == unit.lower():
                fund = container["ff"][ff]
                break
        if fund:
            # Convert the contained processor to extensive (the fund is injected also). Also contained children
            convert_intensive_processor_to_extensive(contained, fund)

    # Inherit all flows and funds from contained
    for ff in contained["ff"]:
        if contained["ff"][ff]["name"] not in container["ff"]:
            ff_prime = copy.deepcopy(contained["ff"][ff])
            ff_prime["value"] = 0.0
            container["ff"][ff] = ff_prime
        else:
            ff_prime = container["ff"][ff]

        ff_prime["value"] += contained["ff"][ff]["value"]

"""
            ff[ff_name] = {"name": ff_name,
                           "type": ff_type,
                           "value": current_line[cols["FF_Value"]],
                           "unit": current_line[cols["FF_Unit"]] if not intensive_processors_specified
                              else current_line[cols["FF_Unit"]] + "/" + intensive_processors_unit,
                           "scale": current_line[cols["Scale"]],
                           "source": current_line[cols["FF_Source"]],
                           "comment": current_line[cols["FF_Comment"]]
                           }
"""


def combine_two_layers_of_processors(sh_in, sh_out, registry):
    """
    Elaborate a new set of processors based on the combination of two existing sets of processors
    A table specifying which processor sets are combined is passed as input (sh_in)
    A set of processors is created

    :param sh_in:
    :param sh_out: Just to show errors and warnings. No content is modified
    :param registry:
    :return:
    """
    m = binary_mask_from_worksheet(sh_in, True)  # True for cells containing numbers
    # Locate the matrix with numbers. Assume this defines the labels to consider, they will be around the matrix
    t = obtain_rectangular_submatrices(m)
    t = t[0]  # Take just the first element
    # print(t[0])
    v = worksheet_to_numpy_array(sh_in)
    for t in [(t[0], t[1], t[2], t[3]), (t[0] + 1, t[1], t[2], t[3]), (t[0], t[1], t[2] + 1, t[3])]:
        f = v[t[0]:t[1], t[2]:t[3]].astype(np.float64)
        row_sum = np.sum(f, axis=1)  # A column vector. If "all ones", the container will be along rows
        col_sum = np.sum(f, axis=0)  # A row vector. If "all ones", the container will be along columns
        container_situation = None
        if np.allclose(row_sum, 1, 1e-2):
            container_situation = "in_rows"
        if np.allclose(col_sum, 1, 1e-2):
            if container_situation:
                show_message(sh_out, t[0] - 1, t[1] - 1,
                             "ERROR: both rows and columns should not sum to ones")
            container_situation = "in_columns"
        if container_situation:
            break
    if not container_situation:
        show_message(sh_out, t[0] - 1, t[1] - 1,
                     "Warning: neither the sum of rows nor of columns is summing to ones",
                     type="warning")

    # Read cell containing the specification of the processors.
    # The FIRST is the CONTAINED (child), the SECOND the CONTAINER (parent).
    # They are specified separated by "/" or "-".
    # Iterate to find both in the registry, obtain "registry_entries"
    RegistryEntry = collections.namedtuple('RegistryEntry', 'name registry_name registry_entry')
    registry_entries = []
    for s in re.split("\/|\-|_", sh_in.title[8:]):  # v[t[0]-1, t[2]-1]
        if "Upscaled_"+s.strip() in registry:
            registry_entries.append(RegistryEntry(s.strip(), "Upscaled_"+s.strip(), registry["Upscaled_"+s.strip()]))
        elif s.strip() in registry:
            registry_entries.append(RegistryEntry(s.strip(), s.strip(), registry[s.strip()]))
    if len(registry_entries) != 2:
        show_message(sh_out, 0, 0, "ERROR, there should be two registry entries recognized. Either the registry "
                                   "does not contain these processors or they are written incorrectly")

    # Identify taxonomic ranks. These ranks identify the processors
    # Concatenate all the taxonomic_ranks
    t_ranks = [(k, v) for k, v in registry_entries[0].registry_entry["taxonomic_ranks"].items()] + \
              [(k, v) for k, v in registry_entries[1].registry_entry["taxonomic_ranks"].items()]
    n_ranks_first_processor = len(registry_entries[0].registry_entry["taxonomic_rank_levels"])
    # Find the rows and cols containing taxon specifications
    lst = []
    for rup in range(t[0] - 1, -1, -1):
        # Take a slice of labels. Using "set" avoid repetitions. Elaborate a tuple
        lst.append((set(v[rup, t[2]:t[3]]), "row", rup))
    for cleft in range(t[2] - 1, -1, -1):
        # Take a slice of labels. Using "set" avoid repetitions. Elaborate a tuple
        lst.append((set(v[t[0]:t[1], cleft]), "col", cleft))
    col_idx = []
    row_idx = []
    for l in lst:
        found = False
        for e in l[0]:
            for t_rank in t_ranks:
                if e and cell_content_to_str(e) in t_rank[1]:
                    found = True
                    break
            if found:
                break
        if found:
            if l[1] == "col":
                col_idx.append(l[2])
            else:
                row_idx.append(l[2])

    # Sweep the matrix of numbers
    warned_not_founds = set()
    new_procs = create_dictionary()  # Newly obtained container processors
    for r in range(t[0], t[1]):
        for c in range(t[2], t[3]):
            if v[r, c] < 1e-3:
                continue

            # Clear taxon identification for each processor
            taxa0 = {}
            taxa1 = {}
            # Find the taxa, for each processor type involved
            for rup in row_idx:
                e = cell_content_to_str(v[rup, c])
                found = False
                for i, t_rank in enumerate(t_ranks):
                    if e and cell_content_to_str(e) in t_rank[1]:
                        if i < n_ranks_first_processor:
                            idx = i
                            taxa = taxa0
                            entry = 0
                        else:
                            idx = i - n_ranks_first_processor
                            taxa = taxa1
                            entry = 1
                        found = True
                        # Add taxon to the processor type
                        taxa[registry_entries[entry].registry_entry["taxonomic_rank_levels"][idx]] = e
                if not found:
                    cell_pos = (rup + 1, c + 1)
                    print(sh_in.title + " ("+str(cell_pos[0])+", "+str(cell_pos[1])+") not found")
                    if e:
                        if cell_pos not in warned_not_founds:
                            warned_not_founds.add(cell_pos)
                            show_message(sh_out, cell_pos[0], cell_pos[1], "WARNING: subtype '"+e+"' not found. Processors to be combined cannot be identified properly.", "warning")

            for cleft in col_idx:
                e = cell_content_to_str(v[r, cleft])
                found = False
                for i, t_rank in enumerate(t_ranks):
                    if e and e in t_rank[1]:
                        if i < n_ranks_first_processor:
                            idx = i
                            taxa = taxa0
                            entry = 0
                        else:
                            idx = i - n_ranks_first_processor
                            taxa = taxa1
                            entry = 1
                        found = True
                        # Add taxon to the processor type
                        taxa[registry_entries[entry].registry_entry["taxonomic_rank_levels"][idx]] = e
                if not found:
                    cell_pos = (r + 1, cleft + 1)
                    print(sh_in.title + " ("+str(cell_pos[0])+", "+str(cell_pos[1])+") not found")
                    if e:
                        if cell_pos not in warned_not_founds:
                            warned_not_founds.add(cell_pos)
                            show_message(sh_out, cell_pos[0], cell_pos[1], "WARNING: subtype '"+e+"' not found. Processors to be combined cannot be identified properly.", "warning")
            # Identify processors
            taxa = [taxa0, taxa1]
            processors = []
            for entry in [0, 1]:
                try:
                    p = registry_entries[entry].registry_entry["processors"][tuple([registry_entries[entry].name]+[taxa[entry][rn] for rn in registry_entries[entry].registry_entry["taxonomic_rank_levels"]])]
                    processors.append(p)
                except:  # Not found. The combination may be referring to a processor which has not been detailed previously
                    pass

            if len(processors) != 2:
                show_message(sh_out, r + 1, c + 1, "WARNING: did not find two processors to combine", "warning")
                continue

            S = processors[0]
            T = processors[1]
            if T["full_name"] in new_procs:
                T2 = new_procs[T["full_name"]]
            else:
                T2 = clone_processor(T)  # Clone processor and its children
                new_procs[T["full_name"]] = T2
            S2 = clone_processor(S)  # Clone processor and its children
            add_child_to(contained=S2, container=T2, factor=v[r, c])

    container = registry_entries[1]
    registry["Upscaled_"+container.name] = \
        {"processors_taxonomy": container.registry_entry["processors_taxonomy"],  # All the taxonomic types of processors, except the first level
         "processors": new_procs,  # The processors, also indexed by the full taxonomic types
         "taxonomic_rank_levels": container.registry_entry["taxonomic_rank_levels"],  # Names of the taxonomic ranks
         "taxonomic_ranks": container.registry_entry["taxonomic_ranks"]  # Names contained in each of the taxonomic ranks
         }


def pivot_table(sh_writable, df, rows, cols, aggs, values, show_totals):
    """
    Given Pivot Table parameters, elaborate it and put it in the output worksheet

    :param sh_writable:
    :param df:
    :param rows:
    :param cols:
    :param aggs:
    :param values:
    :param show_totals:
    :return: Nothing. Results are written into the worksheet "sh_writable"
    """
    df2 = None
    if len(rows) > 0 and len(cols) > 0 and values:
        if df.shape[0] == 0:
            show_message(sh_writable, 1, 1, "Warning: no data from the specified query", type="warning")
        else:
            df2 = pd.pivot_table(df,
                                 values=values,
                                 index=rows,
                                 columns=cols,
                                 aggfunc=aggs, fill_value=np.NaN, margins=show_totals,
                                 dropna=True, margins_name="Total")

            if df2.shape[0] * df2.shape[1] > 40000:
                df2 = None
                show_message(sh_writable, 1, 1, "ERROR: the resulting pivot table cannot be shown because there are "
                                                "too many cells (" + str(df2.shape[0]*df2.shape[1]) +
                                                "). Please, reconsider the parameters.")
            else:
                # df2 = df2.swaplevel(0, len(df2.columns.names)-1, axis=1)
                i = 0
                # Number of levels of the columns Index
                nidx_rows = len(df2.columns.names)
                # Count the number of levels of the rows Index
                if isinstance(df2.index, pd.MultiIndex):
                    nidx_cols = len(df2.index.names)
                else:
                    nidx_cols = 1
                # Reference point
                start = (nidx_rows, nidx_cols)
                # Reset worksheet
                sh_out = reset_worksheet(sh_writable)
                # Put the values at (nr+1, nc+1)
                for r in range(df2.shape[0]):
                    for c in range(df2.shape[1]):
                        sh_out.cell(row=r + start[0] + 1, column=c + start[1] + 1, value=df2.iloc[r, c])
                # Put the columns Index from (:, nc+1) on
                for c in range(df2.shape[1]):
                    for r, l in enumerate(df2.columns.values[c]):
                        sh_out.cell(row=r + 1, column=c + start[1] + 1, value=str(l))
                # Put the rows Index from (nr+1, :) on
                for r in range(df2.shape[0]):
                    if isinstance(df2.index, pd.MultiIndex):
                        for c, l in enumerate(df2.index.values[r]):
                            sh_out.cell(row=r + start[0] + 1, column=c + 1, value=str(l))
                    else:
                        sh_out.cell(row=r + start[0] + 1, column=0 + 1, value=str(df2.index[r]))
                        # TODO Check consecutive values, to merge, in horizontal or in vertical
                # Mark the Pivot Table
                sh_out.cell(row=1, column=1, value="Pivot table")
    return df2


def read_and_calculate_pivot_table(sh, sh_writable, registry, dfs, df=None):
    # Read parameters
    row_c = None
    col_c = None
    agg_c = None
    sh_c = None
    rows = []
    cols = []
    sh_name = None
    some_error = False
    for c in range(sh.max_column):
        cname = sh.cell(row=0 + 1, column=c + 1).value
        if not cname:
            continue
        if strcmp(cname, "Sheet") and df is None:
            sh_c = c
            # Sheet from where to read the input dataframe
            sh_name = sh.cell(row=1 + 1, column=c + 1).value
            if sh_name in dfs:
                df = dfs[sh_name]
            else:
                df = list_processors(sh_name[5:], sh_writable.cell(row=1 + 1, column=sh_c + 1), registry, dfs)
                if df is None:
                    some_error = True
                    show_message(sh_writable, 1 + 1, sh_c + 1,
                                "ERROR: could not find or elaborate '" + sh_name + "'")
                # TODO Check if "cname" corresponds to a worksheet (read sh.sheets)
                # If it exists, it may be a worksheet not generating
        elif strcmp(cname, "Rows") or strcmp(cname, "Columns"):
            if strcmp(cname, "Rows"):
                row_c = c
                lst = rows
            else:
                col_c = c
                lst = cols
            r = 1
            while r < sh.max_row and sh.cell(row=r + 1, column=c + 1).value:
                # Check that the row exists in the input data
                v = sh.cell(row=r + 1, column=c + 1).value
                if v.lower() in ("time", "period", "date", "dates", "time_interval", "year"):  # For datasets (not for processors, which do not time)
                    v = "TIME_PERIOD"
                if not strcmp(v, "value"):
                    if df is not None:
                        found = False
                        for d in df.columns:
                            if d.lower() == v.lower():
                                found = True
                                v = d
                                lst.append(v)
                                break
                        if not found:
                            some_error = True
                            show_message(sh_writable, r + 1, c + 1, "ERROR: dimension not found")
                r += 1
        elif strcmp(cname, "AggFunc"):
            agg_c = c
            aggs = []
            r = 1
            while r < sh.max_row and sh.cell(row=r + 1, column=c + 1).value:
                # Check that the aggregator function exists
                fname = sh.cell(row=r + 1, column=c + 1).value
                if strcmp(fname, "sum"):
                    aggs.append(np.sum)
                elif strcmp(fname, "mean") or strcmp(fname, "average") or strcmp(fname, "avg"):
                    aggs.append(np.mean)
                elif strcmp(fname, "std") or strcmp(fname, "stddev") or strcmp(fname, "standarddeviation"):
                    aggs.append(np.std)
                elif strcmp(fname, "max") or strcmp(fname, "maximum"):
                    aggs.append(np.amax)
                elif strcmp(fname, "min") or strcmp(fname, "minimum"):
                    aggs.append(np.amin)
                else:
                    some_error = True
                    show_message(sh_writable, r + 1, c + 1, "ERROR: aggregation function not found. Please specify one or more of 'sum', 'mean' (or 'average', 'avg'), 'std' (or 'stddev'), 'max' (or 'maximum'), 'min' (or 'minimum')")
                r += 1
        elif strcmp(cname, "ShowTotals"):
            show_totals = False
            r = 1
            while r < sh.max_row and sh.cell(row=r + 1, column=c + 1).value:
                # TODO Check that the column exists in the input data
                v = sh.cell(row=r + 1, column=c + 1).value
                if isinstance(v, int):
                    cnv = int(v)
                    show_totals = cnv == 1
                elif isinstance(v, str):
                    show_totals = strcmp(v, "true")
                elif isinstance(v, bool):
                    show_totals = True if v else False
                else:
                    some_error = True
                    show_message(sh_writable, r + 1, c + 1, "ERROR: value not recognized as boolean")
                r += 1

    if df is None:
        show_message(sh_writable, 1, 1, "ERROR: no valid input data for Pivot Table")

    # Values
    for cn in df.columns:
        if strcmp(cn, "value"):
            values = cn
            break

    if not values:
        show_message(sh_writable, 1 + 1, sh_c + 1,
                     "ERROR: could not find a values column in '" + sh_name + "' worksheet, for input to a PivotTable")

    elif len(rows) == 0:
        show_message(sh_writable, 0 + 1, row_c + 1, "ERROR: at least one row is needed")
    elif len(cols) == 0:
        show_message(sh_writable, 0 + 1, col_c + 1, "ERROR: at least one column is needed")
    elif len(aggs) == 0:
        show_message(sh_writable, 0 + 1, agg_c + 1, "ERROR: at least one aggregation function is needed")
    elif not some_error:
        df2 = pivot_table(sh_writable, df, rows, cols, aggs, values, show_totals)
        if sh_writable.title.lower().startswith("dataset_eurostat_"):
            dfs["PivotTable_"+sh_writable.title[17:]] = df2
        elif sh_writable.title.lower().startswith("dataset_ssp_"):
            dfs["PivotTable_"+sh_writable.title[8:]] = df2


def list_processors(name, cell, registry, dfs):
    if name not in dfs:
        s = "upscaled_" + name
        processors = None  # What layer of processors to enumerate
        for k in registry:
            if s.lower().find(k.lower()) == 0:
                processors = registry[k]
                if s.lower().find(k.lower() + "s") == 0:
                    s = s[len(k) + 2:]
                else:
                    s = s[len(k) + 1:]
                break
        bottom_level = None
        for k in registry:
            if s.lower() == k.lower():
                bottom_level = k
                break
            elif s.endswith("s") and s[:-1].lower() == k.lower():  # Allow plural termination in "s" ("es" is not considered...)
                bottom_level = k
                break

        if not processors:
            msg = "ERROR: could not find Top containing processor level in '" + name + \
                  ". May be there is some misspelling in the processor names"
            print(msg)
            show_message(cell.parent, cell.row, cell.col_idx, msg)

        if not bottom_level:
            msg = "ERROR: could not find requested bottom level in '" + name + \
                  ". May be there is some misspelling in the processor names"
            print(msg)
            show_message(cell.parent, cell.row, cell.col_idx, msg)

        if not processors or not bottom_level:
            return None

        # Find the specified set of processors, and enumerate at the specified level
        df = list_processors_set(registry, processors, bottom_level)
        # Store the Dataframe for Pivot Tables
        dfs[name] = df

    return dfs[name]


def read_mapping_and_join(sh, sh_writable, metadatasets, dfs, maps, source_manager):
    """
    Read a mapping from a dataset dimension into a MuSIASEM taxonomy
    The mapping has to be MANY to ONE
    The mapping has to be complete (all elements from left side must be covered)
    The codes are intrinsically hierarchic 
    
    :param sh: 
    :param sh_writable: 
    :param metadatasets: Registry of Metadata on Datasets 
    :param dfs: Registry of Datasets
    :param maps: Registry of mappings
    :param source_manager: DataSourceManager
    :return: 
    """
    """
* Comprobar mapeo. Si quedan categorías sin mapear, avisar. El mapa debería cubrir el lado izquierdo. SI no, error. Si quedan categorías del lado derecho sin cubrir, avisar (no error)
* Ver si los nodos superiores del dataset contienen valores
    """
    # TODO Read elements below these two columns, with the constraint
    # TODO that values in the left side must be present, and values to the right should be
    v = cell_content_to_str(sh.cell(row=1, column=1).value)
    if v:
        # TODO Split, two parts: dataset name, dimension name
        try:
            dset_name, dim_name = v.split(".")
        except ValueError:
            show_message(sh_writable, 1, 1, "Error: the source code list must be <dataset_name>.<dimension_name>")
    taxon_name = cell_content_to_str(sh.cell(row=1, column=2).value)

    # Do we have the metadata for dset_name?
    if dset_name not in metadatasets:
        # Find source of the dataset, from the list of supported datasets
        lst = source_manager.get_datasets()
        ds = {t[1]: t[0] for t in lst}
        if dset_name in ds:
            source = ds[dset_name]
        else:
            if dset_name.startswith("ssp_"):
                # Obtain SSP
                source = "SSP"
            else:
                source = "Eurostat"
        _, metadata = get_statistical_dataset_structure(source, dset_name, None, source_manager)
        metadatasets[dset_name] = metadata
    else:
        metadata = metadatasets[dset_name]

    if not metadata:
        show_message(sh_writable, 1, 1, "Error: could not find metadata for dataset '" + dset_name + "'")
    else:
        # Obtain the source code list
        if dim_name not in metadata[0]:
            show_message(sh_writable, 1, 1, "Error: could not find dimension '"+dim_name+"' in dataset '"+dset_name+"'")
        else:
            # Obtain the source code list
            src_code_list = [c for c in metadata[0][dim_name].code_list]
            # Read the pairs and the destination code list
            # TODO the destination code list is currently declared in the table itself. In the future it could be declared outside
            dst_code_list = set()
            lst = []
            prev_dest = None
            for r in range(1, sh.max_row):
                orig = cell_content_to_str(sh.cell(row=r + 1, column=1).value)
                dest = cell_content_to_str(sh.cell(row=r + 1, column=2).value)
                if not dest:
                    dest = prev_dest
                else:
                    prev_dest = dest
                if orig and dest:
                    lst.append((orig, dest))
                    dst_code_list.add(dest)
                else:
                    show_message(sh_writable, r + 1, 1, "Warning: origin, destination or both codes are missing. Ignored", type="warning")

            # Parse
            mapped, unmapped = map_codelists(src_code_list, list(dst_code_list), lst)
            # TODO Complain if there are unmapped entries from the source
            # Store the map (it is a multidict: the same source dataset + dimension could be mapped several times)
            if dset_name+"."+dim_name in maps:
                d = maps[dset_name+"."+dim_name]
            else:
                d = create_dictionary()
                maps[dset_name+"."+dim_name] = d
            d[taxon_name] = (list(dst_code_list), mapped)

            # Try the join with existing dataset
            if dset_name in dfs:

                # Create a Dataframe with the mapping
                df_dst = pd.DataFrame(mapped, columns=['sou_rce', taxon_name])
                for di in dfs[dset_name].columns:
                    if strcmp(dim_name, di):
                        dim_name = di
                        break
                dfs[dset_name] = pd.merge(dfs[dset_name], df_dst, how='left', left_on=dim_name, right_on='sou_rce')
                del dfs[dset_name]['sou_rce']


def process_file(input_file):
    """
    Receives a binary with a Spreadsheet file, processes it and returns a new Spreadsheet file containing the results of the
    processing

    Command and variables are case insensitive

    :param input_file: Input Spreadsheet file as byte array ("bytes")
    :return: Output Spreadsheet file as byte array ("bytes")
    """
    the_registry.processors = create_dictionary()
    the_registry.datasets = create_dictionary()  # Dictionary of Dataframes, for PivotTable generation
    the_registry.metadatasets = create_dictionary()
    the_registry.maps = create_dictionary()

    # Construct DataSourceManager and the supported Data Sources

    if backend.data_engine:
        print("Antes de procesar. Existe DATA ENGINE")
    else:
        print("Antes de procesar. NO existe DATA ENGINE")

    dsm = DataSourceManager(session_factory=DBSession)
    dsm.register_datasource_manager(Eurostat())
    if 'FAO_DATASETS_DIR' in app.config:
        fao_dir = app.config['FAO_DATASETS_DIR']
    else:
        fao_dir = "/home/rnebot/DATOS/FAOSTAT/"
    dsm.register_datasource_manager(FAOSTAT(datasets_directory=fao_dir,
                                           metadata_session_factory=DBSession,
                                           data_engine=backend.data_engine))

    sources = dsm.get_supported_sources()

    re_enum = re.compile(r"^Dataset_\L<words>_Enumerate", flags=re.IGNORECASE, words=sources)
    re_meta = re.compile(r"^Metadata_\L<words>", flags=re.IGNORECASE, words=sources)
    re_data = re.compile(r"^Dataset_\L<words>", flags=re.IGNORECASE, words=sources)

    registry = None
    dfs = None
    n_combinations = 0  # Count number of combinations

    # Instance for reading, ignoring formulas reading numbers!
    xl_in = openpyxl.load_workbook(io.BytesIO(input_file), data_only=True)
    # This is also for reading, but does not read numbers when there are formulas
    xl_out = openpyxl.load_workbook(io.BytesIO(input_file))

    # Is it a case study to be stored or just backend_tests? If there is a "Metadata" worksheet -> store
    persist = False
    in_sheets = []
    for sh_name in xl_in.get_sheet_names():
        in_sheets.append(sh_name)
        sh = xl_in.get_sheet_by_name(sh_name)
        print(sh_name+": "+str(sh.max_row)+" x "+str(sh.max_column))
        if strcmp(sh_name, "Metadata"):
            persist = True

    # A case study
    cs = None

    # Process sheets in sequence

    # The commands
    # * From the metadata elaborate a "metadata.xml" in a single cell
    # * Download dataset
    # * Read processors of type ...
    # * Upscale combining two types of processors
    # * Downscale (apply fund to processor types to instantiate them)
    # * Pivot table (elaborate a pivot table using some of the tables. Some parameters needed, has to be thought)

    metadata_processed = False
    for sh_name in in_sheets:
        sh = xl_in.get_sheet_by_name(sh_name)
        sh_writable = xl_out.get_sheet_by_name(sh_name)
        print("Currently processing: "+sh_name)
        if strcmp(sh.title, "Metadata"):
            if metadata_processed:
                pass  # To copy, no action is needed
            else:
                # Process the metadata. Fill the new sheet. Prepare and return a CaseStudy object
                cs = process_metadata(sh)
                metadata_processed = True
        # MUSIASEM COMMANDS
        elif strcmp(sh.title[:11], "Processors_"):
            # TODO Read a set of Processors of the same type
            # There will be a list of Processor types (given by the sheet name)
            # Processors of a type can be totally instantiated or partially instantiated
            # Processors will be qualified by a tuple of types, hierarchically organized
            # Processors will not have
            name = read_processors(sh, sh_writable, the_registry.processors, the_registry.datasets)
            # Generate and Store the Dataframe for Pivot Tables
            df = list_processors_set(the_registry.processors, the_registry.processors[name], name)
            the_registry.datasets["Processors_"+name] = df  # Register with two names
            the_registry.datasets["List_"+name+"_"+name] = df
        elif strcmp(sh.title[:8], "Upscale_"):
            # TODO Look for a table linking two lists of Processors of different types
            # Doing this, each processor of the parent level will contain all the children processors, recursively
            # And the flows and funds are passed into the parent level also, combining or aggregating
            combine_two_layers_of_processors(sh, sh_writable, the_registry.processors)
            n_combinations += 1
        elif strcmp(sh.title[:5], "List_"):
            df = list_processors(sh.title[5:], sh_writable.cell(row=1, column=1), the_registry.processors, the_registry.datasets)
            if df is not None:
                # Reset worksheet
                sh_out = reset_worksheet(sh_writable)
                # Put the dataframe in a new worksheet
                for c, t in enumerate(df.columns):
                    sh_out.cell(row=0 + 1, column=c + 1, value=t)
                for r in range(df.shape[0]):
                    for c in range(df.shape[1]):
                        sh_out.cell(row=r+1+1, column=c+1, value=str(df.iloc[r, c]))
        elif strcmp(sh.title[:len("NameMapping_")], "NameMapping_"):
            # Read the header. It identifies two taxonomic ranks to be mapped
            read_mapping_and_join(sh, sh_writable, the_registry.metadatasets, the_registry.datasets, the_registry.maps, dsm)
        elif strcmp(sh.title[:len("PivotTable_")], "PivotTable_"):
            read_and_calculate_pivot_table(sh, sh_writable, the_registry.processors, the_registry.datasets)
        elif strcmp(sh.title[:len("Sequence_")], "Sequence_"):
            # TODO Mail to Michele: diagrams, slides, what I have done (angular, analysis and new commands), help
            # TODO needed (user's guide, ideas for the presentation)
            # TODO sequence declaration using existing processors
            # Three columns: source, destination and factor
            # At least the source or the destination must be factors, already existing
            # If one of the factors does not exist, it is automatically created
            # Instead of the literal value accompanying the "ff", a list of outgoing or incoming links will be attached
            # The link will appear in the source and in the destination
            # There will be also a central registry of links, containing the list of links, and the list of facts with
            # the corresponding processor or fact inside the processor (should be capable of recovering the processor)
            # A command "update values" will sweep all the facts involved in sequences (in the registry of links)
            # calculating the value using the opposite value and applying the factor, the accumulating until all
            # associated links are processed (for all the links of a fact, the role -source or destination- should be
            # the same)
            pass
        # EXTERNAL DATASET COMMANDS
        elif re_enum.search(sh.title):  # strcmp(sh.title, "Dataset_Eurostat_Enumerate") or strcmp(sh.title, "Dataset_SSP_Enumerate") or strcmp(sh.title, "Dataset_FAOSTAT_Enumerate"):
            # If the worksheet contains something at cell (1,1) assume
            if not sh_writable.cell(row=1, column=1).value:
                reset_cells_format(sh_writable)
                for s in sources:
                    if s.lower() in sh.title.lower():
                        get_codes_all_statistical_datasets(s, sh_writable, dsm)
                        break
                # if strcmp(sh.title, "Dataset_Eurostat_Enumerate"):
                #     get_codes_all_statistical_datasets("EuroStat", sh_writable, ds)
                # elif strcmp(sh.title, "Dataset_SSP_Enumerate"):
                #     get_codes_all_statistical_datasets("SSP", sh_writable, ds)
                # elif strcmp(sh.title, "Dataset_FAOSTAT_Enumerate"):
                #     get_codes_all_statistical_datasets("FAOSTAT", sh_writable, ds)

                # TODO Automatic width (DOES NOT WORK)
                # for v in sh_writable.column_dimensions.values():
                #     v.bestFit = True
        elif re_meta.search(sh.title):  # strcmp(sh.title[:len("Metadata_Eurostat_")], "Metadata_Eurostat_") or strcmp(sh.title[:len("Metadata_SSP_")], "Metadata_SSP_"):
            # Extract worksheet, source and dataset name
            for s in sources:
                if s.lower() in sh.title.lower():
                    source = s
                    dataset_name = sh.title[len("Metadata__")+len(s):]
                    wks_name = "Dataset_" + s + "_" + dataset_name
                    break
            # if strcmp(sh.title[:len("Metadata_Eurostat_")], "Metadata_Eurostat_"):
            #     source = "Eurostat"
            #     dataset_name = sh.title[len("Metadata_Eurostat_"):]
            #     wks_name = "Dataset_Eurostat_" + dataset_name
            # else:
            #     source = "SSP"
            #     dataset_name = sh.title[len("Metadata_SSP_"):]
            #     wks_name = "Dataset_SSP_" + dataset_name

            reset_cells_format(sh_writable)
            # Returns a list of dimensions. A dimension is a tuple (name, list of (code, description))
            # Also metadata, which is tuple (dimensions, attributes, measures)
            # Each of this is a NamedTuple SDMXConcept
            dims, metadata = get_statistical_dataset_structure(source, dataset_name, sh_writable, dsm)
            the_registry.metadatasets[dataset_name] = metadata  # Store the metadata
            if wks_name not in sh.parent.get_sheet_names():
                sh_writable = create_after_worksheet(sh_writable, wks_name)
                # Add default header
                for i, d in enumerate(dims):
                    sh_writable.cell(row=1, column=2*i+1, value=d[0])
                    if d[1]:
                        # Add code list
                        for r, c in enumerate(d[1]):
                            sh_writable.cell(row=r + 2, column=2*i + 1, value=c[0])
                            sh_writable.cell(row=r + 2, column=2*i + 2, value=c[1])
                # Add columns for PIVOT
                sh_writable.cell(row=1, column=2 * i + 3, value="Rows")
                sh_writable.cell(row=1, column=2 * i + 4, value="Columns")
                sh_writable.cell(row=1, column=2 * i + 5, value="AggFunc")
                sh_writable.cell(row=1, column=2 * i + 6, value="ShowTotals")

        elif re_data.search(sh.title):  # strcmp(sh.title[:len("Dataset_Eurostat_")], "Dataset_Eurostat_") or strcmp(sh.title[:len("Dataset_SSP_")], "Dataset_SSP_"):
            # Extract source and dataset name
            for s in sources:
                if s.lower() in sh.title.lower():
                    source = s
                    dataset_name = sh.title[len("Dataset__")+len(s):]
                    break

            # if strcmp(sh.title[:len("Dataset_Eurostat_")], "Dataset_Eurostat_"):
            #     source = "Eurostat"
            #     dataset_name = sh.title[len("Dataset_Eurostat_"):]
            # else:
            #     source = "SSP"
            #     dataset_name = sh.title[len("Dataset_SSP_"):]

            # Get metadata
            dims, metadata = get_statistical_dataset_structure(source, dataset_name, None, dsm)
            the_registry.metadatasets[dataset_name] = metadata  # Store the metadata
            # Different behavior if the header of the worksheet is empty (fill it with all the metadata) or
            # not (request a dataset slice)
            first_row_empty = True
            some_error = False
            for c in range(sh.max_column):
                v = sh.cell(row=1, column=c + 1).value
                if v and v.strip() != "":
                    first_row_empty = False
            if not first_row_empty:
                # Check if the request has been solved. If it has, do nothing
                request_solved = False
                for c in range(sh.max_column):
                    p = sh.cell(row=0 + 1, column=c + 1).value
                    if p and p.lower().startswith("value"):
                        request_solved = True

                if not request_solved:
                    reset_cells_format(sh_writable)
                    # Find possible additional columns
                    columns_from_maps = create_dictionary()
                    for d in metadata[0]:  # Dimensions
                        if dataset_name + "." + d in the_registry.maps:
                            for t in the_registry.maps[dataset_name + "." + d]:
                                columns_from_maps[t] = d

                    # Gather parameters. A parameter per column
                    params = create_dictionary()
                    dim_cols = []
                    dim_names = [d[0].lower() for d in dims]
                    dim_names_upper = ["'" + d.upper() + "'" for d in dim_names]
                    dims_codes = create_dictionary()
                    for d in dims:
                        if d[1]:
                            dims_codes[d[0]] = set([c[0] for c in d[1]])
                    pivot_cols = []
                    for c in range(sh.max_column):
                        p = sh.cell(row=0 + 1, column=c + 1).value
                        if p and (p.lower().strip() in dim_names or p.lower().strip() in columns_from_maps):
                            dim_cols.append(p.lower())
                            r = 1
                            lst = []
                            is_mapped = False
                            if p in dims_codes:
                                cl = [n.lower() for n in dims_codes[p]]
                            elif p in columns_from_maps:
                                cl = [n.lower() for n in the_registry.maps[dataset_name+"."+columns_from_maps[p]][p][0]]
                                is_mapped = True
                            else:
                                cl = None
                            while r < sh.max_row and sh.cell(row=r + 1, column=c + 1).value != "":
                                v = sh.cell(row=r + 1, column=c + 1).value
                                if v:
                                    val = cell_content_to_str(v)
                                    if cl:
                                        if val.lower() in cl:
                                            lst.append(val)
                                        else:
                                            some_error = True
                                            show_message(sh_writable, r + 1, c + 1, "Error: unrecognized code from dimension '"+p+"'")
                                    else:
                                        lst.append(val)
                                r += 1
                            if len(lst) > 0:
                                lst = lst if len(lst) > 1 else lst[0]
                                if is_mapped:
                                    d = columns_from_maps[p]
                                    mapped = the_registry.maps[dataset_name+"."+d][p][1]
                                    p = d  # Change the mapped dimension to the dimension name in the dataset
                                    lst = obtain_reverse_codes(mapped, lst)

                                if p in params:
                                    lst = list(set(params[p]).intersection(lst))

                                params[p] = lst
                        elif p and p.lower().strip() in ("rows", "columns", "aggfunc", "showtotals"):
                            pivot_cols.append(p.lower().strip())
                        elif p and p.strip() != "":
                            some_error = True
                            show_message(sh_writable, 0 + 1, c + 1, "Error: unrecognized column name. It must be a dimension name from dataset '" + dataset_name + "' (" + ', '.join(dim_names_upper) + ") or a parameter for the Pivot Table ('Rows', 'Columns', 'AggFunc', 'ShowTotal')")
                    is_dataset_requested = len(dim_cols) > 0
                    is_pivot_requested = "rows" in pivot_cols and "columns" in pivot_cols and "aggfunc" in pivot_cols
                    if not some_error:
                        if is_dataset_requested:
                            # TODO Consider trees
                            # TODO Reverse map for filter: if elements of a MuSIASEM map are specified, generate the linked "native" codes, for the query to the external dataset
                            # >>>> Elaborate the dataset <<<<
                            ds = dsm.get_dataset_filtered(source, dataset_name, params)
                            df = ds.data
                            # df = get_statistical_dataset(source, dataset_name, params)
                            if df is not None and df.shape[0] > 0:
                                # Try to find a map for each of the dimensions
                                for d in metadata[0]:  # Dimensions
                                    if dataset_name+"."+d in the_registry.maps:
                                        for t in the_registry.maps[dataset_name + "." + d]:
                                            mapped = the_registry.maps[dataset_name + "." + d][t][1]
                                            df_dst = pd.DataFrame(mapped, columns=['sou_rce', t])
                                            for di in df.columns:
                                                if strcmp(d, di):
                                                    d = di
                                                    break
                                            # Upper case column before merging
                                            df[d] = df[d].str.upper()
                                            df = pd.merge(df, df_dst, how='left', left_on=d, right_on='sou_rce')
                                            del df['sou_rce']

                                # TODO If the dataset is filtered by one of the maps, do it now

                                # Store the resulting Dataframe
                                the_registry.datasets[sh.title] = df

                                # Check if PivotTable or enumeration
                                if not is_pivot_requested:
                                    if df.shape[0] > 30000:
                                        show_message(sh_writable, 1, 1, "ERROR: the resulting dataset '"+sh.title +
                                                     "' cannot be enumerated because there are too many rows (" + str(df.shape[0]) +
                                                     "). Anyway, PivotTable operations on the dataset can be performed.")
                                    else:
                                        # Reset worksheet
                                        sh_out = reset_worksheet(sh_writable)
                                        # Put the dataframe in a new worksheet
                                        for c, t in enumerate(df.columns):
                                            sh_out.cell(row=0 + 1, column=c + 1, value=t)
                                        for r in range(df.shape[0]):
                                            for c in range(df.shape[1]):
                                                sh_out.cell(row=r+1+1, column=c+1, value=str(df.iloc[r, c]))
                                else:
                                    read_and_calculate_pivot_table(sh, sh_writable, the_registry.processors, the_registry.datasets, df)
                            else:
                                show_message(sh_writable, 1, 1, "ERROR: the request did not return data. "
                                                                "The dataset is off, the query was too restrictive or "
                                                                "it was not possible to connect to "+source)
                        else:
                            show_message(sh_writable, 1, 1, "Error: no dimension names where recognized")
            else:
                # Add default header
                for i, d in enumerate(dims):
                    sh_writable.cell(row=1, column=2*i+1, value=d[0])
                    if d[1]:
                        # Add code list
                        for r, c in enumerate(d[1]):
                            sh_writable.cell(row=r + 2, column=2*i + 1, value=c[0])
                            sh_writable.cell(row=r + 2, column=2*i + 2, value=c[1])
                i = 2 * i + 1
                # Add columns for maps, if found
                for d in metadata[0]:  # Dimensions
                    if dataset_name + "." + d in the_registry.maps:
                        mps = the_registry.maps[dataset_name + "." + d]
                        for t in mps:
                            codes = sorted(mps[t][0])
                            sh_writable.cell(row=1, column=i + 1, value=t)
                            # Add map code list
                            for r, c in enumerate(codes):
                                sh_writable.cell(row=r + 2, column=i + 1, value=c)
                            i += 1

                # Add columns for PIVOT
                sh_writable.cell(row=1, column=i + 3, value="Rows")
                sh_writable.cell(row=1, column=i + 4, value="Columns")
                sh_writable.cell(row=1, column=i + 5, value="AggFunc")
                sh_writable.cell(row=1, column=i + 6, value="ShowTotals")

    names = {}  # Dictionary with all the names, and what they are (flow, fund, processor, taxonomic rank, taxon)
    msm = None  # Structured view

    # Write and convert to bytes array
    from io import BytesIO
    bio = BytesIO()
    xl_out.save(bio)
    return bio.getbuffer()


# ----------------------------
# DEBUG
# ----------------------------
"""
Load: openpyxl.open_workbook -> wb (Workbook instance)
Save: wb.save
sheets: wb.get_sheet_names
a sheet: wb.get_sheet_by_name -> ws (Worksheet instance)
add sheet: wb.create_sheet(name, index) (create_chartsheet, copy_worksheet)
remove sheet: wb.remove_sheet(Worksheet instance),
worksheet name: ws.title
read/write cell: ws.cell(row, column) -> cell (Cell instance); row and column start in 1, not in 0
read/write cell value: cell.value
read/write cell fill format: cell.fill
read/write cell comment: cell.comment
read merged cells:
    for ra in ws.merged_cell_ranges:
        t = openpyxl.utils.range_boundaries(ra)  # min col, min row, max col, max row (max's included)

merge cells
columns width: ws.column_dimensions
row height: ws.row_dimensions

"""
if __name__ == '__main__':
    import collections
    import io
    from backend.models.statistical_datasets import *
    from backend.models.musiasem_methodology_support import *

    # SDMX Concept can be: dimension, attribute or measure. Stored in "metadatasets" associated to a dataset by its name
    app_cfg = collections.namedtuple('App', 'config')
    app = app_cfg({})
    app.config["SSP_FILES_DIR"] = "/home/rnebot/GoogleDrive/AA_MAGIC/Data/SSP/"

    # TODO Change to MONETDB
    recreate_db = False
    monetdb = False
    if monetdb:
        data_engine = create_monet_database_engine("monetdb://monetdb:monetdb@localhost:50000/", "db")
    else:
        data_engine = create_pg_database_engine("postgresql://postgres:postgres@localhost:5432/", "magic_data", recreate_db=recreate_db)

    engine = create_pg_database_engine("postgresql://postgres:postgres@localhost:5432/", "magic_nis", recreate_db=recreate_db)

    DBSession.configure(bind=engine)  # reconfigure the sessionmaker used by this scoped_session
    tables = ORMBase.metadata.tables
    connection = engine.connect()
    table_existence = [engine.dialect.has_table(connection, tables[t].name) for t in tables]
    connection.close()
    if False in table_existence:
        ORMBase.metadata.bind = engine
        ORMBase.metadata.create_all()

    # fn = "/home/rnebot/GoogleDrive/AA_MAGIC/DataRepository_Milestone/test_based_on_template.xlsx"
    fn = "/home/rnebot/GoogleDrive/AA_MAGIC/DataRepository_Milestone/test_based_on_template_MR_2.xlsx"
    fn = "/home/rnebot/GoogleDrive/AA_MAGIC/DataRepository_Milestone/test_based_on_template_MR_2_v2.xlsx"
    fn = "/home/rnebot/GoogleDrive/AA_MAGIC/DataRepository_Milestone/test_datasets_5a.xlsx"
    fn = "/home/rnebot/Downloads/3rdstep_test_dataset_0.xlsx"
    fn = "/home/rnebot/GoogleDrive/AA_MAGIC/DataRepository_Milestone/2ndstep_test_datasets_4.xlsx"
    fn = "/home/rnebot/GoogleDrive/AA_MAGIC/DataRepository_Milestone/backend_tests/test_datasets_7.xlsx"
    fn = "/home/rnebot/GoogleDrive/AA_MAGIC/DataRepository_Milestone/backend_tests/test_datasets_12.xlsx"
    fn = "/home/rnebot/GoogleDrive/AA_MAGIC/DataRepository_Milestone/test_datasets_6.xlsx"
    fn = "/home/rnebot/GoogleDrive/AA_MAGIC/DataRepository_Milestone/test_datasets_faostat.xlsx"
    # fn = "/home/rnebot/GoogleDrive/AA_MAGIC/DataRepository_Milestone/backend_tests/test_datasets_14.xlsx"
    with open(fn, "rb") as f:
        b = f.read()

    b = process_file(b)

    # b.save("/home/rnebot/GoogleDrive/AA_MAGIC/DataRepository_Milestone/test_based_on_template_output.xlsx")

    with open("/home/rnebot/GoogleDrive/AA_MAGIC/DataRepository_Milestone/test_based_on_template_output_2.xlsx", "wb") as f:
        f.write(b)
