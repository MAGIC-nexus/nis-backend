# -*- coding: utf-8 -*-


import json


import sqlalchemy.orm
import sqlalchemy.schema
from nis.model import DBSession, ORMBase, Diagram
import io
import openpyxl
import openpyxl.utils
from openpyxl.comments import Comment
import pandas as pd
import numpy as np
import requests
from multidict import MultiDict, CIMultiDict
import pint  # Units management
import re
import collections
import copy
import pandasdmx
from nis import app
from nis.eurostat_bulk import get_eurostat_filtered_dataset_into_dataframe

# GLOBAL VARIABLES
case_sensitive = False
ureg = pint.UnitRegistry()

# ------------------------------------------------------------------------------------------------------------


def create_estat_request():
    # EuroStat datasets
    if 'CACHE_FILE_LOCATION' in app.config:
        cache_name = app.config['CACHE_FILE_LOCATION']
    else:
        cache_name = "/tmp/sdmx_datasets_cache"
    r = pandasdmx.Request("ESTAT", cache={"backend": "sqlite", "include_get_headers": True,
                                          "cache_name": cache_name})
    r.timeout = 180
    return r

estat = create_estat_request()

# #################################################################
# CASE SeNsItIvE or INSENSITIVE names (flows, funds, processors, ...)
#

class CaseInsensitiveDict(collections.MutableMapping):
    """
    A dictionary with case insensitive Keys.
    Prepared also to support TUPLES as keys, required because compound keys are required
    """
    def __init__(self, data=None, **kwargs):
        from collections import OrderedDict
        self._store = OrderedDict()
        if data is None:
            data = {}
        self.update(data, **kwargs)

    def __setitem__(self, key, value):
        # Use the lowercased key for lookups, but store the actual
        # key alongside the value.
        if not isinstance(key, tuple):
            self._store[key.lower()] = (key, value)
        else:
            self._store[tuple([k.lower() for k in key])] = (key, value)

    def __getitem__(self, key):
        if not isinstance(key, tuple):
            return self._store[key.lower()][1]
        else:
            return self._store[tuple([k.lower() for k in key])][1]

    def __delitem__(self, key):
        if not isinstance(key, tuple):
            del self._store[key.lower()]
        else:
            del self._store[tuple([k.lower() for k in key])]

    def __iter__(self):
        return (casedkey for casedkey, mappedvalue in self._store.values())

    def __len__(self):
        return len(self._store)

    def lower_items(self):
        """Like iteritems(), but with all lowercase keys."""
        return (
            (lowerkey, keyval[1])
            for (lowerkey, keyval)
            in self._store.items()
        )

    def __eq__(self, other):
        if isinstance(other, collections.Mapping):
            other = CaseInsensitiveDict(other)
        else:
            return NotImplemented
        # Compare insensitively
        return dict(self.lower_items()) == dict(other.lower_items())

    # Copy is required
    def copy(self):
        return CaseInsensitiveDict(self._store.values())

    def __repr__(self):
        return str(dict(self.items()))


def create_dictionary(multi_dict=False):
    """
    Factory to create dictionaries used by the prototype

    It reads the "case_sensitive" global variable

    :param multi_dict: True to create a "MultiDict", capable of storing several values
    :return:
    """

    if not multi_dict:
        if case_sensitive:
            return {}
        else:
            return CaseInsensitiveDict()
    else:
        if case_sensitive:
            return MultiDict()
        else:
            return CIMultiDict()


def strcmp(s1, s2):
    """
    Compare two strings for equality or not, considering a flag for case sensitiveness or not

    It also removes leading and trailing whitespace from both strings, so it is not sensitive to this possible
    difference, which can be a source of problems

    :param s1:
    :param s2:
    :return:
    """
    if case_sensitive:
        return s1.strip() == s2.strip()
    else:
        return s1.strip().lower() == s2.strip().lower()
#
#
# #################################################################


def get_codes_all_statistical_datasets(source, sh_out):
    """
    Obtain a list of datasets available from a source
    If no source is specified, all the sources are queried
    For each dataset, the source, the name, the periods available, an example command and a description are obtained

    :param source:
    :param dataset_filter:
    :return: A Dataframe with the list of datasets
    """
    if source.lower() == "eurostat":
        import xmltodict
        # Make a table of datasets, containing three columns: ID, description, URN
        # List of datasets
        xml = requests.get("http://ec.europa.eu/eurostat/SDMX/diss-web/rest/dataflow/ESTAT/all/latest")
        t = xml.content.decode("utf-8")
        j = xmltodict.parse(t)
        sh_out.cell(row=0 + 1, column=0 + 1).value = "Dataset ID"
        sh_out.cell(row=0 + 1, column=1 + 1).value = "Description"
        sh_out.cell(row=0 + 1, column=2 + 1).value = "URN"
        for r, k in enumerate(j["mes:Structure"]["mes:Structures"]["str:Dataflows"]["str:Dataflow"]):
            for n in k["com:Name"]:
                if n["@xml:lang"] == "en":
                    desc = n["#text"]
                    break
            dsd_id = k["str:Structure"]["Ref"]["@id"]
            sh_out.cell(row=r + 1 + 1, column=0 + 1).value = k["@id"]
            sh_out.cell(row=r + 1 + 1, column=1 + 1).value = desc
            sh_out.cell(row=r + 1 + 1, column=2 + 1).value = k["@urn"]
            # print(dsd_id + "; " + desc + "; " + k["@id"] + "; " + k["@urn"])


def get_statistical_dataset_structure(source, dataset, sh_out=None):
    """
    Obtain the DSD containing the dimensions, attributes, measures, code lists

    :param source:
    :param dataset:
    :return: List of dimension names (for the header)
    """
    if source.lower() == "eurostat":
        refs = dict(references='all')
        dsd_response = estat.datastructure("DSD_" + dataset, params=refs)
        dsd = dsd_response.datastructure["DSD_" + dataset]
        metadata = dsd_response.write()
        # Dimensions and Attributes
        Concept = collections.namedtuple('Concept', 'name istime description code_list')
        dims = create_dictionary()  # Each dimension has a name, a description and a code list
        attrs = create_dictionary()
        meas = create_dictionary()
        for d in dsd.dimensions:
            istime = str(dsd.dimensions.get(d)).split("|")[0].strip() == "TimeDimension"
            dims[d] = Concept(d, istime, "", None)
        for a in dsd.attributes:
            attrs[a] = Concept(a, False, "", None)
        for m in dsd.measures:
            meas[m] = None
        for l in metadata.codelist.index.levels[0]:
            first = True
            # Read code lists
            cl = create_dictionary()
            for m, v in list(zip(metadata.codelist.ix[l].index, metadata.codelist.ix[l]["name"])):
                if not first:
                    cl[m] = v
                else:
                    first = False

            if metadata.codelist.ix[l]["dim_or_attr"][0] == "D":
                istime = str(dsd.dimensions.get(l)).split("|")[0].strip() == "TimeDimension"
                dims[l] = Concept(l, istime, "", cl)
            else:
                attrs[l] = Concept(l, False, "", cl)
        # Make a table of dimensions and code lists, containing three columns: dimension name, code, code_description
        if sh_out:
            sh_out.cell(row=0 + 1, column=0 + 1, value="Dimension name")
            sh_out.cell(row=0 + 1, column=1 + 1, value="Code")
            sh_out.cell(row=0 + 1, column=2 + 1, value="Code description")
        r = 1
        lst_dim = []
        time_dim = False
        for l in dims:
            lst_dim_codes = []
            if dims[l].istime:
                time_dim = True
            else:
                lst_dim.append((dims[l].name, lst_dim_codes))

            if dims[l].code_list:
                for c in dims[l].code_list:
                    if sh_out:
                        sh_out.cell(row=r + 1, column=0 + 1, value=l + (" (TimeDimension)" if dims[l].istime else ""))
                        sh_out.cell(row=r + 1, column=1 + 1, value=c)
                        sh_out.cell(row=r + 1, column=2 + 1, value=dims[l].code_list[c])

                    lst_dim_codes.append((c, dims[l].code_list[c]))

                    r += 1
            else:
                if sh_out:
                    sh_out.cell(row=r + 1, column=0 + 1, value=l + (" (TimeDimension)" if dims[l].istime else ""))
                r += 1
        if time_dim:
            lst_dim.append(("startPeriod", None))
            lst_dim.append(("endPeriod", None))

        return lst_dim


def get_statistical_dataset(source, dataset, dataset_params):
    """
    Obtain a dataset given some parameters
    :param source:
    :param dataset: name of the dataset to retrieve. To obtain a list, call "obtain_datasets"
    :param dataset_params: list of (key, value) pairs filtering the dataset to be obtained. The possible parameters depend
    :return: pd.Dataframe containing the resulting dataset as a facts table, ready for OLAP analysis (like Pivot Table)
    """
    if source.lower() == "eurostat":
        method = 1
        if method == 1:
            df = get_eurostat_filtered_dataset_into_dataframe(dataset, dataset_params, update=False)
            return df
        else:
            params = {}
            if "startPeriod" in dataset_params:
                params["startPeriod"] = dataset_params["startPeriod"]
                del dataset_params["startPeriod"]
            if "endPeriod" in dataset_params:
                params["endPeriod"] = dataset_params["endPeriod"]
                del dataset_params["endPeriod"]

            d = estat.get(resource_type="data", resource_id=dataset, key=dataset_params,
                          params=params)
            df = d.write(d.msg)
            if isinstance(df, pd.DataFrame):
                # Convert to a table of facts, which could be processed by a PivotTable
                col_names = []
                c = 0
                if isinstance(df.columns, pd.MultiIndex):
                    for c, d in enumerate(df.columns.names):
                        col_names.append(d)
                    c += 1
                else:
                    pass  # What to do in this case?

                if isinstance(df.index, pd.MultiIndex):
                    for d in df.index.names:
                        col_names.append(d)
                        c += 1
                else:
                    col_names.append(df.index.name)
                    c += 1

                col_names.append("VALUE")

                data = np.zeros((df.shape[0]*df.shape[1], len(col_names))).astype(object)
                r = 0
                for row in range(df.shape[0]):
                    for col in range(df.shape[1]):
                        # Get values for the columns
                        c = 0
                        if isinstance(df.columns, pd.MultiIndex):
                            for c, l in enumerate(df.columns.values[col]):
                                data[r, c] = str(l)
                            c += 1
                        else:
                            pass  # What to do in this case?

                        if isinstance(df.index, pd.MultiIndex):
                            for l in df.index.values[row]:
                                data[r, c] = str(l)
                                c += 1
                        else:
                            data[r, c] = str(df.index[row])
                            c += 1
                        # Value
                        data[r, c] = df.iloc[row, col]
                        r += 1
                # Create a Dataframe
                df = pd.DataFrame(data, columns=col_names)
                cn = df.columns[-1]
                # Convert "Value" column (the last column) to numeric
                df[cn] = df[cn].apply(lambda x: pd.to_numeric(x, errors='coerce'))
                return df
            else:
                # ERROR: it did not return a Dataset
                return None
    else:
        return None

# ------------------------------------------------------------------------


def process_metadata(sh_in):
    """
    Analyze metadata

    :param sh_in: Input/output sheet containing metadata
    :return:
    """
    def read_metadata():
        for r in range(sh_in.max_row):
            k = None
            for c in range(sh_in.max_column):
                cell = sh_in.cell(row=r + 1, column=c + 1)
                if c == 0:  # Key
                    k = cell.value
                else:
                    v = cell.value
                    if k in d:
                        l_values = d[k]
                    else:
                        l_values = []
                        d[k] = l_values
                    l_values.append(v)

    def write_and_validate_metadata():
        # Process each field then remove from the dictionary
        r = 0  # First row
        # Simple DC fields not covered:
        #  type (controlled),
        #  format (controlled),
        #  rights (controlled),
        #  publisher,
        #  contributor,
        #  relation
        #
        # XML Dublin Core: http://www.dublincore.org/documents/dc-xml-guidelines/
        # Exhaustive list: http://dublincore.org/documents/dcmi-type-vocabulary/

        # Fields: ("<field label in excel file>", "<field name in Dublin Core>", Mandatory?, Controlled?)
        # TODO For Excel fields going to the same DC field AND with no controlled vocabulary, it is not possible to
        #      do the reverse transformation, from DC to Excel, immediately. Only if some kind of syntax is enforced
        #      for each of these fields
        k_list = [("Case study name", "title", True, False),
                  ("Title", "title", True, False),
                  ("Subject, topic and/or keywords", "subject", False, True),
                  ("Description", "description", False, False),
                  ("Level", "description", False, True),
                  ("Dimensions", "subject", True, True),
                  ("Reference documentation", "source", False, False),
                  ("Authors", "creator", True, False),
                  ("Date of elaboration", "date", True, False),
                  ("Temporal situation", "coverage", True, False),
                  ("Geographical location", "coverage", True, True),
                  ("DOI", "identifier", False, False),
                  ("Language", "language", True, True)
                  ]
        for k in k_list:
            if k[0] in d:
                first = True
                for c, v in enumerate(d[k[0]]):
                    if isinstance(v, int) or (isinstance(v, str) and v.strip() != ""):
                        # TODO Control possible values of "v" if k[3] is TRUE. v in controlled_vocabulary?
                        # sh_in.cell(row=r + 1, column=0 + 1).value = k[0]
                        # sh_in.cell(row=r + 1, column=c + 1 + 1).value = v
                        dc.append((k[1], v))
                    else:
                        if first and k[2]:
                            # sh_in.cell(row=r + 1, column=0 + 1).value = k[0]
                            s = "'" + k[0] + "' was not specified"
                            # sh_in.cell(row=r + 1, column=2 + 1).value = s
                            issues.append(s)
                            break
                    first = False
            else:
                if k[2]:
                    # sh_in.cell(row=r + 1, column=0 + 1).value = k[0]
                    s = "'" + k[0] + "' field should be present. Inserted automatically."
                    # sh_in.cell(row=r + 1, column=2 + 1).value = s
                    issues.append(s)
            r += 1
        return dc, issues

    def generate_dublin_core_xml():
        s = """
<?xml version="1.0"?>

<caseStudyMetadata
  xmlns="http://magic-nexus.org/dmp/"
  xmlns:dc="http://purl.org/dc/elements/1.1/">
        """
        for t in dc:
            s += "  <dc:" + t[0] + ">" + str(t[1]) + "</dc:" + t[0] + ">\n"

        s += "</caseStudyMetadata>"

        return s

    # -----------------------------------------------------------------

    # Elaborate a record
    d = create_dictionary()
    # Read information into dictionary
    read_metadata()
    # Put information back, validated (controlled vocabularies only)
    issues = []
    dc = []
    write_and_validate_metadata()

    xml = generate_dublin_core_xml()

    sh_in.cell(row=30, column=1).value = xml


def clone_processor(p):
    """
    Create a new processor with the same properties as the original
    Recursive copy child processors

    :param p:
    :return:
    """
    return copy.deepcopy(p)


def adimensionally_scale_processor(p, number):
    """
    Multiply ALL flows and funds of a processor and its descendants by "number"
    If a unit is specified, for each intensive processor with the same

    :param p:
    :param number:
    :return:
    """
    # If a unit is specified and the processor is intensive
    if not p["intensive"]:
        print("ERROR: in order to be able to multiply by a factor, the processor has to be intensive, not extensive.")
        return

    if "factors" in p:
        p["factors"] += " * " + str(number)
    else:
        p["factors"] = str(number)

    flows_funds = p["ff"]
    for ff in flows_funds:
        # Store the different factors as an expression which can be evaluated later (using "eval"). Append up-scales
        if flows_funds[ff]["unit"]:  # Only if there is a unit. If it is adimensional, do not apply the factor
            flows_funds[ff]["value"] *= number
        # TODO Register the scaling operation

    # Recurse into descendants
    if "children" in p:
        for ch in p["children"]:
            adimensionally_scale_processor(ch, number)


def create_after_worksheet(sh, title):
    """
    Create a new worksheet after the worksheet passed as parameter

    :param sh:
    :param title:
    :return: The new worksheet
    """
    def get_worksheet_index():
        for i, sh_name in enumerate(wb.get_sheet_names()):
            if sh_name == sh.title:
                break
        return i

    wb = sh.parent
    i = get_worksheet_index()
    return wb.create_sheet(title, i+1)


def reset_worksheet(sh):
    """
    Reset a worksheet by deleting and creating it from scratch

    :param sh:
    :return: The new worksheet
    """
    def get_worksheet_index():
        for i, sh_name in enumerate(wb.get_sheet_names()):
            if sh_name == sh.title:
                break
        return i

    wb = sh.parent
    i = get_worksheet_index()
    tmp = sh.title
    wb.remove_sheet(sh)
    return wb.create_sheet(tmp, i)


def id_str(v):
    """
    Auxiliary function

    :param v:
    :return:
    """
    if v:
        if isinstance(v, float) or isinstance(v, int):
            return str(int(v))
        else:
            return str(v)
    else:
        return None


def read_processors(sh_in, sh_out, registry):
    """
    Read a set of processors
    They will come with flows and funds (with value, unit, date, source, comment), types and name
    They need also a relation with the processor which originates the specialization, the combination matrix (or other structure) and so

    :param sh_in:
    :param sh_out:
    :param registry:
    :return: String with the name of the entry added to the registry
    """
    # Analyze column names, to know what property is in each
    first_taxonomic_rank = sh_in.title[11:]
    cols = create_dictionary()  # Name to Index map
    intensive_processors_specified = None  # True if intensive processors are being specified
    intensive_processors_unit = None  # Unit name of the fund that can realize a processor (intensive to extensive)
    error_count = 0
    some_known = False  # The first columns will be taxonomic ranks, with free names
    type_cols = []  # Indices of columns with type information
    type_cols_names = []
    taxonomic_ranks = create_dictionary()  # Dictionary of sets (but dictionaries are used to allow case insensitivity if desired)
    # TODO If "FF_Type" is found, Var_L1 and the other are ignored, show a warning
    for c in range(sh_in.max_column):
        cell = sh_in.cell(row=0 + 1, column=c + 1)
        col_name = cell.value
        if not col_name:
            continue
        col_name = col_name.replace(" ", "")
        if strcmp(col_name, "FF_Type") or strcmp(col_name, "Var_Type"):
            col_name = "FF_Type"
            some_known = True
        elif strcmp(col_name, "Var_L1"):
            col_name = "Var_L1"
            some_known = True
        elif strcmp(col_name, "Var_L2"):
            col_name = "Var_L2"
            some_known = True
        elif strcmp(col_name, "Var_L3"):
            col_name = "Var_L3"
            some_known = True
        elif strcmp(col_name, "FF_Name") or strcmp(col_name, "Var"):
            col_name = "FF_Name"
            some_known = True
        elif strcmp(col_name, "FF_Value") or strcmp(col_name, "Value"):
            col_name = "FF_Value"
            some_known = True
        elif strcmp(col_name[:8], "FF_Unit_") or strcmp(col_name[:5], "Unit_") or \
             strcmp(col_name[:8], "FF_Unit/") or strcmp(col_name[:5], "Unit/"):
            col_name = "FF_Unit_" + (col_name[8:] if strcmp(col_name[:1], "F") else col_name[5:])
            unit_col_name = col_name  # Store specially
            intensive_processors_specified = True
            intensive_processors_unit = col_name[8:]
            some_known = True
            # TODO Consider the link of the unit to the kind of fund, for instance "Land Use"
        elif strcmp(col_name[:8], "FF_Unit") or strcmp(col_name[:5], "Unit"):
            col_name = "FF_Unit"
            unit_col_name = col_name
            intensive_processors_specified = False
            some_known = True
        elif strcmp(col_name, "Scale") or strcmp(col_name, "Data_Scale"):
            col_name = "Scale"
            some_known = True
        elif strcmp(col_name, "FF_Source") or strcmp(col_name, "Source") or strcmp(col_name, "Data_Source"):
            col_name = "FF_Source"
            some_known = True
        elif strcmp(col_name[:10], "FF_Comment") or strcmp(col_name[:7], "Comment"):
            col_name = "FF_Comment"
            some_known = True
        elif strcmp(col_name, ""):
            continue  # Just ignore it
        else:  # Unknown column name
            if some_known:  # If some known column was processed previously --> warning (the column will be ignored)
                if not col_name.endswith("(not recognized)"):  # Avoid concatenating
                    col_name += " (not recognized)"
            else:
                # Assume a taxonomic rank is being defined in this column. Store it as a part of the type
                taxonomic_ranks[col_name] = create_dictionary()
                type_cols.append(c)
                type_cols_names.append(col_name)

        if col_name in cols:
            # TODO Error sh_out.write(1 + error_count + 1, 0 + 1, "'" + col_name + "' column can appear only one time")
            error_count += 1

        cols[col_name] = c
        # TODO Before: sh_out.write(0, c, col_name)

    if "FF_Type" not in cols and "Var_L1" in cols and "Var_L2" in cols and "Var_L3" in cols:
        cols["FF_Type"] = -1  # Just to pass the following test

    # There should exist at least one column indicating the type of the processor
    if len(type_cols) == 0:
        print("ERROR: no column identifying the processor type or name was found")

    # Check presence of mandatory columns
    mandatory = ["FF_Type", "FF_Name", "FF_Value", unit_col_name]
    for m in mandatory:
        if m not in cols:
            # TODO Error: sh_out.cell(1+error_count+1, 0+1, "There is no '" + m + "' column")
            error_count += 1
    if error_count > 0:
        # TODO Error: sh_out.cell(2+error_count+1, 0+1, "The processing has been stopped")
        return

    # Dictionary used to implement a shortcut to avoid typing values in a column: if the value of the current
    # row is equal to the value in the previous row, it is not needed to specify it.
    previous_line = {}
    current_line = {}

    proc_taxonomy = create_dictionary()  # Register processors taxonomy. All the different types
    procs = create_dictionary()  # Register processor instances (intensive or extensive)
    ff_taxonomy = create_dictionary()  # Register all flow fund types
    ffs = create_dictionary()

    some_error = False
    # Read each line, which will be both a processor and a flow/fund
    for r in range(1, sh_in.max_row):
        error_count = 0 + 1  # 1 is for an offset, errors are located in columns AFTER the information
        # Read values of current line
        for c in range(sh_in.max_column):
            cell = sh_in.cell(row=r + 1, column=c + 1)
            v = cell.value
            if (isinstance(v, str) and v.strip() == "") or not v:  # If empty, assume the last one
                if c in previous_line:
                    v = previous_line[c]
                    sh_out.cell(row=r + 1, column=c + 1, value=v)  # Copy to out worksheet

            current_line[c] = v
            previous_line[c] = v  # Overwrite last value for the column "c"

        # PROCESS ROW
        # Full taxon
        # TODO Check: there cannot exist an empty type value followed by a type to its right
        taxon = tuple([first_taxonomic_rank]+[id_str(current_line[c]) for c in type_cols if current_line[c] != ""])
        if taxon not in proc_taxonomy:
            proc_taxonomy[taxon] = None
        # Add subtypes to their corresponding taxonomic ranks
        for i, c in enumerate(type_cols):
            # Store a new entry (current_line[c]) in the taxonomic rank (type_cols_names[i])
            if id_str(current_line[c]) not in taxonomic_ranks[type_cols_names[i]]:
                taxonomic_ranks[type_cols_names[i]][id_str(current_line[c])] = None

        # Second, the processor, which includes the types and the name
        if taxon in procs:
            p = procs[taxon]
        else:
            p = {"intensive": intensive_processors_specified,
                 "relative_to_unit": intensive_processors_unit,
                 "full_name": taxon,
                 "name": taxon[-1]  # Last level MAY BE defining the processor
                 }
            procs[taxon] = p
        # Third, the fund or flow columns
        # Flows in the registry of flows for this type of processor
        ff_name = current_line[cols["FF_Name"]]
        if ff_name in ffs:
            ff_type = ffs[ff_name]
        else:
            ff_type = create_dictionary()
            ffs[ff_name] = ff_type

        # Flows and funds in a processor
        if "ff" in p:
            ff = p["ff"]
        else:
            ff = create_dictionary()
            p["ff"] = ff

        # Prepare the FF type

        # If the columns VAR_L1, VAR_L2, VAR_L3 were defined, convert to FF_Type
        if "FF_Type" not in cols and "Var_L1" in cols and "Var_L2" in cols and "Var_L3" in cols:
            v = [sh_in.cell(row=0 + 1, column=cols["Var_L" + str(i)] + 1).value for i in range(1, 4)]
            int_ext = None
            if strcmp(v[0][:3], "Int"):
                int_ext = "Int_"
            elif strcmp(v[0][:3], "Ext"):
                int_ext = "Ext_"
            in_out = None
            if strcmp(v[1][:3], "Inp"):
                int_ext = "In_"
            elif strcmp(v[1][:3], "Out"):
                int_ext = "Out_"
            flow_fund = None
            if strcmp(v[2][:3], "Flow"):
                int_ext = "Flow"
            elif strcmp(v[2][:3], "Fund"):
                int_ext = "Fund"
            if int_ext and in_out and flow_fund:
                ff_type = int_ext + in_out + flow_fund
            else:
                # TODO Improve
                sh_out.cell(row=r + 1, column=len(sh_in.ncols) + error_count + 1,
                            value="Variables defining a flow or fund type must be all defined correctly")
                error_count += 1
                some_error = True
        else:
            ff_type = current_line[cols["FF_Type"]]

        if ff_type not in ["Int_In_Flow", "Int_In_Fund", "Int_Out_Flow", "Ext_In_Flow", "Ext_Out_Flow"]:
            # TODO Improve
            sh_out.cell(row=r + 1, column=len(sh_in.ncols) + error_count + 1,
                        value="FF type not allowed ('Int_Out_Fund', 'Ext_In_Fund', 'Ext_Out_Fund' combinations do not apply)")
            error_count += 1
            some_error = True

        # Add the flow or fund
        if ff_name not in p["ff"]:
            if current_line[cols[unit_col_name]] and current_line[cols[unit_col_name]].strip() != "-":
                unit = current_line[cols[unit_col_name]]
                if strcmp(unit, "m3"):
                    unit = "m**3"
                elif strcmp(unit, "m2"):
                    unit = "m**2"

                unit = unit if not intensive_processors_specified else unit + "/" + intensive_processors_unit
            else:
                unit = ""
            ff[ff_name] = {"name": ff_name,
                           "type": ff_type,
                           "value": current_line[cols["FF_Value"]],
                           "unit": unit,
                           "scale": current_line[cols["Scale"]],
                           "source": current_line[cols["FF_Source"]],
                           "comment": current_line[cols["FF_Comment"]]
                           }
        else:
            sh_out.cell(row=r + 1, column=len(sh_in.ncols) + error_count + 1,
                        value="'" + ff_name + "' in processor '"+str(taxon)+"' is repeated. Row skipped.")
            error_count += 1
            some_error = True

    # Loop finished!

    # Last step is to register the results for following operations, if no error occurred
    if not some_error:
        registry[first_taxonomic_rank] = {"processors_taxonomy": proc_taxonomy,  # All the taxonomic types of processors, except the first level
                                          "processors": procs,  # The processors, also indexed by the full taxonomic types
                                          "taxonomic_rank_levels": type_cols_names,  # Names of the taxonomic ranks
                                          "taxonomic_ranks": taxonomic_ranks  # Names contained in each of the taxonomic ranks
                                          }
        return first_taxonomic_rank
    else:
        return None

# -----------------------------------------------
# COMBINE
# -----------------------------------------------


def obtain_rectangular_submatrices(mask, region=None):
    """
    Obtain rectangular submatrices of mask
    IMPORTANT: currently it only obtains ONE region

    :param mask: The original matrix, numpy.NDArray, containing only 0/1 (1 is "some content")
    :param region: A tuple (top, bottom, left, right) with indices to search. bottom and right are not included
    :return: The list of rectangular regions as tuples (top, bottom, left, right)
    """

    def nonzero_sequences(a):
        # Create an array that is 1 where a is non-zero, and pad each end with an extra 0.
        isnonzero = np.concatenate(([0], a != 0, [0]))
        absdiff = np.abs(np.diff(isnonzero))
        # Runs start and end where absdiff is 1.
        ranges = np.where(absdiff == 1)[0].reshape(-1, 2)
        return ranges

    lst = []
    if not region:
        region = (0, mask.shape[0], 0, mask.shape[1])  # All the mask
    submask = mask[region[0]:region[1], region[2]:region[3]]
    offset_col, offset_row = (region[2], region[0])
    # Accumulation of elements by row (resulting in a column vector)
    row_sum = np.sum(submask, axis=1)
    # Accumulation of elements by column (resulting in a row vector)
    col_sum = np.sum(submask, axis=0)

    # Ranges
    rs = nonzero_sequences(row_sum.flatten())
    cs = nonzero_sequences(col_sum.flatten())
    lst.append((rs[0][0], rs[0][1], cs[0][0], cs[0][1]))

    return lst


def worksheet_to_numpy_array(sh_in):
    """
    Obtain a replica of the worksheet into a Numpy NDArray, with combined cells (combined cells are repeated)

    :param sh_in:
    :return: The numpy array with the values of the worksheet
    """
    m = np.zeros((sh_in.max_row, sh_in.max_column)).astype(object)
    for r in range(sh_in.max_row):
        for c in range(sh_in.max_column):
            v = sh_in.cell(row=r + 1, column=c + 1).value
            if v:
                m[r, c] = v
            else:
                m[r, c] = 0.0

    # Merged cells
    for ra in sh_in.merged_cell_ranges:
        t = openpyxl.utils.range_boundaries(ra)  # min col, min row, max col, max row (max's included)
        mc = (t[1]-1, t[3]-1, t[0]-1, t[2]-1)  # Rearrange and subtract one
        v = m[mc[0], mc[2]]
        m[mc[0]:mc[1]+1, mc[2]:mc[3]+1] = v

    return m


def binary_mask_from_worksheet(sh_in, only_numbers=True):
    """
    Sweep the worksheet, considering merged cells, elaborate a mask for those cells which
    are not empty or contain a number

    :param sh_in:
    :param only_numbers:
    :return:
    """
    m = np.zeros((sh_in.max_row, sh_in.max_column), dtype=bool)
    for r in range(sh_in.max_row):
        for c in range(sh_in.max_column):
            v = sh_in.cell(row=r + 1, column=c + 1).value
            if v:
                if only_numbers:
                    if isinstance(v, int) or isinstance(v, float):
                        m[r, c] = 1
                else:
                    m[r, c] = 1

    # Merged cells
    for ra in sh_in.merged_cell_ranges:
        t = openpyxl.utils.range_boundaries(ra)  # min col, min row, max col, max row (max's included)
        mc = (t[1]-1, t[3]-1, t[0]-1, t[2]-1)  # Rearrange and subtract one
        v = m[mc[0], mc[2]]
        m[mc[0]:mc[1]+1, mc[2]:mc[3]+1] = v

    return m


def convert_intensive_processor_to_extensive(p, fund):
    """
    Convert the current processor from intensive to extensive
    The conversion is assumed (even if no flow or fund need the scaling, the processor will be extensive after this call)
    Also apply the conversion to descendants, recursively (children are converted first)

    :param p:
    :param fund: Fund, whose value is the total fund to be applied. The processor will contain scalings, so it
    will be multiplied by this factor
    :return:
    """

    if "children" in p:
        for c in p["children"]:
            convert_intensive_processor_to_extensive(c, fund)

    if p["intensive"]:
        fund["unit_intensive"] = "-"
        fund["value_intensive"] = "1"
        base_unit = 1 / ureg(fund["unit"])
        t = base_unit.dimensionality
        # TODO Is the fund already there? WARNING, because it should be
        # Multiply funds and flows which have the same unit in the denominator
        # print(p["full_name"])
        ff = p["ff"]
        for k in ff:
            # Analyze units of the current flow / fund
            un = ff[k]["unit"]
            dims = ureg(ff[k]["unit"]).to_tuple()
            found = False
            for d in dims[1]:
                v = ureg(d[0])**d[1]
                if v.dimensionality == t:
                    found = True
                    break
            if found:
                # Do a dimensional multiplication of the factors
                v = ff[k]["value"] * ureg(ff[k]["unit"]) * fund["value"] * ureg(fund["unit"])
                # Format using short unit names, then split in two parts (magnitude and unit)
                s = "{:~P}".format(v).split(" ", 1)
                ff[k]["value_intensive"] = ff[k]["value"]
                ff[k]["unit_intensive"] = ff[k]["unit"]
                ff[k]["value"] = float(s[0])
                ff[k]["unit"] = s[1]  # Unit is changed
                # print(un + " -> " + s[1])
            else:
                # print(un + "(not converted)")
                pass

        # Clone the fund, scale it using the accumulated scaling factors, then inject it into the processor
        f2 = copy.deepcopy(fund)
        f2["value"] *= eval(p["factors"])
        ff[f2["name"]] = f2
        # Mark the processor as "extensive"
        p["intensive"] = False
        p["relative_to_unit"] = None


def list_processors_set(registry, entry, processor_type):
    """
    Traverse the entry and obtain a list of flows and funds, accompanied with the preceding typologies, showing for each
    flow and/or fund the hierarchy needed to get to the processor and the processor itself

    :param registry: The registry of processor layers
    :param entry: Which registry entry specifically to read
    :param processor_type: A processor type (the registry will be analyzed to recover related taxonomic ranks)
    :return: pd.DataFrame containing the desired structure
    """

    def add():
        nonlocal idx, reg, lst, distinct_fields
        t = []
        for l in lst:
            for m2 in l:
                if m2[0] not in distinct_fields:
                    distinct_fields[m2[0]] = idx
                    idx += 1
                t.append((m2[0], m2[1]))
        reg.append(t)

    def add_itself_and_children(p):
        nonlocal m, lst
        tr = m[p["full_name"][0]]
        tmp = []  # ("ProcessorType", p["full_name"][0])
        tmp.extend([(st, sti) for st, sti in zip(tr, p["full_name"][1:])])
        if "factors" in p:
            tmp.append((p["full_name"][0] + "_factors", p["factors"]))
        lst.append(tmp)  # Add a pack of fields for the level
        if p["full_name"][0] != processor_type:
            if "children" in p:
                for p2 in p["children"]:
                    add_itself_and_children(p2)
        else:
            lst.append([("ProcessorType", p["full_name"][0])])
            for k in p["ff"]:
                lst.append([(l, p["ff"][k][l]) for l in p["ff"][k] if l in
                            ["name", "unit", "value", "type", "value_intensive", "unit_intensive"]])
                # Add to reg
                add()
                del lst[-1]
            del lst[-1]
        del lst[-1]  # Remove fields for the level

    # Find taxonomic rank names for the different processors types. Elaborate a map of lists
    distinct_fields = dict()
    idx = 0
    m = {}
    lst = []
    reg = []
    for k in registry:
        m[k] = registry[k]["taxonomic_rank_levels"]
    for p in entry["processors"]:
        add_itself_and_children(entry["processors"][p])
    # List of fields
    fn = np.zeros((len(distinct_fields))).astype(object)
    for k in distinct_fields:
        fn[distinct_fields[k]] = k
    # Convert reg to Dataframe
    d = np.zeros((len(reg), len(distinct_fields))).astype(object)
    for r, lst in enumerate(reg):
        for lst2 in lst:
            d[r, distinct_fields[lst2[0]]] = lst2[1]
    # Create a Dataframe
    df = pd.DataFrame(d, columns=fn)
    # Convert "Value" column (the last column) to numeric
    df['value'] = df['value'].apply(lambda x: pd.to_numeric(x, errors='coerce'))

    return df


def add_child_to(contained, container, factor):
    """
    Add contained to container, as CHILD ("children" list or dict)
    CONTAINER inherits all flows and funds. The accounted value is reset to zero when a new flow are fund are created
    The linked values need to be updated on request, so when two FF's (from container and contained) are linked, the value
    at the container will be updated automatically.

    :param contained:
    :param container:
    :param factor:
    :return:
    """
    if "children" in container:
        ch = container["children"]
    else:
        ch = []
        container["children"] = ch

    ch.append(contained)
    # Cascade "factor" to contained
    adimensionally_scale_processor(contained, factor)
    # Cascade "extensive" maker to contained with "intensive" nature
    if not container["intensive"] and contained["intensive"]:
        # Obtain the right number
        unit = contained["relative_to_unit"]
        fund = None
        for ff in container["ff"]:
            if container["ff"][ff]["unit"].lower() == unit.lower():
                fund = container["ff"][ff]
                break
        if fund:
            # Convert the contained processor to extensive (the fund is injected also). Also contained children
            convert_intensive_processor_to_extensive(contained, fund)

    # Inherit all flows and funds from contained
    for ff in contained["ff"]:
        if contained["ff"][ff]["name"] not in container["ff"]:
            ff_prime = copy.deepcopy(contained["ff"][ff])
            ff_prime["value"] = 0.0
            container["ff"][ff] = ff_prime
        else:
            ff_prime = container["ff"][ff]

        ff_prime["value"] += contained["ff"][ff]["value"]

"""
            ff[ff_name] = {"name": ff_name,
                           "type": ff_type,
                           "value": current_line[cols["FF_Value"]],
                           "unit": current_line[cols["FF_Unit"]] if not intensive_processors_specified
                              else current_line[cols["FF_Unit"]] + "/" + intensive_processors_unit,
                           "scale": current_line[cols["Scale"]],
                           "source": current_line[cols["FF_Source"]],
                           "comment": current_line[cols["FF_Comment"]]
                           }
"""


def combine_two_layers_of_processors(sh_in, sh_out, registry):
    """
    Elaborate a new set of processors based on the combination of two existing sets of processors
    A table specifying which processor sets are combined is passed as input (sh_in)
    A set of processors is created

    :param sh_in:
    :param sh_out: Just to show errors and warnings. No content is modified
    :param registry:
    :return:
    """
    m = binary_mask_from_worksheet(sh_in, True)  # True for cells containing numbers
    # Locate the matrix with numbers. Assume this defines the labels to consider, they will be around the matrix
    t = obtain_rectangular_submatrices(m)
    t = t[0]  # Take just the first element
    # print(t[0])
    v = worksheet_to_numpy_array(sh_in)
    for t in [(t[0], t[1], t[2], t[3]), (t[0] + 1, t[1], t[2], t[3]), (t[0], t[1], t[2] + 1, t[3])]:
        f = v[t[0]:t[1], t[2]:t[3]].astype(np.float64)
        row_sum = np.sum(f, axis=1)  # A column vector. If "all ones", the container will be along rows
        col_sum = np.sum(f, axis=0)  # A row vector. If "all ones", the container will be along columns
        container_situation = None
        if np.allclose(row_sum, 1, 1e-2):
            container_situation = "in_rows"
        if np.allclose(col_sum, 1, 1e-2):
            if container_situation:
                show_message(sh_out, t[0] - 1, t[1] - 1,
                             "ERROR: both rows and columns should not sum to ones")
            container_situation = "in_columns"
        if container_situation:
            break
    if not container_situation:
        show_message(sh_out, t[0] - 1, t[1] - 1, "ERROR: neither the sum of rows nor of columns is summing to ones")

    # Read cell containing the specification of the processors.
    # The FIRST is the CONTAINED, the SECOND the CONTAINER.
    # They are specified separated by "/" or "-".
    # Iterate to find both in the registry, obtain "registry_entries"
    RegistryEntry = collections.namedtuple('RegistryEntry', 'name registry_name registry_entry')
    registry_entries = []
    for s in re.split("\/|\-|_", sh_in.title[8:]):  # v[t[0]-1, t[2]-1]
        if "Upscaled_"+s.strip() in registry:
            registry_entries.append(RegistryEntry(s.strip(), "Upscaled_"+s.strip(), registry["Upscaled_"+s.strip()]))
        elif s.strip() in registry:
            registry_entries.append(RegistryEntry(s.strip(), s.strip(), registry[s.strip()]))
    if len(registry_entries) != 2:
        show_message(sh_out, 0, 0, "ERROR, there should be two registry entries recognized. Either the registry "
                                   "does not contain these processors or they are written incorrectly")

    # Identify taxonomic ranks. These ranks identify the processors
    # Concatenate all the taxonomic_ranks
    t_ranks = [(k, v) for k, v in registry_entries[0].registry_entry["taxonomic_ranks"].items()] + \
              [(k, v) for k, v in registry_entries[1].registry_entry["taxonomic_ranks"].items()]
    n_ranks_first_processor = len(registry_entries[0].registry_entry["taxonomic_rank_levels"])
    # Find the rows and cols containing taxon specifications
    lst = []
    for rup in range(t[0] - 1, -1, -1):
        # Take a slice of labels. Using "set" avoid repetitions. Elaborate a tuple
        lst.append((set(v[rup, t[2]:t[3]]), "row", rup))
    for cleft in range(t[2] - 1, -1, -1):
        # Take a slice of labels. Using "set" avoid repetitions. Elaborate a tuple
        lst.append((set(v[t[0]:t[1], cleft]), "col", cleft))
    col_idx = []
    row_idx = []
    for l in lst:
        found = False
        for e in l[0]:
            for t_rank in t_ranks:
                if e and id_str(e) in t_rank[1]:
                    found = True
                    break
            if found:
                break
        if found:
            if l[1] == "col":
                col_idx.append(l[2])
            else:
                row_idx.append(l[2])

    # Sweep the matrix of numbers
    warned_not_founds = set()
    new_procs = create_dictionary()  # Newly obtained container processors
    for r in range(t[0], t[1]):
        for c in range(t[2], t[3]):
            if v[r, c] < 1e-3:
                continue

            # Clear taxon identification for each processor
            taxa0 = {}
            taxa1 = {}
            # Find the taxa, for each processor type involved
            for rup in row_idx:
                e = id_str(v[rup, c])
                found = False
                for i, t_rank in enumerate(t_ranks):
                    if e and id_str(e) in t_rank[1]:
                        if i < n_ranks_first_processor:
                            idx = i
                            taxa = taxa0
                            entry = 0
                        else:
                            idx = i - n_ranks_first_processor
                            taxa = taxa1
                            entry = 1
                        found = True
                        # Add taxon to the processor type
                        taxa[registry_entries[entry].registry_entry["taxonomic_rank_levels"][idx]] = e
                if not found:
                    cell_pos = (rup + 1, c + 1)
                    print(sh_in.title + " ("+str(cell_pos[0])+", "+str(cell_pos[1])+") not found")
                    if e:
                        if cell_pos not in warned_not_founds:
                            warned_not_founds.add(cell_pos)
                            show_message(sh_out, cell_pos[0], cell_pos[1], "WARNING: subtype '"+e+"' not found. Processors to be combined cannot be identified properly.", "warning")

            for cleft in col_idx:
                e = id_str(v[r, cleft])
                found = False
                for i, t_rank in enumerate(t_ranks):
                    if e and e in t_rank[1]:
                        if i < n_ranks_first_processor:
                            idx = i
                            taxa = taxa0
                            entry = 0
                        else:
                            idx = i - n_ranks_first_processor
                            taxa = taxa1
                            entry = 1
                        found = True
                        # Add taxon to the processor type
                        taxa[registry_entries[entry].registry_entry["taxonomic_rank_levels"][idx]] = e
                if not found:
                    cell_pos = (r + 1, cleft + 1)
                    print(sh_in.title + " ("+str(cell_pos[0])+", "+str(cell_pos[1])+") not found")
                    if e:
                        if cell_pos not in warned_not_founds:
                            warned_not_founds.add(cell_pos)
                            show_message(sh_out, cell_pos[0], cell_pos[1], "WARNING: subtype '"+e+"' not found. Processors to be combined cannot be identified properly.", "warning")
            # Identify processors
            taxa = [taxa0, taxa1]
            processors = []
            for entry in [0, 1]:
                try:
                    p = registry_entries[entry].registry_entry["processors"][tuple([registry_entries[entry].name]+[taxa[entry][rn] for rn in registry_entries[entry].registry_entry["taxonomic_rank_levels"]])]
                    processors.append(p)
                except:  # Not found. The combination may be referring to a processor which has not been detailed previously
                    pass

            if len(processors) != 2:
                show_message(sh_out, r + 1, c + 1, "WARNING: did not find two processors to combine", "warning")
                continue

            S = processors[0]
            T = processors[1]
            if T["full_name"] in new_procs:
                T2 = new_procs[T["full_name"]]
            else:
                T2 = clone_processor(T)  # Clone processor and its children
                new_procs[T["full_name"]] = T2
            S2 = clone_processor(S)  # Clone processor and its children
            add_child_to(S2, T2, v[r, c])

    container = registry_entries[1]
    registry["Upscaled_"+container.name] = \
        {"processors_taxonomy": container.registry_entry["processors_taxonomy"],  # All the taxonomic types of processors, except the first level
         "processors": new_procs,  # The processors, also indexed by the full taxonomic types
         "taxonomic_rank_levels": container.registry_entry["taxonomic_rank_levels"],  # Names of the taxonomic ranks
         "taxonomic_ranks": container.registry_entry["taxonomic_ranks"]  # Names contained in each of the taxonomic ranks
         }


def show_message(sh, r, c, message, type="error", accumulate=True):
    """
    It serves to show a cell with some type of error (warning or error)
    A message is shown in the comment
    The name of the sheet is changed with a prefix indicating there is at least an issue to be solved

    :param sh:
    :param r:
    :param c:
    :param type:
    :param message:
    :return:
    """
    from openpyxl.styles import PatternFill
    cell = sh.cell(row=r, column=c)
    fill = cell.fill
    if type == "error":
        fill = PatternFill("solid", fgColor="CC0000")
    elif type == "warning":
        fill = PatternFill("solid", fgColor="FFFF33")
    elif type == "info":
        fill = PatternFill("solid", fgColor="87CEEB")
    cell.fill = fill
    if accumulate:
        comment = cell.comment
        if comment:
            comment.text += "\n" + message
        else:
            comment = Comment(message, "NIS")
    else:
        comment = Comment(message, "NIS")
    cell.comment = comment
    # if type == "error":
    #     sh.title = "!" + sh.title


def cleanup_messages(sh):
    """
    Sweep a worksheet

    :param sh:
    :return:
    """
    pass


def pivot_table(sh_writable, df, rows, cols, aggs, values, show_totals):
    """
    Given Pivot Table parameters, elaborate it and put it in the output worksheet

    :param sh_writable:
    :param df:
    :param rows:
    :param cols:
    :param aggs:
    :param values:
    :param show_totals:
    :return:
    """
    if len(rows) > 0 and len(cols) > 0 and values:
        df2 = pd.pivot_table(df,
                             values=values,
                             index=rows,
                             columns=cols,
                             aggfunc=aggs, fill_value=np.NaN, margins=show_totals,
                             dropna=True, margins_name="Total")
        if df2.shape[0] * df2.shape[1] > 40000:
            show_message(sh_writable, 1, 1, "ERROR: the resulting pivot table cannot be shown because there are "
                                            "too many cells (" + str(df2.shape[0]*df2.shape[1]) +
                                            "). Please, reconsider the parameters.")
        else:
            # df2 = df2.swaplevel(0, len(df2.columns.names)-1, axis=1)
            i = 0
            # Number of levels of the columns Index
            nidx_rows = len(df2.columns.names)
            # Count the number of levels of the rows Index
            if isinstance(df2.index, pd.MultiIndex):
                nidx_cols = len(df2.index.names)
            else:
                nidx_cols = 1
            # Reference point
            start = (nidx_rows, nidx_cols)
            # Reset worksheet
            sh_out = reset_worksheet(sh_writable)
            # Put the values at (nr+1, nc+1)
            for r in range(df2.shape[0]):
                for c in range(df2.shape[1]):
                    sh_out.cell(row=r + start[0] + 1, column=c + start[1] + 1, value=df2.iloc[r, c])
            # Put the columns Index from (:, nc+1) on
            for c in range(df2.shape[1]):
                for r, l in enumerate(df2.columns.values[c]):
                    sh_out.cell(row=r + 1, column=c + start[1] + 1, value=str(l))
            # Put the rows Index from (nr+1, :) on
            for r in range(df2.shape[0]):
                if isinstance(df2.index, pd.MultiIndex):
                    for c, l in enumerate(df2.index.values[r]):
                        sh_out.cell(row=r + start[0] + 1, column=c + 1, value=str(l))
                else:
                    sh_out.cell(row=r + start[0] + 1, column=0 + 1, value=str(df2.index[r]))
                    # TODO Check consecutive values, to merge, in horizontal or in vertical


def read_and_calculate_pivot_table(sh, sh_writable, registry, dfs, df=None):
    # Read parameters
    row_c = None
    col_c = None
    agg_c = None
    sh_c = None
    sh_name = None
    for c in range(sh.max_column):
        cname = sh.cell(row=0 + 1, column=c + 1).value
        if not cname:
            continue
        if strcmp(cname, "Sheet") and df is None:
            sh_c = c
            # Sheet from where to read the input dataframe
            sh_name = sh.cell(row=1 + 1, column=c + 1).value
            if sh_name in dfs:
                df = dfs[sh_name]
            else:
                df = list_processors(sh_name[5:], sh_writable.cell(row=1 + 1, column=sh_c + 1), registry, dfs)
                if df is None:
                    show_message(sh_writable, 1 + 1, sh_c + 1,
                                "ERROR: could not find or elaborate '" + sh_name + "'")
                # TODO Check if "cname" corresponds to a worksheet (read sh.sheets)
                # If it exists, it may be a worksheet not generating
        elif strcmp(cname, "Rows"):
            row_c = c
            rows = []
            r = 1
            while r < sh.max_row and sh.cell(row=r + 1, column=c + 1).value:
                # TODO Check that the row exists in the input data
                v = sh.cell(row=r + 1, column=c + 1).value
                if not strcmp(v, "value"):
                    if df is not None:
                        for d in df.columns:
                            if d.lower() == v.lower():
                                v = d
                    rows.append(v)
                r += 1
        elif strcmp(cname, "Columns"):
            col_c = c
            cols = []
            r = 1
            while r < sh.max_row and sh.cell(row=r + 1, column=c + 1).value:
                # TODO Check that the column exists in the input data
                v = sh.cell(row=r + 1, column=c + 1).value
                if not strcmp(v, "value"):
                    if df is not None:
                        for d in df.columns:
                            if d.lower() == v.lower():
                                v = d
                    cols.append(v)
                r += 1
        elif strcmp(cname, "AggFunc"):
            agg_c = c
            aggs = []
            r = 1
            while r < sh.max_row and sh.cell(row=r + 1, column=c + 1).value:
                # TODO Check that the aggregator function exists
                fname = sh.cell(row=r + 1, column=c + 1).value
                if strcmp(fname, "sum"):
                    aggs.append(np.sum)
                elif strcmp(fname, "mean") or strcmp(fname, "average") or strcmp(fname, "avg"):
                    aggs.append(np.mean)
                elif strcmp(fname, "std") or strcmp(fname, "stddev") or strcmp(fname, "standarddeviation"):
                    aggs.append(np.std)
                elif strcmp(fname, "max") or strcmp(fname, "maximum"):
                    aggs.append(np.amax)
                elif strcmp(fname, "min") or strcmp(fname, "minimum"):
                    aggs.append(np.amin)
                r += 1
        elif strcmp(cname, "ShowTotals"):
            show_totals = False
            r = 1
            while r < sh.max_row and sh.cell(row=r + 1, column=c + 1).value:
                # TODO Check that the column exists in the input data
                v = sh.cell(row=r + 1, column=c + 1).value
                if isinstance(v, int):
                    cnv = int(v)
                    show_totals = cnv == 1
                elif isinstance(v, str):
                    show_totals = strcmp(v, "true")
                elif isinstance(v, bool):
                    show_totals = True if v else False
                else:
                    print("ERROR: value not recognized as boolean")
                r += 1

    # Values
    for cn in df.columns:
        if strcmp(cn, "value"):
            values = cn
            break

    if not values:
        show_message(sh_writable, 1 + 1, sh_c + 1,
                     "ERROR: could not find a values column in '" + sh_name + "' worksheet, for input to a PivotTable")

    elif len(rows) == 0:
        show_message(sh_writable, 0 + 1, row_c + 1, "ERROR: at least one row is needed")
    elif len(cols) == 0:
        show_message(sh_writable, 0 + 1, col_c + 1, "ERROR: at least one column is needed")
    elif df is not None:
        pivot_table(sh_writable, df, rows, cols, aggs, values, show_totals)


def reset_comments(sh_writable):
    for r in range(4):
        for c in range(4):
            cell = sh_writable.cell(row=r + 1, column=c + 1)
            cell.comment = None


def list_processors(name, cell, registry, dfs):
    if name not in dfs:
        s = "upscaled_" + name
        processors = None  # What layer of processors to enumerate
        for k in registry:
            if s.lower().find(k.lower()) == 0:
                processors = registry[k]
                if s.lower().find(k.lower() + "s") == 0:
                    s = s[len(k) + 2:]
                else:
                    s = s[len(k) + 1:]
                break
        bottom_level = None
        for k in registry:
            if s.lower() == k.lower():
                bottom_level = k
                break
            elif s.endswith("s") and s[
                                     :-1].lower() == k.lower():  # Allow plural termination in "s" ("es" is not considered...)
                bottom_level = k
                break

        if not processors:
            msg = "ERROR: could not find Top containing processor level in '" + name + \
                  ". May be there is some misspelling in the processor names"
            print(msg)
            show_message(cell.parent, cell.row, cell.col_idx, msg)

        if not bottom_level:
            msg = "ERROR: could not find requested bottom level in '" + name + \
                  ". May be there is some misspelling in the processor names"
            print(msg)
            show_message(cell.parent, cell.row, cell.col_idx, msg)

        if not processors or not bottom_level:
            return None

        # Find the specified set of processors, and enumerate at the specified level
        df = list_processors_set(registry, processors, bottom_level)
        # Store the Dataframe for Pivot Tables
        dfs[name] = df

    return dfs[name]


def process_file(input_file):
    """
    Receives a binary with an Excel file, processes it and returns a new Excel file containing the results of the
    processing

    Command and variables are case insensitive

    :param input_file: Input Excel file as byte array ("bytes")
    :return: Output Excel file as byte array ("bytes")
    """

    registry = create_dictionary()
    dfs = create_dictionary()  # Dictionary of Dataframes, for PivotTable generation
    n_combinations = 0  # Count number of combinations

    # Instance for reading, ignoring formulas reading numbers!
    xl_in = openpyxl.load_workbook(io.BytesIO(input_file), data_only=True)
    # This is also for reading, but does not read numbers when there are formulas
    xl_out = openpyxl.load_workbook(io.BytesIO(input_file))

    # Is it a case study to be stored or just tests? If there is a "Metadata" worksheet -> store
    persist = False
    in_sheets = []
    for sh_name in xl_in.get_sheet_names():
        in_sheets.append(sh_name)
        sh = xl_in.get_sheet_by_name(sh_name)
        print(sh_name+": "+str(sh.max_row)+" x "+str(sh.max_column))
        if strcmp(sh_name, "Metadata"):
            persist = True

    # A case study
    cs = None


    # Process sheets in sequence

    # The commands
    # * From the metadata elaborate a "metadata.xml" in a single cell
    # * Download dataset
    # * Read processors of type ...
    # * Upscale combining two types of processors
    # * Downscale (apply fund to processor types to instantiate them)
    # * Pivot table (elaborate a pivot table using some of the tables. Some parameters needed, has to be thought)

    metadata_processed = False
    for sh_name in in_sheets:
        sh = xl_in.get_sheet_by_name(sh_name)
        sh_writable = xl_out.get_sheet_by_name(sh_name)
        print("Currently processing: "+sh_name)
        if strcmp(sh.title, "Metadata"):
            if metadata_processed:
                pass  # To copy, no action is needed
            else:
                # Process the metadata. Fill the new sheet. Prepare and return a CaseStudy object
                cs = process_metadata(sh)
                metadata_processed = True
        elif strcmp(sh.title[:11], "Processors_"):
            # TODO Read a set of Processors of the same type
            # There will be a list of Processor types (given by the sheet name)
            # Processors of a type can be totally instantiated or partially instantiated
            # Processors will be qualified by a tuple of types, hierarchically organized
            # Processors will not have
            name = read_processors(sh, sh_writable, registry)
            # Generate and Store the Dataframe for Pivot Tables
            df = list_processors_set(registry, registry[name], name)
            dfs["Processors_"+name] = df  # Register with two names
            dfs["List_"+name+"_"+name] = df
        elif strcmp(sh.title[:8], "Upscale_"):
            # TODO Look for a table linking two lists of Processors of different types
            # Doing this, each processor of the parent level will contain all the children processors, recursively
            # And the flows and funds are passed into the parent level also, combining or aggregating
            combine_two_layers_of_processors(sh, sh_writable, registry)
            n_combinations += 1
        elif strcmp(sh.title[:5], "List_"):
            df = list_processors(sh.title[5:], sh_writable.cell(row=1, column=1), registry, dfs)
            if df is not None:
                # Reset worksheet
                sh_out = reset_worksheet(sh_writable)
                # Put the dataframe in a new worksheet
                for c, t in enumerate(df.columns):
                    sh_out.cell(row=0 + 1, column=c + 1, value=t)
                for r in range(df.shape[0]):
                    for c in range(df.shape[1]):
                        sh_out.cell(row=r+1+1, column=c+1, value=str(df.iloc[r, c]))

        elif strcmp(sh.title, "Dataset_Eurostat_Enumerate"):
            # If the worksheet contains something at cell (1,1) assume
            if not sh_writable.cell(row=1, column=1).value:
                reset_comments(sh_writable)
                get_codes_all_statistical_datasets("EuroStat", sh_writable)
                # TODO Automatic width (DOES NOT WORK)
                # for v in sh_writable.column_dimensions.values():
                #     v.bestFit = True
        elif strcmp(sh.title[:len("Metadata_Eurostat_")], "Metadata_Eurostat_"):
            dataset_name = sh.title[len("Metadata_Eurostat_"):]
            wks_name = "Dataset_Eurostat_" + dataset_name
            reset_comments(sh_writable)
            dims = get_statistical_dataset_structure("EuroStat", dataset_name, sh_writable)
            if wks_name not in sh.parent.get_sheet_names():
                sh_writable = create_after_worksheet(sh_writable, wks_name)
                # Add default header
                for i, d in enumerate(dims):
                    sh_writable.cell(row=1, column=2*i+1, value=d[0])
                    if d[1]:
                        # Add code list
                        for r, c in enumerate(d[1]):
                            sh_writable.cell(row=r + 2, column=2*i + 1, value=c[0])
                            sh_writable.cell(row=r + 2, column=2*i + 2, value=c[1])
                # Add columns for PIVOT
                sh_writable.cell(row=1, column=2 * i + 3, value="Rows")
                sh_writable.cell(row=1, column=2 * i + 4, value="Columns")
                sh_writable.cell(row=1, column=2 * i + 5, value="AggFunc")
                sh_writable.cell(row=1, column=2 * i + 6, value="ShowTotals")

        elif strcmp(sh.title[:len("Dataset_Eurostat_")], "Dataset_Eurostat_"):
            # Get metadata
            dataset_name = sh.title[len("Dataset_Eurostat_"):]
            dims = get_statistical_dataset_structure("EuroStat", dataset_name)
            reset_comments(sh_writable)
            # Different behavior if the header of the worksheet is empty (fill it with all the metadata) or
            # not (request a dataset slice)
            first_row_empty = True
            for c in range(sh.max_column):
                v = sh.cell(row=1, column=c + 1).value
                if v and v.strip() != "":
                    first_row_empty = False
            if not first_row_empty:
                # Check if the request has been solved. If it has, do nothing
                request_solved = False
                for c in range(sh.max_column):
                    p = sh.cell(row=0 + 1, column=c + 1).value
                    if p and p.lower().startswith("value"):
                        request_solved = True

                if not request_solved:
                    # Gather parameters. A parameter per column
                    params = create_dictionary()
                    dim_cols = []
                    pivot_cols = []
                    for c in range(sh.max_column):
                        p = sh.cell(row=0 + 1, column=c + 1).value
                        if p and p.lower() in [d[0].lower() for d in dims]:
                            dim_cols.append(p.lower())
                            r = 1
                            lst = []
                            while r < sh.max_row and sh.cell(row=r + 1, column=c + 1).value != "":
                                v = sh.cell(row=r + 1, column=c + 1).value
                                if v:
                                    if isinstance(v, str):
                                        lst.append(v)
                                    else:
                                        lst.append(str(int(v)))  # Integer numbers
                                r += 1
                            if len(lst) > 0:
                                params[p] = lst if len(lst) > 1 else lst[0]
                        else:
                            if p and p.lower().strip() in ("rows", "columns", "aggfunc", "showtotals"):
                                pivot_cols.append(p.lower().strip())
                    is_dataset_requested = len(dim_cols) > 0
                    is_pivot_requested = "rows" in pivot_cols and "columns" in pivot_cols and "aggfunc" in pivot_cols
                    if is_dataset_requested:
                        # Elaborate the dataset
                        df = get_statistical_dataset("EuroStat", sh.title[len("Dataset_Eurostat_"):], params)
                        if df is not None:
                            # Store the Dataframe for Pivot Tables
                            dfs[sh.title] = df
                            # Check if PivotTable or enumeration
                            if not is_pivot_requested:
                                if df.shape[0] > 30000:
                                    show_message(sh_writable, 1, 1, "ERROR: the resulting dataset '"+sh.title +
                                                 "' cannot be enumerated because there are too many rows (" + str(df.shape[0]) +
                                                 "). Anyway, PivotTable operations on the dataset can be performed.")
                                else:
                                    # Reset worksheet
                                    sh_out = reset_worksheet(sh_writable)
                                    # Put the dataframe in a new worksheet
                                    for c, t in enumerate(df.columns):
                                        sh_out.cell(row=0 + 1, column=c + 1, value=t)
                                    for r in range(df.shape[0]):
                                        for c in range(df.shape[1]):
                                            sh_out.cell(row=r+1+1, column=c+1, value=str(df.iloc[r, c]))
                            else:
                                read_and_calculate_pivot_table(sh, sh_writable, registry, dfs, df)
                                sh_writable.cell(row=1, column=1, value="Pivot table")
            else:
                # Add default header
                for i, d in enumerate(dims):
                    sh_writable.cell(row=1, column=2*i+1, value=d[0])
                    if d[1]:
                        # Add code list
                        for r, c in enumerate(d[1]):
                            sh_writable.cell(row=r + 2, column=2*i + 1, value=c[0])
                            sh_writable.cell(row=r + 2, column=2*i + 2, value=c[1])
                # Add columns for PIVOT
                sh_writable.cell(row=1, column=2 * i + 3, value="Rows")
                sh_writable.cell(row=1, column=2 * i + 4, value="Columns")
                sh_writable.cell(row=1, column=2 * i + 5, value="AggFunc")
                sh_writable.cell(row=1, column=2 * i + 6, value="ShowTotals")

        elif strcmp(sh.title[:len("NameMapping_")], "NameMapping_"):
            # Read the header. It identifies two taxonomic ranks to be mapped
            pass
        elif strcmp(sh.title[:len("PivotTable_")], "PivotTable_"):
            read_and_calculate_pivot_table(sh, sh_writable, registry, dfs)
            sh_writable.cell(row=1, column=1, value="Pivot table")

    names = {}  # Dictionary with all the names, and what they are (flow, fund, processor, taxonomic rank, taxon)
    msm = None  # Structured view

    # Write and convert to bytes array
    from io import BytesIO
    bio = BytesIO()
    xl_out.save(bio)
    return bio.getbuffer()


# ----------------------------
# DEBUG
# ----------------------------
"""
Load: openpyxl.open_workbook -> wb (Workbook instance)
Save: wb.save
sheets: wb.get_sheet_names
a sheet: wb.get_sheet_by_name -> ws (Worksheet instance)
add sheet: wb.create_sheet(name, index) (create_chartsheet, copy_worksheet)
remove sheet: wb.remove_sheet(Worksheet instance),
worksheet name: ws.title
read/write cell: ws.cell(row, column) -> cell (Cell instance); row and column start in 1, not in 0
read/write cell value: cell.value
read/write cell fill format: cell.fill
read/write cell comment: cell.comment
read merged cells:
    for ra in ws.merged_cell_ranges:
        t = openpyxl.utils.range_boundaries(ra)  # min col, min row, max col, max row (max's included)

merge cells
columns width: ws.column_dimensions
row height: ws.row_dimensions

"""
if __name__ == '__main__':
    import sys
    import io
    # fn = "/home/rnebot/GoogleDrive/AA_MAGIC/DataRepository_Milestone/test_based_on_template.xlsx"
    fn = "/home/rnebot/GoogleDrive/AA_MAGIC/DataRepository_Milestone/test_based_on_template_MR_2.xlsx"
    fn = "/home/rnebot/GoogleDrive/AA_MAGIC/DataRepository_Milestone/test_based_on_template_MR_2_v2.xlsx"
    fn = "/home/rnebot/GoogleDrive/AA_MAGIC/DataRepository_Milestone/test_musiasem_2.xlsx"
    with open(fn, "rb") as f:
        b = f.read()

    b = process_file(b)

    # b.save("/home/rnebot/GoogleDrive/AA_MAGIC/DataRepository_Milestone/test_based_on_template_output.xlsx")

    with open("/home/rnebot/GoogleDrive/AA_MAGIC/DataRepository_Milestone/test_based_on_template_output.xlsx", "wb") as f:
        f.write(b)

# OpenPyXL
